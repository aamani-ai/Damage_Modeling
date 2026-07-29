#!/usr/bin/env python3
"""Validate the coverage-complete TC-wind x solar model-v2.1 proposal."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import sys
import zipfile
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PROPOSED = ROOT / "docs/cells/tropical_cyclone_wind_solar/proposed"
ARTIFACT = PROPOSED / "tropical_cyclone_wind_solar__model_v2_1__docs_r1__curve_artifact.json"
CAPABILITY = PROPOSED / "tropical_cyclone_wind_solar__model_v2_1__docs_r1__capability.json"
KATS = PROPOSED / "known_answer_tests_tropical_cyclone_wind_solar__model_v2_1__docs_r1.json"
SOURCES = PROPOSED / "SOURCE_REGISTER_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv"
CLAIMS = PROPOSED / "CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv"
PARAMETERS = PROPOSED / "PARAMETER_TIER_TABLE_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv"
VALUES = PROPOSED / "VALUE_CROSSWALK_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv"
OLD_NEW = PROPOSED / "OLD_VS_NEW_COMPARISON_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv"
CURVE_TABLE = PROPOSED / "FULL_PLANT_SCREENING_CURVE_TABLE_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv"
WORKBOOK = PROPOSED / "damage_curve_records_tropical_cyclone_wind_solar__model_v2_1__docs_r1.xlsx"
V20_ARTIFACT = PROPOSED / "tropical_cyclone_wind_solar__model_v2_0__docs_r1__curve_artifact.json"
V20_SHA = "06ee048096f3a54344e18e00cb8831a7a33910e61034f23fd1f4c33415658428"
INDEX = ROOT / "docs/contracts/machine_readable_artifact_index.json"
SCHEMA_DIR = ROOT / "docs/contracts/schemas"
OVERVIEW = PROPOSED / "README_tropical_cyclone_wind_solar__model_v2_1__docs_r1.md"
DOSSIER = PROPOSED / "tropical_cyclone_wind_solar_curve_derivation_dossier__model_v2_1__docs_r1.md"
METADATA = PROPOSED / "tropical_cyclone_wind_solar_damage_code_metadata_spec__model_v2_1__docs_r1.md"
REPORT = PROPOSED / "VALIDATION_REPORT_tropical_cyclone_wind_solar__model_v2_1__docs_r1.md"
GUIDE = ROOT / "docs/extra/guides/tropical_cyclone_wind_solar_v2_1_curve_request_guide.md"
HANDOFF = ROOT / "docs/contracts/hazard_handoff/tropical_cyclone_wind_solar_model_v2_1_screening_proposal.md"

sys.path.insert(0, str(ROOT / "scripts/reference_helpers"))
import tropical_cyclone_wind_solar_v2_curve_eval as v20  # noqa: E402
import tropical_cyclone_wind_solar_v2_1_curve_eval as evaluator  # noqa: E402
from build_tropical_cyclone_wind_solar_v2_1_package import (  # noqa: E402
    direct_gsu_request,
    full_fixed_request,
    full_tracker_request,
)


class ValidationFailure(AssertionError):
    pass


class Checks:
    count = 0


def require(condition: bool, message: str) -> None:
    Checks.count += 1
    if not condition:
        raise ValidationFailure(message)


def close(actual: float, expected: float, tolerance: float, message: str) -> None:
    require(
        math.isclose(actual, expected, rel_tol=0, abs_tol=tolerance),
        f"{message}: expected {expected}, got {actual}",
    )


def load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
    require(bool(rows), f"empty CSV: {path.name}")
    for line_number, row in enumerate(rows, 2):
        require(None not in row, f"{path.name}:{line_number}: extra columns")
        require(all(value is not None for value in row.values()), f"{path.name}:{line_number}: missing columns")
    return rows


def validate_identity(artifact: Mapping[str, Any], capability: Mapping[str, Any]) -> None:
    require(artifact["schema_version"] == "damage_curve_record_bundle.v3", "bundle schema")
    require(artifact["cell_id"] == "tropical_cyclone_wind_solar", "cell identity")
    require(artifact["semantic_damage_model_version"] == "model v2.1", "model identity")
    require(artifact["documentation_revision"] == "docs r1", "docs identity")
    require(artifact["canonical_runtime_artifact"] is False, "proposal became canonical")
    require(artifact["package_inclusion_status"] == "not_included", "proposal became packaged")
    require(artifact["capability_declaration"] == capability, "embedded capability mismatch")
    require(
        capability["pathway_capabilities"][0]["scenario_loss_given_value_basis"]
        == "supported_with_explicit_failure_unit_value_and_exposure_basis",
        "scenario loss is not supported",
    )
    require(capability["pathway_capabilities"][0]["withheld_failure_units"] == [], "withheld units remain")
    require(
        artifact["emit_contract"]["companion_output_schema"]
        == "physical_damage_assembly.v1",
        "assembly schema missing",
    )
    supported_outputs = set(artifact["emit_contract"]["supported_outputs"])
    require("full-plant physical replacement DR" in supported_outputs, "plant DR unsupported")
    require("scenario physical dollars when capacity_kwdc is supplied" in supported_outputs, "scenario dollars unsupported")
    prohibited = " ".join(artifact["emit_contract"]["prohibited_outputs"]).lower()
    require("full-plant" not in prohibited, "full plant still prohibited")
    require("scenario dollar" not in prohibited, "scenario dollars still prohibited")
    require("eal" in prohibited and "pml" in prohibited, "consumer-owned metric boundary missing")
    require(
        "SCENARIO_LOSS_WITHHELD_SYNTHETIC_T4_PROPOSAL"
        not in artifact["evaluation_contract"]["failure_codes"],
        "obsolete scenario-loss failure code remains",
    )
    require(
        "architecture_prohibited_for_direct_withheld_unit_query"
        not in artifact["evaluation_contract"],
        "obsolete withheld-unit query label remains",
    )
    require(
        artifact["evaluation_contract"]["withheld_reason_codes"]
        == ["DERIVED_ASSEMBLY_RULE_ONLY"],
        "obsolete v2.0 unit withholding reasons remain",
    )
    require(
        "EVENT_PHYSICAL_DAMAGE_OUTPUTS_AVAILABLE_BEFORE_PROMOTION"
        in capability["consumer_annual_metrics"]["limitation_flags"],
        "annual-metric block does not distinguish available event damage outputs",
    )


def validate_schemas(artifact: Mapping[str, Any], capability: Mapping[str, Any]) -> None:
    from jsonschema import Draft202012Validator
    from referencing import Registry, Resource

    schema_paths = [
        SCHEMA_DIR / "curve_artifact_bundle.v3.schema.json",
        SCHEMA_DIR / "capability_declaration.v3.schema.json",
        SCHEMA_DIR / "damage_emit.v2.schema.json",
        SCHEMA_DIR / "physical_damage_assembly.v1.schema.json",
    ]
    schemas = [load(path) for path in schema_paths]
    registry = Registry().with_resources(
        [(schema["$id"], Resource.from_contents(schema)) for schema in schemas]
    )
    for schema in schemas:
        Draft202012Validator.check_schema(schema)
    Draft202012Validator(schemas[1], registry=registry).validate(capability)
    Draft202012Validator(schemas[0], registry=registry).validate(artifact)
    examples = [
        evaluator.evaluate_damage_call(artifact, full_fixed_request(1.0, 1.0)),
        evaluator.evaluate_damage_call(artifact, direct_gsu_request(1.0)),
    ]
    Draft202012Validator(schemas[2], registry=registry).validate(examples[0]["damage_emit"])
    Draft202012Validator(schemas[3], registry=registry).validate(examples[0]["physical_damage_assembly"])
    Draft202012Validator(schemas[2], registry=registry).validate(examples[1])


def validate_records(artifact: Mapping[str, Any]) -> None:
    records = artifact["pathways"][0]["curve_records"]
    require(len(records) == 10, "expected ten curve records")
    require(len({record["curve_id"] for record in records}) == 10, "duplicate curve id")
    require(len({record["failure_unit_id"] for record in records}) == 10, "duplicate failure-unit curve")
    common = [record for record in records if record["failure_unit_id"] in evaluator.COMMON_CURVE_UNITS]
    require(len(common) == 5, "common curve count")
    for record in common:
        require(record["x_axis"] == "site_facility_tropical_cyclone_wind_demand_ratio", "common axis")
        states = record["parameters"]["damage_states"]
        scenarios = record["parameters"]["capacity_scenarios"]
        require(len(states) == 4, "common state count")
        require([state["cost_ratio"] for state in states] == sorted(state["cost_ratio"] for state in states), "cost ratios unordered")
        previous = {scenario["scenario_id"]: -1.0 for scenario in scenarios}
        for step in range(401):
            axis = 2.0 * step / 400
            evaluated = v20.evaluate_ordered_damage_state_record(record, axis)
            for scenario_id, values in evaluated.items():
                dr = values["damage_ratio"]
                require(0 <= dr <= 1, "common DR bounds")
                require(dr + 1e-14 >= previous[scenario_id], "common DR nonmonotone")
                previous[scenario_id] = dr
                close(sum(values["state_probabilities"].values()), 1.0, 1e-12, "state closure")
            require(
                evaluated["lower_resistance"]["damage_ratio"] + 1e-14
                >= evaluated["central_screening"]["damage_ratio"]
                >= evaluated["upper_resistance"]["damage_ratio"] - 1e-14,
                "resistance scenarios unordered",
            )


def validate_assembly(artifact: Mapping[str, Any]) -> None:
    expected_units_by_architecture = {
        evaluator.FIXED: {
            "PV_FIXED_TILT_MODULE_FIELD",
            "PV_FIXED_TILT_SUPPORT_STRUCTURE",
            *evaluator.COMMON_CURVE_UNITS,
        },
        evaluator.TRACKER: {
            "PV_TRACKER_MODULE_FIELD",
            "PV_TRACKER_SBOS_ASSEMBLY",
            *evaluator.COMMON_CURVE_UNITS,
        },
    }
    for architecture, request_factory in (
        (evaluator.FIXED, full_fixed_request),
        (evaluator.TRACKER, full_tracker_request),
    ):
        previous = {scenario: -1.0 for scenario in ("lower_resistance", "central_screening", "upper_resistance")}
        for step in range(101):
            ratio = 2.0 * step / 100
            output = evaluator.evaluate_damage_call(artifact, request_factory(ratio, ratio))
            emit = output["damage_emit"]
            assembly = output["physical_damage_assembly"]
            require(emit["selectors_used"]["array_architecture"] == architecture, "architecture output")
            result_units = {item["failure_unit_id"] for item in emit["failure_unit_results"]}
            require(result_units == expected_units_by_architecture[architecture], "coverage result set")
            require(all(item["status"] == "conditional" for item in emit["failure_unit_results"]), "numeric unit withheld")
            require(all(item["scalar_central_dr"] is not None for item in emit["failure_unit_results"]), "numeric unit null")
            coverage = assembly["coverage"]
            close(coverage["physical_replacement_value_fraction"], 1.0, 0.0, "physical coverage")
            close(
                coverage["intrinsic_curve_value_fraction"] + coverage["derived_support_value_fraction"],
                1.0,
                1e-12,
                "coverage reconciliation",
            )
            for scenario_id, result in assembly["scenario_results"].items():
                dr = result["physical_replacement_dr"]
                require(0 <= dr <= 1, "plant DR bounds")
                require(dr + 1e-14 >= previous[scenario_id], "plant DR nonmonotone")
                previous[scenario_id] = dr
                close(
                    result["physical_loss_2024_usd_per_kwdc"],
                    result["direct_and_civil_loss_2024_usd_per_kwdc"]
                    + result["replacement_support_loss_2024_usd_per_kwdc"],
                    1e-12,
                    "physical loss assembly",
                )
                close(
                    dr,
                    result["physical_loss_2024_usd_per_kwdc"]
                    / evaluator.PHYSICAL_REPLACEMENT_VALUE_PER_KWDC,
                    1e-14,
                    "physical DR denominator",
                )
                close(
                    result["installed_capex_physical_loss_fraction"],
                    result["physical_loss_2024_usd_per_kwdc"]
                    / evaluator.INSTALLED_CAPEX_PER_KWDC,
                    1e-14,
                    "installed denominator",
                )
            scenarios = assembly["scenario_results"]
            require(
                scenarios["lower_resistance"]["physical_replacement_dr"] + 1e-14
                >= scenarios["central_screening"]["physical_replacement_dr"]
                >= scenarios["upper_resistance"]["physical_replacement_dr"] - 1e-14,
                "plant resistance scenarios unordered",
            )
        zero = evaluator.evaluate_damage_call(artifact, request_factory(0.0, 0.0))
        require(
            all(
                result["physical_replacement_dr"] == 0
                for result in zero["physical_damage_assembly"]["scenario_results"].values()
            ),
            "zero input not zero",
        )
        high = evaluator.evaluate_damage_call(artifact, request_factory(2.0, 2.0))
        require(
            high["physical_damage_assembly"]["scenario_results"]["central_screening"]["physical_replacement_dr"] > 0.75,
            "artificial legacy cap remains",
        )
    dollar_output = evaluator.evaluate_damage_call(
        artifact, full_fixed_request(1.0, 1.0, capacity=100000.0)
    )
    central = dollar_output["physical_damage_assembly"]["scenario_results"]["central_screening"]
    close(
        central["scenario_physical_loss_2024_usd"],
        central["physical_loss_2024_usd_per_kwdc"] * 100000.0,
        1e-8,
        "scenario dollars",
    )
    gsu = evaluator.evaluate_damage_call(artifact, direct_gsu_request(1.0))
    require(gsu["failure_unit_results"][0]["status"] == "conditional", "GSU remains withheld")
    require(gsu["failure_unit_results"][0]["scalar_central_dr"] > 0, "GSU remains null/zero")


def validate_kats(artifact: Mapping[str, Any]) -> tuple[int, int]:
    fixture = load(KATS)
    require(fixture["semantic_damage_model_version"] == "model v2.1", "KAT model")
    for test in fixture["runtime_known_answer_tests"]:
        actual = evaluator.evaluate_damage_call(artifact, test["request"])
        require(actual == test["expected"], f"{test['test_id']}: output drift")
    for test in fixture["rejection_tests"]:
        try:
            evaluator.evaluate_damage_call(artifact, test["request"])
        except (evaluator.TropicalCycloneWindSolarV21EvaluationError, v20.TropicalCycloneWindSolarV2EvaluationError) as error:
            require(error.code == test["expected_error_code"], f"{test['test_id']}: wrong error {error.code}")
        else:
            raise ValidationFailure(f"{test['test_id']}: rejection missing")
    return len(fixture["runtime_known_answer_tests"]), len(fixture["rejection_tests"])


def validate_pin(artifact: Mapping[str, Any]) -> None:
    artifact_hash = sha(ARTIFACT)
    pin = {
        "cell_id": artifact["cell_id"],
        "semantic_damage_model_version": artifact["semantic_damage_model_version"],
        "documentation_revision": artifact["documentation_revision"],
        "schema_version": artifact["schema_version"],
        "artifact_sha256": artifact_hash,
    }
    v20.verify_artifact_pin(artifact, pin, artifact_sha256_hex=artifact_hash)
    bad = dict(pin)
    bad["artifact_sha256"] = "0" * 64
    try:
        v20.verify_artifact_pin(artifact, bad, artifact_sha256_hex=artifact_hash)
    except v20.TropicalCycloneWindSolarV2EvaluationError as error:
        require(error.code == "ARTIFACT_PIN_MISMATCH", "pin mismatch code")
    else:
        raise ValidationFailure("bad pin accepted")


def validate_workbook() -> tuple[int, int]:
    require(zipfile.is_zipfile(WORKBOOK), "workbook is not XLSX")
    workbook = load_workbook(WORKBOOK, data_only=False, read_only=True)
    required = {
        "README",
        "Curve_Records",
        "Value_Crosswalk",
        "Parameter_Tiers",
        "Sources",
        "Claims",
        "Old_vs_New",
        "Plant_Curve_Table",
        "KATs",
        "QA",
    }
    require(set(workbook.sheetnames) == required, "workbook sheet set")
    qa = workbook["QA"]
    statuses = [qa.cell(row=row, column=4).value for row in range(2, qa.max_row + 1)]
    require(statuses and all(status == "PASS" for status in statuses), "workbook QA")
    return len(workbook.sheetnames), len(statuses)


def validate_governance(artifact: Mapping[str, Any]) -> tuple[int, int, int, int, int]:
    for path in (
        ARTIFACT,
        CAPABILITY,
        KATS,
        SOURCES,
        CLAIMS,
        PARAMETERS,
        VALUES,
        OLD_NEW,
        CURVE_TABLE,
        WORKBOOK,
        OVERVIEW,
        DOSSIER,
        METADATA,
        REPORT,
        GUIDE,
        HANDOFF,
        SCHEMA_DIR / "physical_damage_assembly.v1.schema.json",
    ):
        require(path.exists(), f"missing {path.name}")
    for field in (
        "source_dossier",
        "source_workbook",
        "known_answer_tests",
        "source_register",
        "claim_parameter_register",
        "value_crosswalk",
        "screening_curve_table",
    ):
        require((ROOT / artifact[field]).exists(), f"broken artifact path {field}")
    sources = csv_rows(SOURCES)
    claims = csv_rows(CLAIMS)
    parameters = csv_rows(PARAMETERS)
    values = csv_rows(VALUES)
    old_new = csv_rows(OLD_NEW)
    curve_table = csv_rows(CURVE_TABLE)
    require({"TCWS21-C001", "TCWS21-C002"}.issubset({row["claim_id"] for row in claims}), "v2.1 claims missing")
    require(len(old_new) == 10, "old/new row count")
    require(len(curve_table) == 246, "full curve-table row count")
    require({row["architecture"] for row in curve_table} == {"fixed_tilt", "tracker"}, "curve-table architectures")
    require({row["scenario_id"] for row in curve_table} == {"lower_resistance", "central_screening", "upper_resistance"}, "curve-table scenarios")
    require(all(row["model_v2_0_full_plant_dr"] == "WITHHELD" for row in old_new), "v2.0 baseline misstated")
    require(sha(V20_ARTIFACT) == V20_SHA, "v2.0 baseline changed")
    require("model v2.1" not in INDEX.read_text(), "artifact index cutover occurred")
    require(not (PROPOSED.parent / "current").exists(), "current pointer created")
    artifact_hash = sha(ARTIFACT)
    report = REPORT.read_text()
    guide = GUIDE.read_text()
    handoff = HANDOFF.read_text()
    for expected_hash in (sha(ARTIFACT), sha(CAPABILITY), sha(KATS), sha(WORKBOOK)):
        require(expected_hash in report, f"validation report missing hash {expected_hash}")
    require(guide.count(artifact_hash) >= 3, "usage guide artifact pins are stale")
    require(artifact_hash in handoff, "Hazard handoff artifact pin is stale")
    require("full-plant physical replacement DR" in artifact["emit_contract"]["supported_outputs"], "plant output contract drift")
    return len(sources), len(claims), len(parameters), len(values), len(old_new)


def main() -> int:
    artifact = load(ARTIFACT)
    capability = load(CAPABILITY)
    validate_identity(artifact, capability)
    validate_schemas(artifact, capability)
    validate_records(artifact)
    validate_assembly(artifact)
    runtime_kats, rejection_kats = validate_kats(artifact)
    validate_pin(artifact)
    workbook_sheets, workbook_qa = validate_workbook()
    source_count, claim_count, parameter_count, value_count, old_new_count = validate_governance(artifact)
    print("PASS tropical_cyclone_wind_solar model v2.1/docs r1 coverage-complete screening proposal")
    print(f"checks={Checks.count}")
    print("schema_validation=bundle v3 + capability v3 + damage emit v2 + physical assembly v1")
    print(f"curve_records={len(artifact['pathways'][0]['curve_records'])}")
    print(f"runtime_kats={runtime_kats}")
    print(f"rejection_kats={rejection_kats}")
    print(f"sources={source_count}")
    print(f"claims={claim_count}")
    print(f"parameters={parameter_count}")
    print(f"value_rows={value_count}")
    print(f"old_vs_new_rows={old_new_count}")
    print("full_plant_curve_table_rows=246")
    print(f"workbook_sheets={workbook_sheets}")
    print(f"workbook_qa_passes={workbook_qa}")
    print(f"artifact_sha256={sha(ARTIFACT)}")
    print(f"capability_sha256={sha(CAPABILITY)}")
    print(f"known_answer_tests_sha256={sha(KATS)}")
    print(f"workbook_sha256={sha(WORKBOOK)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
