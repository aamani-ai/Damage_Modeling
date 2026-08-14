#!/usr/bin/env python3
"""Build the owner-approved TC-wind × canonical Wind Farm v1.1 proposal.

The builder preserves model-v1.0 source records and adds one exact, named
3.3-MW-source → 5-MW-target screening bridge.  It is deliberately not a
generic nearest-neighbour mechanism.
"""

from __future__ import annotations

import csv
import json
from copy import deepcopy
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import tropical_cyclone_wind_wind_curve_eval as evaluator


ROOT = Path(__file__).resolve().parents[2]
CELL = ROOT / "docs/cells/tropical_cyclone_wind_wind"
CURRENT = CELL / "current"
PROPOSED = CELL / "proposed"
STEM = "model_v1_1__docs_r1"
CURRENT_ARTIFACT = CURRENT / "tropical_cyclone_wind_wind__model_v1_0__docs_r1__curve_artifact.json"
CURRENT_CAPABILITY = CURRENT / "tropical_cyclone_wind_wind__model_v1_0__docs_r1__capability.json"
CURRENT_KATS = CURRENT / "known_answer_tests_tropical_cyclone_wind_wind__model_v1_0__docs_r1.json"
ARTIFACT = PROPOSED / f"tropical_cyclone_wind_wind__{STEM}__curve_artifact.json"
CAPABILITY = PROPOSED / f"tropical_cyclone_wind_wind__{STEM}__capability.json"
KATS = PROPOSED / f"known_answer_tests_tropical_cyclone_wind_wind__{STEM}.json"
SOURCES = PROPOSED / f"SOURCE_REGISTER_tropical_cyclone_wind_wind__{STEM}.csv"
CLAIMS = PROPOSED / f"CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_wind__{STEM}.csv"
PARAMETERS = PROPOSED / f"PARAMETER_TIER_TABLE_tropical_cyclone_wind_wind__{STEM}.csv"
VALUES = PROPOSED / f"VALUE_CROSSWALK_tropical_cyclone_wind_wind__{STEM}.csv"
OLD_NEW = PROPOSED / f"OLD_VS_NEW_COMPARISON_tropical_cyclone_wind_wind__{STEM}.csv"
WORKBOOK = PROPOSED / f"damage_curve_records_tropical_cyclone_wind_wind__{STEM}.xlsx"


def load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=False) + "\n")


def write_csv(path: Path, header: list[str], rows: Iterable[Iterable[Any]]) -> None:
    with path.open("w", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(header)
        writer.writerows(rows)


def _proxy_record(source: Mapping[str, Any]) -> dict[str, Any]:
    record = deepcopy(source)
    record.update(
        {
            "curve_id": "TCWW_JAIMES_3P3MW_AS_CANONICAL_5MW_OWNER_PROXY_V1",
            "failure_unit_id": evaluator.PROXY_FAILURE_UNIT,
            "selector_match": {
                "turbine_archetype_id": evaluator.PROXY_ARCHETYPE_ID,
                "rated_power_mw": 5,
                "hub_height_m": 100,
                "rotor_diameter_m": 114,
            },
        }
    )
    return record


def build_capability(current: Mapping[str, Any]) -> dict[str, Any]:
    capability = deepcopy(current)
    capability["canonical_runtime_artifact"] = False
    pathway = capability["pathway_capabilities"][0]
    pathway.update(
        {
            "scenario_loss_given_value_basis": (
                "supported_with_explicit_failure_unit_value_and_exposure_basis"
            ),
            "conditions": [
                "exact source-native Jaimes selectors retain model-v1.0 behavior",
                "canonical 5 MW use requires the exact owner-approved proxy policy, asset profile, and covered-value basis IDs",
                "the proxy evaluates the unchanged Jaimes 3.3 MW / 100 m parameters on tc_peak_gust_3s_10m_kmh",
                "the proxy applies only to WT_TURBINE_EQUIPMENT_ASSEMBLY rotor+nacelle+tower screening scope",
                "covered value is 0.63 of project TIV; the remaining 0.37 is withheld, not zero",
                "source-native selectors retain v1.0 withheld-range behavior",
                "the named 5 MW proxy alone assigns zero in 90–108 km/h and caps DR at max_dr above 252 km/h, with explicit flags",
            ],
            "limitation_flags": [
                "OWNER_APPROVED_SCREENING_PROXY",
                "REQUESTED_5MW_EVALUATED_WITH_3P3MW_SOURCE_CURVE",
                "NO_CAPACITY_RATIO_SCALING",
                "NOT_TARGET_MATCHED_5MW_EVIDENCE",
                "NOT_FIELD_CALIBRATED",
                "NOT_CLAIMS_CALIBRATED",
                "PARTIAL_STRUCTURAL_VALUE_COVERAGE_63PCT",
                "UNCOVERED_PROJECT_VALUE_37PCT_WITHHELD_NOT_ZERO",
                "CURVE_INTRINSIC_SPREAD_NOT_CARRIED",
                "NO_NHC_OR_HUB_HEIGHT_BRIDGE",
                "PROXY_SCREENING_TRANSITION_BAND_ZERO_RULE",
                "PROXY_SCREENING_ABOVE_RANGE_MAX_DR_CAP",
            ],
            "withheld_failure_units": [
                item
                for item in pathway["withheld_failure_units"]
                if item["failure_unit_id"] != evaluator.PROXY_FAILURE_UNIT
            ],
        }
    )
    capability["consumer_annual_metrics"] = {
        "computation_owner": "downstream_consumer",
        "status_before_promotion": "withheld_noncanonical_proposal",
        "status_after_promotion": "consumer_computable_from_validated_frequency_intensity_coupling_value_and_cap_model",
        "prerequisites": [
            "released exact cell/model/docs/schema/SHA pin",
            "validated Hurricane event frequency and intensity coupling",
            "exact canonical asset profile and proxy policy identities",
            "0.63 covered-value cap applied at occurrence and annual grains",
            "compound-event partition and full-grid consumer validation",
        ],
        "limitation_flags": [
            "CURVE_INTRINSIC_SPREAD_NOT_CARRIED",
            "PARTIAL_STRUCTURAL_VALUE_COVERAGE_63PCT",
            "ANNUAL_AND_TAIL_METRICS_CONDITIONAL_ON_SCREENING_PROXY",
        ],
    }
    capability["cap_binding"] = {
        "policy": "consumer_enforced_fail_closed",
        "enforcement_owner": "downstream_consumer",
        "checks_required": [
            "exact proxy policy, asset profile, value basis, model/docs/schema/SHA pin",
            "covered loss <= 0.63 * project TIV for every occurrence",
            "uncovered 0.37 remains withheld and is never emitted as zero loss",
            "no 5/3.3 multiplier or generic nearest-neighbour selector",
            "proxy boundary completion is exactly zero for 90–108 km/h and max_dr above 252 km/h, with flags",
            "annual/TIV cap is applied inside the annual calculation",
        ],
        "action_if_fail": "withhold scenario loss and all annual or tail metrics",
    }
    capability["promotion_gate"] = {
        "status": "ready_for_review",
        "required_before_canonical_use": [
            "old-v-new source-native reproduction",
            "proxy, negative-contract, value-share, and cap known answers",
            "Hazard exact-pin M2-M4 integration and rollback test",
            "owner review of explicit screening grade and partial-value reporting",
        ],
    }
    return capability


def build_artifact() -> tuple[dict[str, Any], dict[str, Any]]:
    artifact = deepcopy(load(CURRENT_ARTIFACT))
    capability = build_capability(load(CURRENT_CAPABILITY))
    artifact.update(
        {
            "damage_code_id": "TROPICAL_CYCLONE_WIND_WIND_JAIMES_SCREENING_V1_1",
            "semantic_damage_model_version": "model v1.1",
            "documentation_revision": "docs r1",
            "lifecycle_state": "proposed_v1_1",
            "promotion_status": "proposed",
            "review_status": "owner_directed_proxy_pending_consumer_validation",
            "model_grade": "screening_owner_approved_target_mismatch_proxy",
            "canonical_runtime_artifact": False,
            "source_dossier": (
                "docs/cells/tropical_cyclone_wind_wind/proposed/"
                "tropical_cyclone_wind_wind_curve_derivation_dossier__model_v1_1__docs_r1.md"
            ),
            "source_workbook": WORKBOOK.relative_to(ROOT).as_posix(),
            "known_answer_tests": KATS.relative_to(ROOT).as_posix(),
            "source_register": SOURCES.relative_to(ROOT).as_posix(),
            "claim_parameter_register": CLAIMS.relative_to(ROOT).as_posix(),
            "value_crosswalk": VALUES.relative_to(ROOT).as_posix(),
        }
    )
    pathway = artifact["pathways"][0]
    source = next(
        record
        for record in pathway["curve_records"]
        if record["curve_id"] == "TCWW_JAIMES_3P3MW_100M_SCREENING"
    )
    pathway["curve_records"].append(_proxy_record(source))
    selector = next(
        item for item in pathway["selector_logic"] if item["field"] == "turbine_archetype_id"
    )
    selector["allowed"].append(evaluator.PROXY_ARCHETYPE_ID)
    selector.update(
        {
            "routing": "exact selector match; canonical 5 MW route additionally requires exact proxy policy, asset profile, and value-basis IDs",
            "nearest_neighbor": "prohibited except the single named owner-approved bridge, which is not inferred",
            "modern_fleet_transfer": "one explicit canonical 5 MW screening bridge only",
        }
    )
    pathway["exposure_contract"].update(
        {
            "whole_farm_default": "prohibited; partial covered-value use requires exact proxy contract",
            "value_rule": (
                "Proxy DR may be applied only to rotor+nacelle+tower = 0.63 of project TIV; "
                "foundation/substation/electrical/civil = 0.37 withheld, not zero."
            ),
        }
    )
    for unit in artifact["failure_units"]:
        if unit["id"] == evaluator.PROXY_FAILURE_UNIT:
            unit.update(
                {
                    "treatment": "secondary_conditional",
                    "denominator": (
                        "canonical Wind Farm rotor+nacelle+tower physical replacement scope; "
                        "0.63 of project installed TIV"
                    ),
                    "covered_value_basis_id": evaluator.PROXY_VALUE_BASIS_ID,
                    "covered_value_share_of_project_tiv": evaluator.PROXY_COVERED_VALUE_SHARE,
                    "withheld_reason_codes": [],
                }
            )
    for coverage in pathway["failure_unit_coverage"]:
        if coverage["failure_unit_id"] == evaluator.PROXY_FAILURE_UNIT:
            coverage.update(
                {
                    "status": "conditional_owner_approved_screening_proxy",
                    "reason_codes": [],
                }
            )
    artifact["derivation_rationale"].update(
        {
            "v1_1_decision": "admit one owner-approved canonical-5-MW screening bridge without changing source parameters",
            "why_minor_version": "supported selector/failure-unit/value behavior expands while existing curves and axis remain unchanged",
            "why_not_capacity_scaling": "rated power is not a damage-ratio multiplier and no evidence supports a 5/3.3 adjustment",
            "owner_approved_proxy_contract": {
                "status": "owner_approved_screening_proxy",
                "proxy_policy_id": evaluator.PROXY_POLICY_ID,
                "canonical_asset_profile_id": evaluator.PROXY_ASSET_PROFILE_ID,
                "source_curve_id": source["curve_id"],
                "source_evidence_identity": "Jaimes 3.3 MW / 100 m / 114 m record",
                "target_identity": "canonical 5 MW / 100 m turbine",
                "numeric_rule": "evaluate unchanged source parameters; no 5/3.3 scaling",
                "screening_completion_rule": {
                    "transition_band_kmh": [90, 108],
                    "transition_treatment": "zero_with_explicit_flag",
                    "transition_eal_upper_bound_all_active_conus_placements_usd": 10564.849317568538,
                    "source_ceiling_kmh": 252,
                    "above_ceiling_treatment": "cap_at_max_dr_with_explicit_flag",
                    "measurement_scope": "1773 active cells; 113526 governed M1 events; 20 canonical turbine nodes",
                },
                "covered_value_basis_id": evaluator.PROXY_VALUE_BASIS_ID,
                "covered_value_share_of_project_tiv": evaluator.PROXY_COVERED_VALUE_SHARE,
                "uncovered_value_share_of_project_tiv": 1.0
                - evaluator.PROXY_COVERED_VALUE_SHARE,
            },
        }
    )
    artifact["evaluation_contract"].update(
        {
            "selector_behavior": "exact source selectors plus one exact owner-approved 5 MW bridge; no inferred nearest neighbour",
            "proxy_screening_completion": {
                "applies_only_when_proxy_policy_id": evaluator.PROXY_POLICY_ID,
                "transition_band_kmh": [90, 108],
                "transition_treatment": "return DR=0 with SCREENING_TRANSITION_BAND_ASSIGNED_ZERO",
                "source_ceiling_kmh": 252,
                "above_ceiling_treatment": "return max_dr with SCREENING_ABOVE_SOURCE_RANGE_CAPPED_AT_MAX_DR",
                "source_native_selectors_unchanged": True,
            },
        }
    )
    artifact["value_linkage"].update(
        {
            "scenario_loss_status": "consumer_computable_only_for_explicit_proxy_covered_value",
            "owner_approved_proxy_value_basis": {
                "id": evaluator.PROXY_VALUE_BASIS_ID,
                "canonical_asset_profile_id": evaluator.PROXY_ASSET_PROFILE_ID,
                "project_tiv_usd_at_activation": 140000000,
                "covered_subsystem_shares": {
                    "rotor": 0.26,
                    "nacelle": 0.21,
                    "tower": 0.16,
                },
                "covered_value_share": evaluator.PROXY_COVERED_VALUE_SHARE,
                "covered_value_usd_at_activation": 88200000,
                "uncovered_value_share": 0.37,
                "uncovered_value_usd_at_activation": 51800000,
                "uncovered_treatment": "withheld_not_zero",
            },
            "approval_required": "exact Hazard consumer pin, cap validation, and owner review before promotion",
        }
    )
    artifact["emit_contract"].update(
        {
            "supported_output": "conditional scalar mean DR for exact source-native records or the named canonical-5-MW turbine-equipment proxy",
            "scenario_loss": "consumer-computable only with the exact 0.63 covered-value basis and caps",
            "annual_and_tail_metrics": "downstream consumer only after frequency/coupling/cap validation",
        }
    )
    artifact["capability_declaration"] = capability
    return artifact, capability


def build_kats(artifact: Mapping[str, Any]) -> dict[str, Any]:
    current = load(CURRENT_KATS)
    proxy_base = {
        "pathway_id": evaluator.SUPPORTED_PATHWAY,
        "failure_unit_id": evaluator.PROXY_FAILURE_UNIT,
        "turbine_archetype_id": evaluator.PROXY_ARCHETYPE_ID,
        "source_model_assumption_set_id": evaluator.SUPPORTED_ASSUMPTION_SET,
        "proxy_policy_id": evaluator.PROXY_POLICY_ID,
        "canonical_asset_profile_id": evaluator.PROXY_ASSET_PROFILE_ID,
        "covered_value_basis_id": evaluator.PROXY_VALUE_BASIS_ID,
    }
    record = next(
        item
        for item in artifact["pathways"][0]["curve_records"]
        if item["selector_match"]["turbine_archetype_id"] == evaluator.PROXY_ARCHETYPE_ID
    )
    proxy_tests = []
    for speed in (90.0, 100.0, 108.0, 160.0, 163.3, 180.0, 200.0, 252.0, 300.0):
        expected_dr, boundary_flag = evaluator.evaluate_proxy_screening_completion(record, speed)
        proxy_tests.append(
            {
                "test_id": f"TCWW11_CANONICAL_5MW_PROXY_V{str(speed).replace('.', 'P')}",
                "input": {**proxy_base, evaluator.AXIS_FIELD: speed},
                "expected": {
                    "status": "supported",
                    "curve_id": record["curve_id"],
                    "failure_unit_damage_ratio": expected_dr,
                    "boundary_flag": boundary_flag,
                    "covered_value_share_of_project_tiv": 0.63,
                    "source_curve_id": "TCWW_JAIMES_3P3MW_100M_SCREENING",
                    "capacity_ratio_scaling": "prohibited",
                },
            }
        )
    negative = [
        {
            "test_id": "TCWW11_PROXY_OPT_IN_REQUIRED",
            "input": {
                key: value
                for key, value in {**proxy_base, evaluator.AXIS_FIELD: 180}.items()
                if key not in {"proxy_policy_id", "canonical_asset_profile_id", "covered_value_basis_id"}
            },
            "expected_error_code": "OWNER_APPROVED_PROXY_OPT_IN_REQUIRED",
        },
        {
            "test_id": "TCWW11_PROXY_POLICY_MISMATCH",
            "input": {**proxy_base, "proxy_policy_id": "GENERIC_NEAREST", evaluator.AXIS_FIELD: 180},
            "expected_error_code": "OWNER_APPROVED_PROXY_IDENTITY_MISMATCH",
        },
        {
            "test_id": "TCWW11_ASSET_PROFILE_MISMATCH",
            "input": {**proxy_base, "canonical_asset_profile_id": "OTHER_WIND_FARM", evaluator.AXIS_FIELD: 180},
            "expected_error_code": "OWNER_APPROVED_PROXY_IDENTITY_MISMATCH",
        },
        {
            "test_id": "TCWW11_GENERIC_4MW_REJECTED",
            "input": {**proxy_base, "turbine_archetype_id": "GENERIC_4MW", evaluator.AXIS_FIELD: 180},
            "expected_error_code": "TURBINE_ARCHETYPE_UNSUPPORTED",
        },
    ]
    reproduction = []
    for test in current["formula_known_answer_tests"]:
        reproduction.append(
            {
                "test_id": "TCWW11_REPRO_" + test["test_id"],
                "input": test["input"],
                "expected": test["expected"],
            }
        )
    return {
        "schema_version": "known_answer_tests.tcww_v1_1.v1",
        "cell_id": artifact["cell_id"],
        "semantic_damage_model_version": "model v1.1",
        "documentation_revision": "docs r1",
        "artifact_schema_version": artifact["schema_version"],
        "absolute_tolerance": 1e-12,
        "v1_0_reproduction_tests": reproduction,
        "proxy_known_answer_tests": proxy_tests,
        "negative_contract_tests": negative,
        "value_crosswalk_tests": [
            {
                "test_id": "TCWW11_CANONICAL_VALUE_CAP",
                "project_tiv_usd": 140000000,
                "covered_value_share": 0.63,
                "expected_covered_value_usd": 88200000,
                "expected_uncovered_value_usd": 51800000,
            },
            {
                "test_id": "TCWW11_UNIT_TIV_CAP",
                "project_tiv_usd": 1,
                "covered_value_share": 0.63,
                "expected_covered_value_usd": 0.63,
                "expected_uncovered_value_usd": 0.37,
            },
        ],
    }


def build_registers() -> None:
    write_csv(
        SOURCES,
        ["source_id", "role", "use", "limits"],
        [
            ["JAIMES_2020_TC_WIND_TURBINE", "numeric source", "unchanged 3.3 MW/100 m curve parameters and native axis", "not target-matched to 5 MW"],
            ["OWNER_DECISION_2026_08_14", "governance", "authorizes one explicit canonical-5-MW screening bridge", "not empirical evidence"],
            ["CONUS_WIND_FARM_REFERENCE_V1", "consumer asset profile", "20×5 MW, 100 MW, $140M and subsystem shares", "screening value grade"],
            ["HAZARD_HURRICANE_WIND_FARM_M2_FULL_POPULATION_2026_08_14", "consumer measurement", "1773 active cells, 113526 events, 20-node boundary and M2 sensitivity", "screening completion evidence; not target-matched damage evidence"],
        ],
    )
    write_csv(
        CLAIMS,
        ["claim_id", "claim", "status", "basis", "revisit_trigger"],
        [
            ["TCWW11-C01", "The 5 MW proxy uses the unchanged Jaimes 3.3 MW numerical curve.", "owner-approved screening assumption", "OWNER_DECISION_2026_08_14", "target-matched 5 MW evidence"],
            ["TCWW11-C02", "No 5/3.3 damage scaling is applied.", "hard contract", "damage ratio has no supported capacity multiplier", "new reviewed physical transfer model"],
            ["TCWW11-C03", "Rotor+nacelle+tower cover 0.63 of canonical project TIV.", "owner-approved value crosswalk", "shared canonical asset profile", "improved component valuation"],
            ["TCWW11-C04", "The other 0.37 is withheld, not zero.", "hard reporting contract", "partial coverage discipline", "additional governed curves"],
            ["TCWW11-C05", "The named proxy assigns zero in the unsupported 90–108 km/h transition band.", "owner-approved screening completion", "full-M1 sensitivity bound = $10,564.85 summed placement EAL", "target-matched low-wind evidence"],
            ["TCWW11-C06", "The named proxy caps speeds above 252 km/h at max_dr=1.", "owner-approved screening completion", "bounded physical damage ratio; no curve extrapolation", "target-matched extreme-wind evidence"],
        ],
    )
    write_csv(
        PARAMETERS,
        ["parameter", "value", "tier", "meaning"],
        [
            ["V_zero_kmh", 90, "source-derived", "unchanged Jaimes 3.3 MW record"],
            ["delta_V50_kmh", 73.3, "source-derived", "unchanged Jaimes 3.3 MW record"],
            ["rho", 4.99, "source-derived", "unchanged Jaimes 3.3 MW record"],
            ["target_rated_power_mw", 5, "owner-approved target identity", "no numeric scaling effect"],
            ["covered_value_share", 0.63, "owner-approved value crosswalk", "rotor+nacelle+tower only"],
            ["proxy_transition_zero_upper_kmh", 108, "owner-approved screening completion", "zero below the source-supported simulation band"],
            ["proxy_source_ceiling_kmh", 252, "source-derived boundary", "proxy caps above this speed at max_dr"],
        ],
    )
    write_csv(
        VALUES,
        ["value_basis_id", "subsystem", "share_of_project_tiv", "treatment", "activation_value_usd", "notes"],
        [
            [evaluator.PROXY_VALUE_BASIS_ID, "rotor", 0.26, "covered_screening_proxy", 36400000, "Damage ratio applies"],
            [evaluator.PROXY_VALUE_BASIS_ID, "nacelle", 0.21, "covered_screening_proxy", 29400000, "Damage ratio applies"],
            [evaluator.PROXY_VALUE_BASIS_ID, "tower", 0.16, "covered_screening_proxy", 22400000, "Damage ratio applies"],
            [evaluator.PROXY_VALUE_BASIS_ID, "covered_total", 0.63, "aggregate_cap", 88200000, "maximum wind-only covered loss"],
            [evaluator.PROXY_VALUE_BASIS_ID, "foundation+substation+electrical+civil", 0.37, "withheld_not_zero", 51800000, "outside this model"],
        ],
    )
    write_csv(
        OLD_NEW,
        ["dimension", "model_v1_0", "model_v1_1", "expected_change"],
        [
            ["source-native exact Jaimes selectors", "supported", "identical", "none"],
            ["canonical 5 MW turbine", "rejected", "supported only with exact proxy IDs", "intentional"],
            ["curve parameters", "three source records", "same three plus byte-equivalent proxy record", "no numerical scaling"],
            ["standard turbine equipment assembly", "withheld", "conditional proxy", "intentional"],
            ["covered project value", "withheld", "0.63 of TIV", "intentional partial coverage"],
            ["remaining project value", "withheld", "0.37 withheld", "none; never zero"],
            ["annual metrics", "withheld", "consumer-computable after Hazard gates", "consumer capability only"],
            ["90–108 km/h", "withheld", "proxy returns flagged zero; source selectors still withhold", "intentional screening completion"],
            [">252 km/h", "withheld", "proxy returns flagged max_dr cap; source selectors still withhold", "intentional screening completion"],
        ],
    )


def build_workbook(artifact: Mapping[str, Any], kats: Mapping[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True)

    def sheet(name: str, title: str, header: list[str], rows: list[list[Any]]) -> None:
        ws = workbook.create_sheet(name)
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(header)))
        ws["A1"].fill = title_fill
        ws["A1"].font = title_font
        ws.append(header)
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A3"
        for index, column in enumerate(ws.columns, start=1):
            letter = get_column_letter(index)
            width = min(70, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            ws.column_dimensions[letter].width = width

    sheet(
        "README",
        "TC-wind × Wind Farm model v1.1 — explicit screening proxy",
        ["Field", "Value"],
        [
            ["Current pointer", "model v1.0 remains current until consumer promotion"],
            ["Target", "CONUS_WIND_FARM_REFERENCE_V1 · 20×5 MW · 100 MW · $140M"],
            ["Source curve", "Jaimes 3.3 MW / 100 m, unchanged"],
            ["Covered value", "rotor+nacelle+tower = 0.63 of project TIV"],
            ["Uncovered value", "0.37 withheld, not zero"],
        ],
    )
    sheet(
        "Curve records",
        "Source records plus one exact owner-approved bridge",
        ["Curve ID", "Failure unit", "Archetype", "V0", "delta V50", "rho", "Proxy policy"],
        [
            [
                record["curve_id"],
                record["failure_unit_id"],
                record["selector_match"]["turbine_archetype_id"],
                record["parameters"]["V_zero_kmh"],
                record["parameters"]["delta_V50_kmh"],
                record["parameters"]["rho"],
                record["selector_match"].get("proxy_policy_id", ""),
            ]
            for record in artifact["pathways"][0]["curve_records"]
        ],
    )
    sheet(
        "Proxy KATs",
        "The 5 MW route exactly reproduces the 3.3 MW source equation",
        ["Test", "Speed km/h", "Expected DR", "Covered share", "Scaling"],
        [
            [
                test["test_id"],
                test["input"][evaluator.AXIS_FIELD],
                test["expected"]["failure_unit_damage_ratio"],
                test["expected"]["covered_value_share_of_project_tiv"],
                test["expected"]["capacity_ratio_scaling"],
            ]
            for test in kats["proxy_known_answer_tests"]
        ],
    )
    sheet(
        "Value crosswalk",
        "Canonical Wind Farm project-TIV shares",
        ["Subsystem", "Share", "Treatment", "Activation value USD"],
        [
            ["rotor", 0.26, "covered", 36400000],
            ["nacelle", 0.21, "covered", 29400000],
            ["tower", 0.16, "covered", 22400000],
            ["covered total", 0.63, "cap", 88200000],
            ["remaining plant", 0.37, "withheld_not_zero", 51800000],
        ],
    )
    workbook.save(WORKBOOK)


def main() -> None:
    artifact, capability = build_artifact()
    kats = build_kats(artifact)
    write_json(ARTIFACT, artifact)
    write_json(CAPABILITY, capability)
    write_json(KATS, kats)
    build_registers()
    build_workbook(artifact, kats)
    print(f"built={ARTIFACT.relative_to(ROOT)}")
    print(f"proxy_tests={len(kats['proxy_known_answer_tests'])}")
    print(f"v1_0_reproduction_tests={len(kats['v1_0_reproduction_tests'])}")


if __name__ == "__main__":
    main()
