#!/usr/bin/env python3
"""Validate the proposed TC-wind x solar model-v2 synthetic-T4 package."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Mapping


ROOT = Path(__file__).resolve().parents[2]
PROPOSED = ROOT / "docs/cells/tropical_cyclone_wind_solar/proposed"
ARTIFACT = PROPOSED / "tropical_cyclone_wind_solar__model_v2_0__docs_r1__curve_artifact.json"
CAPABILITY = PROPOSED / "tropical_cyclone_wind_solar__model_v2_0__docs_r1__capability.json"
KATS = PROPOSED / "known_answer_tests_tropical_cyclone_wind_solar__model_v2_0__docs_r1.json"
SOURCES = PROPOSED / "SOURCE_REGISTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
CLAIMS = PROPOSED / "CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
CLAIM_SUPERSESSION = PROPOSED / "CLAIM_SUPERSESSION_MAP_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
PARAMETERS = PROPOSED / "PARAMETER_TIER_TABLE_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
VALUES = PROPOSED / "VALUE_CROSSWALK_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
OLD_NEW = PROPOSED / "OLD_VS_NEW_COMPARISON_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
WORKBOOK = PROPOSED / "damage_curve_records_tropical_cyclone_wind_solar__model_v2_0__docs_r1.xlsx"
SHARED = ROOT / "docs/method/shared_components/solar_wind_normalized_response/candidate_response_profile_v0_1.json"
INDEX = ROOT / "docs/contracts/machine_readable_artifact_index.json"
BUNDLE_SCHEMA = ROOT / "docs/contracts/schemas/curve_artifact_bundle.v3.schema.json"
CAPABILITY_SCHEMA = ROOT / "docs/contracts/schemas/capability_declaration.v3.schema.json"
EMIT_SCHEMA = ROOT / "docs/contracts/schemas/damage_emit.v2.schema.json"
REQUEST_GUIDE = ROOT / "docs/extra/guides/tropical_cyclone_wind_solar_v2_curve_request_guide.md"
HANDOFF = ROOT / "docs/contracts/hazard_handoff/tropical_cyclone_wind_solar_model_v2_0_synthetic_proposal.md"
VALIDATION_REPORT = PROPOSED / "VALIDATION_REPORT_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md"
DERIVATION_DOSSIER = PROPOSED / "tropical_cyclone_wind_solar_curve_derivation_dossier__model_v2_0__docs_r1.md"
EXPECTED_FINAL_CHECKS = 36425

V0_FILES = {
    PROPOSED / "tropical_cyclone_wind_solar__model_v0_1__docs_r1__curve_artifact.json": "2b3753e8bdcef3e3c91c8afb7ca12d67b15cd236873e97c908d6ccccb4748ae1",
    PROPOSED / "tropical_cyclone_wind_solar__model_v0_1__docs_r1__capability.json": "c8bafb3cde61f85f22c7f3b7a10e7ac4bdcb6787f6a7c45d2be7662130e34a60",
    PROPOSED / "known_answer_tests_tropical_cyclone_wind_solar__model_v0_1__docs_r1.json": "ed59cf93fa0403e9a852c820fc5f3f9c7e7217aeb3aa76d02fecf53e5a605e14",
    PROPOSED / "damage_curve_records_tropical_cyclone_wind_solar__model_v0_1__docs_r1.xlsx": "54e126234cf41da494dec77a6a9458b0d1ffa69ecf43cf413803eebb5c20b1bb",
}
V1_FILES = {
    PROPOSED / "tropical_cyclone_wind_solar__model_v1_0__docs_r1__curve_artifact.json": "bb01300d3e76114203dd826be5bff4bb9f2b98490880327dd57575007a180840",
    PROPOSED / "tropical_cyclone_wind_solar__model_v1_0__docs_r1__capability.json": "5cd4f5501961a9d7f2c21259b4cfabd9e74eef30b5fdd9ceff72729b83ffc4fc",
    PROPOSED / "known_answer_tests_tropical_cyclone_wind_solar__model_v1_0__docs_r1.json": "2e18603a9efb5cbb8bdd1c7f3b162e1a3e0c4b0723df5e1afbdc27def84f7cd2",
    PROPOSED / "damage_curve_records_tropical_cyclone_wind_solar__model_v1_0__docs_r1.xlsx": "748031c226187e3b43d83f6a57b2dbd5554457edc01a06debe16b7ef640f3105",
}

sys.path.insert(0, str(ROOT))
from scripts.reference_helpers.tropical_cyclone_wind_solar_v2_curve_eval import (  # noqa: E402
    ARCHITECTURE_UNITS,
    COMMON_REQUEST_FIELDS,
    COMMON_WITHHELD_UNITS,
    FIXED,
    FIXED_REQUEST_FIELDS,
    FAILURE_CODES,
    PATHWAY,
    PERRY,
    PERRY_REQUEST_FIELDS,
    PERRY_SELECTORS,
    PERRY_UNIT,
    TRACKER,
    TRACKER_REQUEST_FIELDS,
    TropicalCycloneWindSolarV2EvaluationError,
    artifact_sha256,
    evaluate_damage_call,
    evaluate_ordered_damage_state_record,
    verify_artifact_pin,
)
from scripts.reference_helpers.build_tropical_cyclone_wind_solar_v2_package import (  # noqa: E402
    CELL_LOCAL_SYNTHETIC_SOURCE,
    NO_DIRECT_TC_CALIBRATION_SOURCE,
    STRONG_V2_ARTIFACT_SHA256,
    fixed_request,
    perry_request,
    tracker_request,
)


class ValidationFailure(AssertionError):
    pass


class Checks:
    value = 0


def require(condition: bool, message: str) -> None:
    Checks.value += 1
    if not condition:
        raise ValidationFailure(message)


def close(actual: float, expected: float, tolerance: float, label: str) -> None:
    require(math.isclose(actual, expected, rel_tol=0, abs_tol=tolerance), f"{label}: expected {expected}, got {actual}")


def load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        data = list(reader)
    require(bool(data), f"empty CSV: {path.name}")
    for line, row in enumerate(data, 2):
        require(None not in row, f"{path.name}:{line}: extra fields")
        require(all(value is not None for value in row.values()), f"{path.name}:{line}: missing fields")
    return data


def split_ids(value: str) -> set[str]:
    return {item.strip() for item in value.split(";") if item.strip()}


def result(emit: Mapping[str, Any], unit: str) -> Mapping[str, Any]:
    matches = [item for item in emit["failure_unit_results"] if item["failure_unit_id"] == unit]
    require(len(matches) == 1, f"{unit}: result not unique")
    return matches[0]


def validate_identity(artifact: Mapping[str, Any], capability: Mapping[str, Any]) -> None:
    require(artifact["schema_version"] == "damage_curve_record_bundle.v3", "wrong bundle schema")
    require(artifact["schema_status"] == "proposed_draft", "wrong schema status")
    require(artifact["cell_id"] == "tropical_cyclone_wind_solar", "wrong cell")
    require(artifact["damage_code_id"] == "TROPICAL_CYCLONE_WIND_SOLAR_SYNTHETIC_T4_V2_PROPOSED", "wrong code")
    require(artifact["semantic_damage_model_version"] == "model v2.0", "wrong model")
    require(artifact["documentation_revision"] == "docs r1", "wrong docs")
    require(artifact["lifecycle_state"] == "candidate", "lifecycle drift")
    require(artifact["promotion_status"] == "proposed_blocked", "promotion status drift")
    require(
        artifact["review_status"] == "independent_proposal_review_complete_promotion_review_pending",
        "review status drift",
    )
    require(artifact["model_grade"] == "experimental_synthetic_T4_scenario", "model grade drift")
    require(artifact["canonical_runtime_artifact"] is False, "proposal cannot be canonical")
    require(artifact["package_release"] == "unreleased", "proposal cannot claim release")
    require(artifact["package_inclusion_status"] == "not_included", "proposal cannot claim inclusion")
    require(artifact["capability_declaration"] == capability, "embedded capability mismatch")
    require(capability["cell_id"] == artifact["cell_id"], "capability cell mismatch")
    require(capability["canonical_runtime_artifact"] is False, "capability cannot be canonical")
    require(capability["promotion_gate"]["status"] == "blocked", "promotion must remain blocked")
    require(capability["pathway_capabilities"][0]["scenario_loss_given_value_basis"] == "withheld", "dollar loss enabled")
    evaluation_contract = artifact["evaluation_contract"]
    require(
        evaluation_contract["architecture_required_for_array_and_perry_routes"] is True,
        "architecture requirement for numerical routes drifted",
    )
    require(
        evaluation_contract["architecture_prohibited_for_direct_withheld_unit_query"] is True,
        "direct-withheld architecture exception drifted",
    )
    require("architecture_required" not in evaluation_contract, "ambiguous architecture requirement returned")
    require(set(evaluation_contract["failure_codes"]) == set(FAILURE_CODES), "failure-code registry drift")
    emit_contract = artifact["emit_contract"]
    require("result_grain" not in emit_contract, "universal result grain cannot describe direct withheld queries")
    require(
        emit_contract["result_grain_by_route"]
        == {
            "numerical_architecture_route": "pathway_id x array_architecture x failure_unit_id x array_zone_or_source_site",
            "direct_withheld_unit_query": "pathway_id x failure_unit_id",
        },
        "route-specific result grain drifted",
    )
    require(
        emit_contract["event_lineage_fields_carried_separately"] == ["event_id", "event_family_id"],
        "event-lineage grain contract drifted",
    )
    shared = artifact["audit_comparison_profile"]
    require(shared["runtime_dependency"] is False and shared["runtime_approved"] is False, "shared profile became runtime")


def validate_paths(artifact: Mapping[str, Any]) -> None:
    for field in (
        "source_dossier",
        "source_workbook",
        "known_answer_tests",
        "source_register",
        "claim_parameter_register",
        "claim_supersession_map",
        "value_crosswalk",
    ):
        path = artifact[field]
        require(isinstance(path, str) and path.startswith("docs/"), f"bad path for {field}")
        require((ROOT / path).exists(), f"missing reference: {path}")
    for path in (ARTIFACT, CAPABILITY, KATS, SOURCES, CLAIMS, CLAIM_SUPERSESSION, PARAMETERS, VALUES, OLD_NEW, WORKBOOK, SHARED):
        require(path.exists(), f"missing generated file {path}")


def validate_schemas(artifact: Mapping[str, Any], capability: Mapping[str, Any]) -> str:
    from jsonschema import Draft202012Validator
    from referencing import Registry, Resource

    schemas = [load(BUNDLE_SCHEMA), load(CAPABILITY_SCHEMA), load(EMIT_SCHEMA)]
    registry = Registry().with_resources(
        [(schema["$id"], Resource.from_contents(schema)) for schema in schemas]
    )
    for schema in schemas:
        Draft202012Validator.check_schema(schema)
    Draft202012Validator(schemas[1], registry=registry).validate(capability)
    Draft202012Validator(schemas[0], registry=registry).validate(artifact)
    samples = [
        evaluate_damage_call(artifact, fixed_request(1.0)),
        evaluate_damage_call(artifact, tracker_request(1.0)),
        evaluate_damage_call(artifact, perry_request(30.0)),
    ]
    gsu = {
        "event_id": "TC-EVENT-VALIDATE",
        "event_family_id": "TC-FAMILY-VALIDATE",
        "pathway_id": PATHWAY,
        "failure_unit_id": "PV_GSU_SUBSTATION",
    }
    samples.append(evaluate_damage_call(artifact, gsu))
    for emit in samples:
        Draft202012Validator(schemas[2], registry=registry).validate(emit)
    return "bundle v3 + capability v3 + four damage-emit v2 modes validated"


def validate_records(artifact: Mapping[str, Any]) -> tuple[int, int]:
    require(len(artifact["pathways"]) == 1, "expected one pathway")
    pathway = artifact["pathways"][0]
    require(pathway["pathway_id"] == PATHWAY, "wrong pathway")
    require(pathway["hazard_axis"]["routing_field"] == "array_architecture", "wrong routing field")
    contracts = pathway["hazard_axis"]["architecture_input_contracts"]
    require(set(contracts) == {PERRY, FIXED, TRACKER}, "architecture contracts incomplete")
    expected_payloads = {
        PERRY: {
            "source_compatibility": {"perry_event_max_gust_mps", *PERRY_SELECTORS},
        },
        FIXED: {
            "preferred": {
                "tc_fixed_event_to_design_net_pressure_ratio",
                "tc_wind_field_bridge_id",
                "tc_directional_history_bridge_id",
                "tc_duration_cycling_bridge_id",
                "aerodynamic_demand_bridge_id",
                "array_zone",
                "array_spatial_object_id",
            },
            "screening_proxy": {
                "tc_array_height_3s_gust_mps",
                "qualified_design_array_height_3s_gust_mps",
                "tc_wind_field_bridge_id",
                "tc_directional_history_bridge_id",
                "tc_duration_cycling_bridge_id",
                "aerodynamic_demand_bridge_id",
                "array_zone",
                "array_spatial_object_id",
            },
        },
        TRACKER: {
            "qualified_exact_system": set(TRACKER_REQUEST_FIELDS - COMMON_REQUEST_FIELDS),
        },
    }
    for architecture, expected_modes in expected_payloads.items():
        actual_payloads = contracts[architecture]["accepted_payloads"]
        require(len(actual_payloads) == len(expected_modes), f"{architecture}: payload mode count")
        actual_modes = {
            item["mode"]: set(item["required_fields"]) for item in actual_payloads
        }
        require(actual_modes == expected_modes, f"{architecture}: published request contract drift")
    representative = {
        PERRY: perry_request(30.0),
        FIXED: fixed_request(1.0),
        TRACKER: tracker_request(1.0),
    }
    for architecture, request in representative.items():
        require(set(request) <= {PERRY: PERRY_REQUEST_FIELDS, FIXED: FIXED_REQUEST_FIELDS, TRACKER: TRACKER_REQUEST_FIELDS}[architecture], f"{architecture}: helper emits undeclared fields")
        for payload in contracts[architecture]["accepted_payloads"]:
            if payload["mode"] in {"source_compatibility", "preferred", "qualified_exact_system"}:
                require(set(payload["required_fields"]) <= set(request), f"{architecture}: helper misses contract fields")
    records = pathway["curve_records"]
    require(len(records) == 5, "expected five curve records")
    require(len({item["curve_id"] for item in records}) == 5, "duplicate curve IDs")
    require({item["failure_unit_id"] for item in records} == set().union(*ARCHITECTURE_UNITS.values()), "curve units changed")

    v1 = load(PROPOSED / "tropical_cyclone_wind_solar__model_v1_0__docs_r1__curve_artifact.json")
    prior_perry = v1["pathways"][0]["curve_records"][0]
    current_perry = next(item for item in records if item["failure_unit_id"] == PERRY_UNIT)
    require(current_perry == prior_perry, "Perry record is not byte-value compatible")

    shared = load(SHARED)
    require(shared["runtime_approved"] is False, "shared candidate cannot be runtime approved")
    require(shared["comparison_only"] is True, "shared profile is not comparison-only")
    require(shared["origin_strong_wind_artifact_sha256"] == STRONG_V2_ARTIFACT_SHA256, "shared origin pin drift")
    strong_path = ROOT / "docs/cells/strong_wind_solar/proposed/strong_wind_solar__model_v2_0__docs_r1__curve_artifact.json"
    require(sha(strong_path) == STRONG_V2_ARTIFACT_SHA256, "pinned strong-wind source changed")
    require(artifact["audit_comparison_profile"]["sha256"] == sha(SHARED), "shared profile SHA binding drift")
    generic = [item for item in records if item["curve_form"] == "ordered_damage_state_lognormal"]
    require(len(generic) == 4, "expected four generic records")
    comparison_by_unit = {
        item["failure_unit_id"]: item for item in shared["comparison_records"]
    }
    require(set(comparison_by_unit) == {item["failure_unit_id"] for item in generic}, "shared comparison units drift")
    for item in shared["comparison_records"]:
        require(not ({"pathway_id", "curve_id", "x_axis"} & set(item)), "shared comparison leaked cell-local identity")
    for record in generic:
        comparison = comparison_by_unit[record["failure_unit_id"]]
        expected_parameters = json.loads(json.dumps(record["parameters"]))
        for state in expected_parameters["damage_states"]:
            state.pop("source_ids", None)
        for scenario in expected_parameters["capacity_scenarios"]:
            scenario.pop("source_ids", None)
        require(comparison["parameters"] == expected_parameters, "shared audit fingerprint differs from cell-local parameters")
        require(comparison["curve_form"] == record["curve_form"] and comparison["y_axis"] == record["y_axis"], "shared comparison semantics drift")
    for record in generic:
        require(record["pathway_id"] == PATHWAY, "record pathway mismatch")
        require(record["x_axis"] == "architecture_specific_tropical_cyclone_wind_demand_index", "record axis mismatch")
        require("zero_below" not in record["parameters"], "positive hard-zero threshold reintroduced")
        require(record["parameters"]["beta_ln"] in {0.30, 0.275}, "unexpected beta")
        states = record["parameters"]["damage_states"]
        costs = [state["cost_ratio"] for state in states]
        require(costs[0] == 0 and costs[-1] == 1, "state endpoints changed")
        require(all(a <= b for a, b in zip(costs, costs[1:])), "state costs unordered")
        require(all(state["tier"] == "T4_placeholder_or_expert_judgment" for state in states), "state tier changed")
        scenarios = record["parameters"]["capacity_scenarios"]
        require(len(scenarios) == 3, "scenario count changed")
        require({item["scenario_id"] for item in scenarios} == {"lower_resistance", "central_screening", "upper_resistance"}, "scenario set changed")
        for scenario in scenarios:
            require(scenario["tier"] == "T4_placeholder_or_expert_judgment", "scenario tier changed")
            medians = scenario["state_medians"]
            require(len(medians) == len(states) - 1, "median count mismatch")
            require(all(a < b for a, b in zip(medians, medians[1:])), "medians unordered")
        previous = {scenario["scenario_id"]: -1.0 for scenario in scenarios}
        for step in range(401):
            x = 2.0 * step / 400
            evaluated = evaluate_ordered_damage_state_record(record, x)
            require(set(evaluated) == {item["scenario_id"] for item in scenarios}, "evaluated scenario keys drift")
            for scenario_id, values in evaluated.items():
                dr = values["damage_ratio"]
                require(-1e-14 <= dr <= 1 + 1e-14, "DR outside bounds")
                require(dr + 1e-14 >= previous[scenario_id], "DR nonmonotone")
                previous[scenario_id] = dr
                probs = values["state_probabilities"]
                require(set(probs) == {state["state_id"] for state in states}, "state probability keys drift")
                require(all(-1e-14 <= p <= 1 + 1e-14 for p in probs.values()), "probability outside bounds")
                close(sum(probs.values()), 1.0, 1e-12, "state closure")
                recomposed = sum(
                    probs[state["state_id"]] * state["cost_ratio"] for state in states
                )
                close(dr, recomposed, 1e-14, "DR probability-cost recomposition")
            require(
                evaluated["lower_resistance"]["damage_ratio"] + 1e-14
                >= evaluated["central_screening"]["damage_ratio"]
                >= evaluated["upper_resistance"]["damage_ratio"] - 1e-14,
                "resistance ordering failed",
            )
        at_zero = evaluate_ordered_damage_state_record(record, 0.0)
        for values in at_zero.values():
            close(values["damage_ratio"], 0.0, 0.0, "zero demand DR")
        at_positive = evaluate_ordered_damage_state_record(record, 0.05)
        require(all(values["damage_ratio"] > 0 for values in at_positive.values()), "positive demand rounded to policy zero")
    return len(records), len(generic)


def validate_coverage(artifact: Mapping[str, Any], capability: Mapping[str, Any]) -> None:
    unit_rows = artifact["failure_units"]
    units = {item["id"] for item in unit_rows}
    expected = set().union(*ARCHITECTURE_UNITS.values(), COMMON_WITHHELD_UNITS)
    require(len(unit_rows) == len(units) == len(expected), "failure-unit IDs are not unique")
    require(units == expected, "failure-unit registry changed")
    treatments = {item["id"]: item["treatment"] for item in unit_rows}
    require(treatments["PV_REPLACEMENT_SUPPORT"] == "exposure_modifier", "support treatment drift")
    require(all(treatments[unit] == "withheld" for unit in COMMON_WITHHELD_UNITS if unit != "PV_REPLACEMENT_SUPPORT"), "withheld treatment drift")
    coverage = artifact["pathways"][0]["failure_unit_coverage"]
    require(len(coverage) == len(expected), "coverage table size drift")
    require(len({item["failure_unit_id"] for item in coverage}) == len(coverage), "duplicate coverage row")
    require({item["failure_unit_id"] for item in coverage} == expected, "coverage table units drift")
    withheld = capability["pathway_capabilities"][0]["withheld_failure_units"]
    require({item["failure_unit_id"] for item in withheld} == set(COMMON_WITHHELD_UNITS), "withheld units changed")
    require("PV_GSU_SUBSTATION" in set(COMMON_WITHHELD_UNITS), "GSU must be withheld")
    gsu = {
        "event_id": "TC-EVENT-GSU",
        "event_family_id": "TC-FAMILY-GSU",
        "pathway_id": PATHWAY,
        "failure_unit_id": "PV_GSU_SUBSTATION",
    }
    gsu_emit = evaluate_damage_call(artifact, gsu)
    gsu_result = result(gsu_emit, "PV_GSU_SUBSTATION")
    require(gsu_result["status"] == "withheld", "GSU status changed")
    require(gsu_result["scalar_central_dr"] is None, "GSU became numeric")
    require("NO_ARRAY_DR_OR_EXPOSURE_INHERITANCE" in gsu_result["withheld_reason_codes"], "GSU guard missing")
    require(gsu_emit["hazard_input_used"]["array_axis_applied"] is False, "GSU inherited array axis")
    require("array_architecture" not in gsu_emit["selectors_used"], "GSU inherited array architecture")


def validate_kats(artifact: Mapping[str, Any]) -> tuple[int, int, int]:
    fixture = load(KATS)
    require(fixture["semantic_damage_model_version"] == "model v2.0", "KAT model changed")
    expected_runtime_ids = {
        "FIXED_DIRECT_0_0", "FIXED_DIRECT_0_5", "FIXED_DIRECT_1_0", "FIXED_DIRECT_1_5", "FIXED_DIRECT_2_0",
        "FIXED_SPEED_PROXY_0_81", "TRACKER_0_0", "TRACKER_0_75", "TRACKER_1_0", "TRACKER_1_5", "TRACKER_2_0",
        "PERRY_COMPAT_LOW", "PERRY_COMPAT_INTERIOR", "PERRY_COMPAT_HIGH", "GSU_WITHHELD",
    }
    expected_rejection_ids = {
        "REJECT_WRONG_PATHWAY", "REJECT_MISSING_ARCH", "REJECT_UNBRIDGED_10M",
        "REJECT_FIXED_MISSING_DURATION_BRIDGE", "REJECT_TRACKER_ANGLE_MISMATCH",
        "REJECT_TRACKER_COMMAND_ONLY", "REJECT_CROSS_ARCH_UNIT", "REJECT_AXIS_ABOVE_2",
        "REJECT_COMPOUND_WITHOUT_ACK", "REJECT_VALUE_PAYLOAD", "REJECT_PERRY_TAIL",
        "REJECT_UNKNOWN_FIELD", "REJECT_EXPOSURE_ALIAS", "REJECT_NUMERIC_COMPOUND_BOOL",
        "REJECT_PERRY_COMPOSITE_OVERLAP", "REJECT_TRACKER_MISSING_QUALIFICATION_SHA",
        "REJECT_FOREIGN_ROUTE_FIELD", "REJECT_FIXED_DIRECT_WITH_PROXY_COMPANION",
    }
    require({item["test_id"] for item in fixture["runtime_known_answer_tests"]} == expected_runtime_ids, "runtime KAT registry drift")
    require({item["test_id"] for item in fixture["rejection_tests"]} == expected_rejection_ids, "rejection KAT registry drift")
    require({item["test_id"] for item in fixture["artifact_pin_tests"]} == {"PIN_EXACT", "PIN_INCOMPLETE", "PIN_MISMATCH"}, "pin KAT registry drift")
    tolerance = 1e-12
    for test in fixture["runtime_known_answer_tests"]:
        emit = evaluate_damage_call(artifact, test["input"])
        expected = test["expected"]
        require(emit["hazard_input_used"] == expected["hazard_input_used"], test["test_id"] + " hazard lineage")
        require(emit["emit_mode"] == expected["emit_mode"], test["test_id"] + " emit mode")
        require(emit["selectors_used"] == expected["selectors_used"], test["test_id"] + " selectors")
        require(expected["required_flags"] == emit["input_quality"]["limitation_flags"], test["test_id"] + " flags")
        require(emit["hazard_input_used"]["event_id"] == test["input"]["event_id"], test["test_id"] + " event id")
        require(emit["hazard_input_used"]["event_family_id"] == test["input"]["event_family_id"], test["test_id"] + " event family")
        require({item["failure_unit_id"] for item in emit["failure_unit_results"]} == {item["failure_unit_id"] for item in expected["failure_unit_results"]}, test["test_id"] + " result set")
        for item in expected["failure_unit_results"]:
            actual = result(emit, item["failure_unit_id"])
            require(actual["status"] == item["status"], test["test_id"] + " status")
            require(actual["withheld_reason_codes"] == item["withheld_reason_codes"], test["test_id"] + " reasons")
            if item["scalar_central_dr"] is None:
                require(actual["scalar_central_dr"] is None, test["test_id"] + " null DR")
            else:
                close(actual["scalar_central_dr"], item["scalar_central_dr"], tolerance, test["test_id"] + " central")
            require(set(actual["scenario_drs"]) == set(item["scenario_drs"]), test["test_id"] + " scenarios")
            for scenario_id, value in item["scenario_drs"].items():
                close(actual["scenario_drs"][scenario_id], value, tolerance, test["test_id"] + scenario_id)
            expected_probabilities = item["state_probabilities_by_scenario"]
            require(set(actual["state_probabilities_by_scenario"]) == set(expected_probabilities), test["test_id"] + " probability scenarios")
            for scenario_id, probability_map in expected_probabilities.items():
                actual_map = actual["state_probabilities_by_scenario"][scenario_id]
                require(set(actual_map) == set(probability_map), test["test_id"] + scenario_id + " state keys")
                for state_id, probability in probability_map.items():
                    close(actual_map[state_id], probability, tolerance, test["test_id"] + scenario_id + state_id)
                close(sum(actual_map.values()), 1.0, tolerance, test["test_id"] + scenario_id + " closure")
            if actual["status"] == "conditional" and emit["emit_mode"] == "state_ensemble":
                close(actual["scalar_central_dr"], actual["scenario_drs"]["central_screening"], tolerance, test["test_id"] + " central parity")
        route = test["input"].get("array_architecture")
        flags = set(emit["input_quality"]["limitation_flags"])
        if route == PERRY:
            require("PERRY_SOURCE_COMPATIBILITY_ROUTE" in flags, test["test_id"] + " Perry flag")
            require("EXPERIMENTAL_SYNTHETIC_T4_SCENARIO" not in flags, test["test_id"] + " Perry synthetic leak")
        elif route in {FIXED, TRACKER}:
            require("EXPERIMENTAL_SYNTHETIC_T4_SCENARIO" in flags, test["test_id"] + " synthetic flag")
            require("TC_BRIDGE_CONTENT_NOT_RESOLVED_BY_REFERENCE_EVALUATOR" in flags, test["test_id"] + " bridge-content flag")
            require("PERRY_SOURCE_COMPATIBILITY_ROUTE" not in flags, test["test_id"] + " Perry flag leak")
        else:
            require(test["test_id"] == "GSU_WITHHELD" and "axis_value" not in emit["hazard_input_used"], "withheld route axis leak")
    for test in fixture["rejection_tests"]:
        try:
            evaluate_damage_call(artifact, test["input"])
        except TropicalCycloneWindSolarV2EvaluationError as exc:
            require(exc.code == test["expected_error_code"], test["test_id"] + f": got {exc.code}")
        else:
            raise ValidationFailure(test["test_id"] + ": expected rejection")

    digest = artifact_sha256(ARTIFACT)
    exact = {
        "cell_id": artifact["cell_id"],
        "semantic_damage_model_version": artifact["semantic_damage_model_version"],
        "documentation_revision": artifact["documentation_revision"],
        "schema_version": artifact["schema_version"],
        "artifact_sha256": digest,
    }
    verify_artifact_pin(artifact, exact, artifact_sha256_hex=digest)
    incomplete = dict(exact)
    incomplete.pop("artifact_sha256")
    try:
        verify_artifact_pin(artifact, incomplete, artifact_sha256_hex=digest)
    except TropicalCycloneWindSolarV2EvaluationError as exc:
        require(exc.code == "ARTIFACT_PIN_INCOMPLETE", "incomplete pin code changed")
    else:
        raise ValidationFailure("incomplete pin accepted")
    mismatch = dict(exact)
    mismatch["artifact_sha256"] = "0" * 64
    try:
        verify_artifact_pin(artifact, mismatch, artifact_sha256_hex=digest)
    except TropicalCycloneWindSolarV2EvaluationError as exc:
        require(exc.code == "ARTIFACT_PIN_MISMATCH", "mismatch pin code changed")
    else:
        raise ValidationFailure("mismatched pin accepted")
    try:
        evaluate_damage_call(artifact, fixed_request(1.0), verified_artifact_sha256_hex="not-a-digest")
    except TropicalCycloneWindSolarV2EvaluationError as exc:
        require(exc.code == "ARTIFACT_PIN_MISMATCH", "invalid verified digest code changed")
    else:
        raise ValidationFailure("invalid verified digest accepted")
    return len(fixture["runtime_known_answer_tests"]), len(fixture["rejection_tests"]), len(fixture["artifact_pin_tests"])


def validate_registers(artifact: Mapping[str, Any]) -> tuple[int, int, int, int, int, int]:
    sources = rows(SOURCES)
    claims = rows(CLAIMS)
    supersession = rows(CLAIM_SUPERSESSION)
    parameters = rows(PARAMETERS)
    values = rows(VALUES)
    old_new = rows(OLD_NEW)
    require(len(sources) == 71, "source register count drift")
    require(len(claims) == 87, "claim register count drift")
    require(len(supersession) == 8, "claim supersession count drift")
    require(len(parameters) == 53, "parameter register count drift")
    require(len(values) == 18, "value crosswalk changed")
    require(len(old_new) >= 11, "old/new comparison incomplete")
    source_ids = {row["source_id"] for row in sources}
    require(len(source_ids) == len(sources), "duplicate source IDs")
    require({"SHARED_SOLAR_WIND_PROXY_V0_1", CELL_LOCAL_SYNTHETIC_SOURCE, NO_DIRECT_TC_CALIBRATION_SOURCE, "GOVERNANCE_CONTRACT"} <= source_ids, "control sources missing")
    tiers = {
        "T1_claims_or_field_calibrated",
        "T2_public_lab_standard_or_physics",
        "T3_engineering_proxy_or_adjacent_empirical",
        "T4_placeholder_or_expert_judgment",
    }
    for row in sources:
        require(row["evidence_tier"] in tiers, f"invalid source tier {row['source_id']}")
        require(all(row[field] for field in ("citation", "exact_locator", "permitted_inference", "prohibited_inference", "decision", "status")), f"incomplete source {row['source_id']}")
    claim_ids = {row["claim_id"] for row in claims}
    require(len(claim_ids) == len(claims), "duplicate claim IDs")
    require({"TCWS-C216", "TCWS2-C007", "TCWS2-C012", "TCWS2-C013", "TCWS2-C016"} <= claim_ids, "load-bearing claims missing")
    for row in claims:
        require(row["evidence_tier"] in tiers, f"invalid claim tier {row['claim_id']}")
        require(not (split_ids(row["source_ids"]) - source_ids), f"unresolved claim source {row['claim_id']}")
        require(row["permitted_inference"] and row["prohibited_inference"] and row["update_trigger"], f"incomplete claim {row['claim_id']}")
    historical_ids = {"TCWS-C003", "TCWS-C025", "TCWS-C107", "TCWS-C111", "TCWS-C207", "TCWS-C215", "TCWS-C216", "TCWS-C217"}
    claims_by_id = {row["claim_id"]: row for row in claims}
    require(all(claims_by_id[claim_id]["adoption_status"] == "historical_scope_only" for claim_id in historical_ids), "historical claims remain unscoped")
    require({row["prior_claim_id"] for row in supersession} == historical_ids, "supersession map claim set drift")
    require(all(row["superseding_claim_id"] and row["retained_truth"] for row in supersession), "incomplete supersession row")
    for row in parameters:
        require(row["tier"] in tiers, f"invalid parameter tier {row['parameter']}")
        require(not (split_ids(row["source_ids"]) - source_ids), f"unresolved parameter source {row['parameter']}")
    generic_records = [item for item in artifact["pathways"][0]["curve_records"] if item["curve_form"] == "ordered_damage_state_lognormal"]
    expected_numeric: dict[tuple[str, str], str] = {}
    for record in generic_records:
        curve = record["curve_id"]
        expected_numeric[(curve, f"{curve} | beta_ln")] = str(record["parameters"]["beta_ln"])
        for state in record["parameters"]["damage_states"]:
            expected_numeric[(curve, f"{curve} | {state['state_id']} cost_ratio")] = str(state["cost_ratio"])
        for scenario in record["parameters"]["capacity_scenarios"]:
            expected_numeric[(curve, f"{curve} | {scenario['scenario_id']} medians")] = json.dumps(scenario["state_medians"])
    actual_numeric = {
        (row["curve_id"], row["parameter"]): row for row in parameters if row["curve_id"].startswith("TCWS2_")
    }
    require(set(actual_numeric) == set(expected_numeric), "synthetic parameter row identity drift")
    for key, expected_value in expected_numeric.items():
        row = actual_numeric[key]
        require(row["value"] == expected_value, f"synthetic parameter value drift {key}")
        require(row["tier"] == "T4_placeholder_or_expert_judgment", f"synthetic number is not T4 {key}")
        require(split_ids(row["source_ids"]) == {CELL_LOCAL_SYNTHETIC_SOURCE, NO_DIRECT_TC_CALIBRATION_SOURCE}, f"synthetic provenance drift {key}")
    parameter_by_name = {row["parameter"]: row for row in parameters}
    require(parameter_by_name["TCWS2-P000 | package curve-record count"]["value"] == "5", "package record count row drift")
    require(parameter_by_name["TCWS2-P000B | generic synthetic record count"]["value"] == "4", "generic record count row drift")
    require(parameter_by_name["Perry_compatibility_source_derived_record_count"]["value"] == "1", "Perry count row drift")
    require(parameter_by_name["fit_method"]["value"] == "equal_record_weighted_PAVA_then_block_edge_linearization", "PAVA wording regressed")
    embedded = artifact["parameter_tier_table"]
    require(all(item["tier"] in tiers for item in embedded), "embedded parameter tier invalid")
    require(next(item for item in embedded if item["parameter"] == "generic ordered-state form")["tier"] == "T4_placeholder_or_expert_judgment", "ordered-state form over-tiered")
    require(all(row["include_in_direct_denominator"].lower() == "false" for row in values), "runtime value binding enabled")
    value_text = "\n".join(" ".join(row.values()) for row in values)
    require("model v0.1 scenario loss" not in value_text and "no model v0.1 monetary loss" not in value_text, "stale v0.1 value wording")
    require("no TC-wind curve exists" not in value_text, "stale no-curve value wording")
    require(any(row["comparison_id"] == "LEGACY_ASSET_DR" and row["comparability"] == "incomparable" for row in old_new), "legacy incomparability missing")
    return len(sources), len(claims), len(supersession), len(parameters), len(values), len(old_new)


def validate_workbook() -> tuple[int, int, int]:
    require(WORKBOOK.exists() and zipfile.is_zipfile(WORKBOOK), "invalid workbook")
    from openpyxl import load_workbook

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    expected = [
        "README",
        "Dashboard",
        "Scope_Pathway",
        "Architecture_Axes",
        "Curve_Records",
        "Curve_Data",
        "State_Definitions",
        "Perry_Compatibility",
        "Site_Condition_Double_Count",
        "KATs",
        "Value_Crosswalk",
        "Sources",
        "Claim_Register",
        "Claim_Supersession",
        "Parameter_Tiers",
        "Legacy_Comparison",
        "QA_Checks",
    ]
    require(wb.sheetnames == expected, "workbook sheet order changed")
    formulas = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                    require(not any(token in cell.value for token in ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?")), f"formula error token {ws.title}!{cell.coordinate}")
    require(formulas == 190, "workbook formula count drift")
    qa = wb["QA_Checks"]
    statuses = [qa.cell(row, 4).value for row in range(2, qa.max_row + 1)]
    require(len(statuses) == 19 and all(value == "PASS" for value in statuses), "workbook QA failure")
    for row_number in range(2, qa.max_row + 1):
        require(qa.cell(row_number, 2).value == qa.cell(row_number, 3).value, f"workbook QA observed/expected drift row {row_number}")
    perry = wb["Perry_Compatibility"]
    require(perry.max_row == 14, "Perry workbook equality rows changed")
    artifact = load(ARTIFACT)
    pathway = artifact["pathways"][0]
    records = pathway["curve_records"]
    record_rows = list(wb["Curve_Records"].iter_rows(min_row=2, values_only=True))
    require(len(record_rows) == 5, "workbook curve-record row count")
    require({row[0] for row in record_rows} == {record["curve_id"] for record in records}, "workbook curve IDs drift")
    generic = [record for record in records if record["curve_form"] == "ordered_damage_state_lognormal"]
    curve_data = list(wb["Curve_Data"].iter_rows(min_row=2, values_only=True))
    require(len(curve_data) == 164, "workbook curve-data row count")
    cursor = 0
    for record in generic:
        for step in range(41):
            row = curve_data[cursor]
            cursor += 1
            x = step / 20
            require(row[0] == record["curve_id"], "workbook curve-data ID drift")
            close(float(row[1]), x, 1e-15, "workbook curve-data x")
            evaluated = evaluate_ordered_damage_state_record(record, x)
            for column, scenario_id in zip((2, 3, 4), ("lower_resistance", "central_screening", "upper_resistance"), strict=True):
                close(float(row[column]), evaluated[scenario_id]["damage_ratio"], 1e-14, "workbook curve-data DR")
            require(row[5] == f'=IF(AND(C{cursor + 1}>=D{cursor + 1},D{cursor + 1}>=E{cursor + 1}),"PASS","FAIL")', "workbook curve-data QA formula drift")
    expected_states = {
        (record["curve_id"], state["state_id"], state["cost_ratio"], state["tier"], state["description"])
        for record in generic for state in record["parameters"]["damage_states"]
    }
    actual_states = set(wb["State_Definitions"].iter_rows(min_row=2, values_only=True))
    require(actual_states == expected_states, "workbook state definitions drift")
    perry_record = next(record for record in records if record["failure_unit_id"] == PERRY_UNIT)
    for row_number, (x, y) in enumerate(perry_record["parameters"]["points"], 2):
        close(float(perry.cell(row_number, 1).value), x, 0.0, "workbook Perry x")
        close(float(perry.cell(row_number, 2).value), y, 0.0, "workbook Perry v1")
        close(float(perry.cell(row_number, 3).value), y, 0.0, "workbook Perry v2")
    kats = load(KATS)
    expected_kat_rows = {
        (test["test_id"], "runtime") for test in kats["runtime_known_answer_tests"]
    } | {(test["test_id"], "rejection") for test in kats["rejection_tests"]}
    actual_kat_rows = {(row[0], row[1]) for row in wb["KATs"].iter_rows(min_row=2, values_only=True)}
    require(actual_kat_rows == expected_kat_rows, "workbook KAT registry drift")

    def mapping_rows(sheet_name: str) -> list[dict[str, str]]:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        return [
            {header: "" if value is None else str(value) for header, value in zip(headers, row, strict=True)}
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]

    for sheet_name, path in (
        ("Sources", SOURCES),
        ("Claim_Register", CLAIMS),
        ("Claim_Supersession", CLAIM_SUPERSESSION),
        ("Parameter_Tiers", PARAMETERS),
        ("Value_Crosswalk", VALUES),
        ("Legacy_Comparison", OLD_NEW),
    ):
        require(mapping_rows(sheet_name) == rows(path), f"workbook {sheet_name} differs from CSV")
    return len(wb.sheetnames), formulas, len(statuses)


def validate_hash_preservation() -> int:
    count = 0
    for path, expected in {**V0_FILES, **V1_FILES}.items():
        require(path.exists(), f"prior file missing {path.name}")
        require(sha(path) == expected, f"prior file changed {path.name}")
        count += 1
    return count


def validate_index_and_repo_boundary() -> None:
    index = load(INDEX)
    require(
        not [item for item in index["artifacts"] if (ROOT / item["path"]).resolve() == ARTIFACT.resolve()],
        "model-v2.0 proposal entered artifact index",
    )
    require(
        not [
            item for item in index["artifacts"]
            if item["cell_id"] == "tropical_cyclone_wind_solar"
            and item["semantic_damage_model_version"] == "model v2.0"
        ],
        "model-v2.0 proposal identity entered artifact index",
    )
    require(artifact_sha256(ARTIFACT) == sha(ARTIFACT), "artifact digest helper mismatch")
    forbidden = {"hurricane_sites_manual.csv", "ceferino_supplement.docx"}
    require(not [path for path in PROPOSED.rglob("*") if path.name.lower() in forbidden], "raw source vendored")
    text = ARTIFACT.read_text()
    require("straight_line_convective" not in text, "convective runtime semantics leaked")
    require("runtime_approved\": true" not in SHARED.read_text().lower(), "shared candidate promoted")


def validate_local_links() -> int:
    docs = [
        PROPOSED / "README_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "CHANGE_CLASSIFICATION_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "DECISION_LOG_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "tropical_cyclone_wind_solar_curve_derivation_dossier__model_v2_0__docs_r1.md",
        PROPOSED / "tropical_cyclone_wind_solar_damage_code_metadata_spec__model_v2_0__docs_r1.md",
        PROPOSED / "SEVEN_STEP_AUDIT_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "BOUNDED_EVIDENCE_SEARCH_LOG_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "LEGACY_AND_ADJACENT_MODEL_AUDIT_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "NEIGHBORING_WIND_AND_COMPOUND_BOUNDARY_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "SITE_CONDITION_ADAPTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "PRESSURE_TEST_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "PROMOTION_GATE_MATRIX_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "workbook_sheet_manifest_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        PROPOSED / "VALIDATION_REPORT_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        ROOT / "docs/method/shared_components/solar_wind_normalized_response/README.md",
        ROOT / "docs/contracts/hazard_handoff/tropical_cyclone_wind_solar_model_v2_0_synthetic_proposal.md",
        ROOT / "docs/extra/guides/tropical_cyclone_wind_solar_v2_curve_request_guide.md",
    ]
    count = 0
    pattern = re.compile(r"\[[^\]]+\]\(([^)]+)\)")
    for doc in docs:
        require(doc.exists(), f"missing doc {doc}")
        for raw in pattern.findall(doc.read_text()):
            target = raw.strip().split("#", 1)[0]
            if not target or target.startswith(("http://", "https://", "mailto:")):
                continue
            resolved = (doc.parent / target).resolve()
            require(resolved.exists(), f"broken link {doc.relative_to(ROOT)} -> {target}")
            count += 1
    return count


def validate_documented_bindings(
    *,
    record_count: int,
    generic_count: int,
    runtime_kats: int,
    rejection_kats: int,
    pin_kats: int,
    source_count: int,
    claim_count: int,
    supersession_count: int,
    parameter_count: int,
    value_count: int,
    old_new_count: int,
    sheets: int,
    formulas: int,
    workbook_qa: int,
    prior_hashes: int,
    links: int,
) -> None:
    artifact_digest = sha(ARTIFACT)
    capability_digest = sha(CAPABILITY)
    kat_digest = sha(KATS)
    workbook_digest = sha(WORKBOOK)
    shared_digest = sha(SHARED)

    guide = REQUEST_GUIDE.read_text()
    documented_guide_pins = re.findall(
        r'"artifact_sha256"\s*:\s*"([0-9a-f]{64})"', guide
    )
    require(len(documented_guide_pins) == 6, "request guide must contain six exact artifact pins")
    require(set(documented_guide_pins) == {artifact_digest}, "request guide artifact pin is stale or mixed")

    handoff = HANDOFF.read_text()
    documented_handoff_pins = re.findall(
        r"^artifact_sha256:\s*([0-9a-f]{64})$", handoff, flags=re.MULTILINE
    )
    require(documented_handoff_pins == [artifact_digest], "Hazard handoff artifact pin is stale")

    report = VALIDATION_REPORT.read_text()
    require(
        f"\nchecks={EXPECTED_FINAL_CHECKS}\n" in report,
        "validation report overall check count is stale",
    )
    report_hashes = set(re.findall(r"`([0-9a-f]{64})`", report))
    require(
        report_hashes
        == {artifact_digest, capability_digest, kat_digest, workbook_digest, shared_digest},
        "validation report machine hashes are stale or incomplete",
    )
    reported_counts = {
        f"curve_records={record_count}",
        f"generic_synthetic_T4_records={generic_count}",
        f"runtime_kats={runtime_kats}",
        f"rejection_kats={rejection_kats}",
        f"pin_kats={pin_kats}",
        f"sources={source_count}",
        f"claims={claim_count}",
        f"claim_supersession_rows={supersession_count}",
        f"parameters={parameter_count}",
        f"value_rows={value_count}",
        f"old_vs_new_rows={old_new_count}",
        f"workbook_sheets={sheets}",
        f"workbook_formulas={formulas}",
        f"workbook_qa_passes={workbook_qa}",
        f"preserved_prior_hashes={prior_hashes}",
        f"local_links={links}",
    }
    for line in reported_counts:
        require(f"\n{line}\n" in report, f"validation report count is stale: {line}")
    dossier = DERIVATION_DOSSIER.read_text()
    require(f"sha256: {shared_digest}" in dossier, "derivation dossier shared-profile SHA is stale")


def main() -> None:
    if "--allow-incomplete" in sys.argv[1:]:
        raise SystemExit("--allow-incomplete is not supported for the v2 proposal")
    artifact = load(ARTIFACT)
    capability = load(CAPABILITY)
    validate_identity(artifact, capability)
    validate_paths(artifact)
    schema_note = validate_schemas(artifact, capability)
    record_count, generic_count = validate_records(artifact)
    validate_coverage(artifact, capability)
    runtime_kats, rejection_kats, pin_kats = validate_kats(artifact)
    source_count, claim_count, supersession_count, parameter_count, value_count, old_new_count = validate_registers(artifact)
    sheets, formulas, workbook_qa = validate_workbook()
    prior_hashes = validate_hash_preservation()
    validate_index_and_repo_boundary()
    links = validate_local_links()
    validate_documented_bindings(
        record_count=record_count,
        generic_count=generic_count,
        runtime_kats=runtime_kats,
        rejection_kats=rejection_kats,
        pin_kats=pin_kats,
        source_count=source_count,
        claim_count=claim_count,
        supersession_count=supersession_count,
        parameter_count=parameter_count,
        value_count=value_count,
        old_new_count=old_new_count,
        sheets=sheets,
        formulas=formulas,
        workbook_qa=workbook_qa,
        prior_hashes=prior_hashes,
        links=links,
    )
    require(
        Checks.value + 1 == EXPECTED_FINAL_CHECKS,
        "validator check-count constant drifted; review and refresh the validation report",
    )
    print("PASS tropical_cyclone_wind_solar model v2.0/docs r1 synthetic-T4 proposal")
    print(f"checks={Checks.value}")
    print(f"schema_validation={schema_note}")
    print(f"curve_records={record_count}")
    print(f"generic_synthetic_T4_records={generic_count}")
    print(f"runtime_kats={runtime_kats}")
    print(f"rejection_kats={rejection_kats}")
    print(f"pin_kats={pin_kats}")
    print(f"sources={source_count}")
    print(f"claims={claim_count}")
    print(f"claim_supersession_rows={supersession_count}")
    print(f"parameters={parameter_count}")
    print(f"value_rows={value_count}")
    print(f"old_vs_new_rows={old_new_count}")
    print(f"workbook_sheets={sheets}")
    print(f"workbook_formulas={formulas}")
    print(f"workbook_qa_passes={workbook_qa}")
    print(f"preserved_prior_hashes={prior_hashes}")
    print(f"local_links={links}")
    print(f"artifact_sha256={sha(ARTIFACT)}")
    print(f"capability_sha256={sha(CAPABILITY)}")
    print(f"known_answer_tests_sha256={sha(KATS)}")
    print(f"workbook_sha256={sha(WORKBOOK)}")
    print(f"shared_profile_sha256={sha(SHARED)}")


if __name__ == "__main__":
    main()
