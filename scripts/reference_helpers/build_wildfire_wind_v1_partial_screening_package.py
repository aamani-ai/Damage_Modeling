#!/usr/bin/env python3
"""Build the governed wildfire_wind model-v1 partial screening artifacts.

The numerical profiles are cell-local Tier-4 assumptions.  The cited sources
constrain mechanism, ordering, and post-fire disposition; they do not calibrate
the FSim-class-to-economic-DR ordinates.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


REPO = Path(__file__).resolve().parents[2]
OUT = REPO / "docs/cells/wildfire_wind/proposed"
MODEL = "model v1.0"
DOCS = "docs r1"
PATHWAY = "wildfire_thermal_attack"
FSIM_PRODUCT = "USFS_RDS_2016_0034_3_270M"
ASSUMPTION_SET = "WW_T4_PARTIAL_ELECTRICAL_SCREENING_2026_08_08"
FLAGS = [
    "NONCANONICAL_PROPOSAL",
    "SCREENING_ENGINEERING_PROXY_T4",
    "PARTIAL_FAILURE_UNIT_COVERAGE",
    "FSIM_CLASS_IS_NOT_LOCAL_EQUIPMENT_HEAT_FLUX",
    "NOT_FIELD_OR_CLAIMS_CALIBRATED",
    "NO_AUTOMATIC_MITIGATION_CREDIT",
    "CURVE_INTRINSIC_SPREAD_NOT_CARRIED",
]

CURVES = {
    "WT_PAD_ELECTRICAL": {
        "curve_id": "WWV1_PAD_ELECTRICAL_FSIM_T4",
        "points": [[0, 0.0], [1, 0.001], [2, 0.006], [3, 0.03], [4, 0.12], [5, 0.35], [6, 0.70]],
        "subsystem": "REPEATED_TURBINE_PAD_ELECTRICAL",
        "component": "PAD_OR_TURBINE_STEPUP_TRANSFORMER_SWITCHGEAR_AND_LOCAL_TERMINATIONS",
        "denominator": "direct replacement value of the same exposed pad-electrical failure unit",
        "profile_note": "Lower T4 response than exposed controls because common apparatus has a steel exterior; polymeric, cable, seal, and auxiliary contents keep risk nonzero.",
    },
    "WT_GSU_PROTECTION_CONTROL_DC": {
        "curve_id": "WWV1_GSU_PROTECTION_CONTROL_DC_FSIM_T4",
        "points": [[0, 0.0], [1, 0.004], [2, 0.02], [3, 0.08], [4, 0.25], [5, 0.60], [6, 0.90]],
        "subsystem": "FACILITY_GSU_SUBSTATION",
        "component": "PROTECTION_RELAY_CONTROL_SCADA_COMMUNICATIONS_STATION_SERVICE_AND_DC",
        "denominator": "direct replacement value of the same GSU protection-control-DC package",
        "profile_note": "Higher T4 response reflects polymeric, electronic, communications, battery, and cable content plus fire/heat replacement guidance; it is not a measured probability curve.",
    },
}

WITHHELD = {
    "WT_TURBINE_FIRE_ASSEMBLY": ["NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT", "FULL_TURBINE_ECONOMIC_RESPONSE_NOT_CALIBRATED"],
    "WT_COLLECTION_NETWORK": ["NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT", "LINE_EXPOSURE_AND_CONSTRUCTION_STATE_REQUIRED"],
    "WT_GSU_MAIN_TRANSFORMER": ["NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT", "TRANSFORMER_THERMAL_STATE_AND_DISPOSITION_MODEL_REQUIRED"],
    "WT_GSU_SWITCHGEAR_BUS": ["NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT", "SWITCHGEAR_RECONDITION_OR_REPLACE_MODEL_REQUIRED"],
    "WT_GSU_CABLE_TERMINATIONS": ["NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT", "CABLE_BODY_TERMINATION_AND_INGRESS_SPLIT_REQUIRED"],
    "WT_CONTROL_MET_OM": ["NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT", "MIXED_BUILDING_AND_INSTRUMENT_INVENTORY_REQUIRES_SPLIT"],
    "WT_FOUNDATION": ["NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT", "DIRECT_THERMAL_RESPONSE_NOT_CALIBRATED"],
    "WT_CIVIL_INFRA": ["NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT", "MIXED_CIVIL_BUCKET_REQUIRES_SPLIT"],
    "SUPPORT_FIELDWORK": ["NO_INDEPENDENT_FRAGILITY", "SUPPORT_COST_ALLOCATE_ONCE_AFTER_DISPOSITION"],
    "SUPPORT_TRANSPORT_LOGISTICS": ["NO_INDEPENDENT_FRAGILITY", "SUPPORT_COST_ALLOCATE_ONCE_AFTER_DISPOSITION"],
}


def dump_json(name: str, payload: object) -> None:
    (OUT / name).write_text(json.dumps(payload, indent=2) + "\n")


def write_csv(name: str, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    with (OUT / name).open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def capability() -> dict:
    withheld = [
        {"failure_unit_id": unit, "reason_codes": reasons}
        for unit, reasons in WITHHELD.items()
    ]
    return {
        "schema_version": "capability_declaration.v3",
        "cell_id": "wildfire_wind",
        "canonical_runtime_artifact": False,
        "pathway_capabilities": [{
            "pathway_id": PATHWAY,
            "failure_unit_scalar_dr": "conditional",
            "scenario_loss_given_value_basis": "withheld",
            "curve_intrinsic_spread": "not_carried",
            "populated_emit_modes": ["scalar_mean"],
            "conditions": [
                "exact USFS FSim RDS-2016-0034-3 product identity and source-native conditional flame-length class are required",
                "class state must be an exact integer from 1 through 6; state 0 is a damage-code no-event control only",
                f"screening_assumption_set_id must equal {ASSUMPTION_SET}",
                "the output applies only to the named failure unit and is not a whole-wind-farm damage ratio",
                "firebrand, destructive residue, internal turbine fire, BI, frequency, and tail metrics are outside this pathway",
            ],
            "limitation_flags": FLAGS + ["NUMERICAL_IDENTITY_TO_WILDFIRE_SOLAR_IS_AUDIT_FINGERPRINT_ONLY"],
            "withheld_failure_units": withheld,
        }],
        "consumer_annual_metrics": {
            "computation_owner": "downstream_consumer",
            "status_before_promotion": "withheld_noncanonical_proposal",
            "status_after_promotion": "withheld",
            "prerequisites": [
                "canonical promotion and exact model docs schema and SHA pin",
                "validated wildfire frequency and source-class spatial coupling",
                "reviewed local-attack bridge or explicit continued screening-use approval",
                "complete same-unit value ownership and exposure binding",
                "compound-event and dependent-spread aggregation controls",
            ],
            "limitation_flags": FLAGS + ["ANNUAL_AND_TAIL_METRICS_WITHHELD"],
        },
        "cap_binding": {
            "policy": "consumer_enforced_fail_closed",
            "enforcement_owner": "downstream_consumer",
            "checks_required": [
                "reject whole-plant DR and annual or tail metrics from this proposal",
                "bind each DR only to the same named failure-unit value and local exposure fraction",
                "prohibit the mixed 72 USD/kW electrical row and full-project TIV as implicit values",
                "represent a shared GSU protection-control-DC package once",
                "preserve nulls and reason codes for every withheld unit and pathway",
            ],
            "action_if_fail": "withhold scenario loss and all annual or tail metrics",
        },
        "promotion_gate": {
            "status": "blocked",
            "required_before_canonical_use": [
                "independent wildfire-science review of FSim screening-class use and limitations",
                "independent wind-electrical engineering review of the two failure-unit boundaries and ordering",
                "independent review or structured elicitation of all Tier-4 ordinates",
                "site-level pad and GSU value ownership and exposure fixtures",
                "Hazard adapter and no-fallback known-answer tests",
                "shadow comparison rollback plan and explicit promotion decision",
            ],
        },
    }


def failure_units() -> list[dict]:
    result = []
    for unit, spec in CURVES.items():
        result.append({
            "id": unit,
            "subsystem": spec["subsystem"],
            "component": spec["component"],
            "treatment": "primary_nonzero",
            "y_axis": "failure_unit_damage_ratio",
            "denominator": spec["denominator"],
            "exposure_grain": "one physical pad-electrical unit" if unit == "WT_PAD_ELECTRICAL" else "one physical shared GSU protection-control-DC package",
            "profile_note": spec["profile_note"],
        })
    for unit, reasons in WITHHELD.items():
        result.append({
            "id": unit,
            "subsystem": "WITHHELD_OR_SUPPORT",
            "component": unit,
            "treatment": "withheld",
            "y_axis": "not_applicable_support_cost" if unit.startswith("SUPPORT_") else "failure_unit_damage_ratio",
            "denominator": "same-unit direct replacement value when later qualified",
            "withheld_reason_codes": reasons,
        })
    return result


def artifact(cap: dict) -> dict:
    coverage = [
        {"failure_unit_id": unit, "status": "conditional_numeric_screening_curve", "reason_codes": []}
        for unit in CURVES
    ] + [
        {"failure_unit_id": unit, "status": "withheld_not_zero", "reason_codes": reasons}
        for unit, reasons in WITHHELD.items()
    ]
    records = []
    for unit, spec in CURVES.items():
        records.append({
            "curve_id": spec["curve_id"],
            "pathway_id": PATHWAY,
            "failure_unit_id": unit,
            "curve_form": "piecewise_linear",
            "x_axis": "fsim_conditional_flame_length_class_state_exact_integer_only",
            "y_axis": "failure_unit_damage_ratio",
            "parameters": {"points": spec["points"]},
            "valid_range": [0, 6],
            "interpolation_policy": "linear_between_source_knots",
            "extrapolation_policy": "reject noninteger unknown or out-of-range states; evaluator does not execute interpolation",
            "selector_match": {
                "source_wildfire_product_id": FSIM_PRODUCT,
                "screening_assumption_set_id": ASSUMPTION_SET,
            },
            "source_parameter_refs": ["WW1-A001", "WW1-S001", "WW1-S002", "WW1-S003", "WW1-S004"],
            "metadata_flags": FLAGS + ["CELL_LOCAL_T4_ORDINATES"],
        })
    tiers = [
        {"parameter": "FSim product and six conditional flame-length class semantics", "curve_id": "all", "value": FSIM_PRODUCT, "param_role": "axis_semantics", "tier": "T2_public_lab_standard_or_physics", "source_ids": ["WW1-S001"], "reasoning": "Official source-product semantics only; no equipment-demand conversion.", "update_trigger": "source product changes"},
        {"parameter": "GSU polymeric ignition ordering", "curve_id": "all", "value": "controls/electronics more vulnerable than steel-enclosed apparatus", "param_role": "relative_vulnerability_order", "tier": "T3_engineering_proxy_or_adjacent_empirical", "source_ids": ["WW1-S002", "WW1-S003", "WW1-S004"], "reasoning": "Primary substation modeling and equipment guidance support the ordering, not the numeric ordinates.", "update_trigger": "wind-site fire inspection data"},
        {"parameter": "pad electrical ordinates", "curve_id": CURVES["WT_PAD_ELECTRICAL"]["curve_id"], "value": str(CURVES["WT_PAD_ELECTRICAL"]["points"]), "param_role": "curve_shape_and_level", "tier": "T4_placeholder_or_expert_judgment", "source_ids": ["WW1-A001"], "reasoning": "Owner-authorized cell-local screening assumptions; numerically identical neighboring profile is audit comparison only.", "update_trigger": "qualified field/lab/claims calibration or structured elicitation"},
        {"parameter": "GSU protection-control-DC ordinates", "curve_id": CURVES["WT_GSU_PROTECTION_CONTROL_DC"]["curve_id"], "value": str(CURVES["WT_GSU_PROTECTION_CONTROL_DC"]["points"]), "param_role": "curve_shape_and_level", "tier": "T4_placeholder_or_expert_judgment", "source_ids": ["WW1-A001"], "reasoning": "Owner-authorized cell-local screening assumptions; not an ignition-probability or claims curve.", "update_trigger": "qualified field/lab/claims calibration or structured elicitation"},
        {"parameter": "noninteger and out-of-range handling", "curve_id": "all", "value": "reject", "param_role": "boundary_or_cap", "tier": "T4_placeholder_or_expert_judgment", "source_ids": ["WW1-A001", "GOVERNANCE_CONTRACT"], "reasoning": "Avoids inventing false continuity between source-native categorical classes.", "update_trigger": "governed continuous local-demand bridge"},
    ]
    return {
        "schema_version": "damage_curve_record_bundle.v3",
        "schema_status": "proposed_draft",
        "cell_id": "wildfire_wind",
        "damage_code_id": "WILDFIRE_WIND_PARTIAL_ELECTRICAL_SCREENING_V1",
        "semantic_damage_model_version": MODEL,
        "documentation_revision": DOCS,
        "lifecycle_state": "release_candidate",
        "promotion_status": "proposed",
        "review_status": "pressure_tested_pending_independent_review",
        "model_grade": "screening_engineering_proxy_t4",
        "package_release": "unreleased",
        "package_baseline": "library v2.5",
        "package_inclusion_status": "not_included",
        "canonical_runtime_artifact": False,
        "source_dossier": "docs/cells/wildfire_wind/proposed/wildfire_wind_curve_derivation_dossier__model_v1_0__docs_r1.md",
        "source_workbook": "docs/cells/wildfire_wind/proposed/damage_curve_records_wildfire_wind__model_v1_0__docs_r1.xlsx",
        "known_answer_tests": "docs/cells/wildfire_wind/proposed/known_answer_tests_wildfire_wind__model_v1_0__docs_r1.json",
        "source_register": "docs/cells/wildfire_wind/proposed/SOURCE_REGISTER_wildfire_wind__model_v1_0__docs_r1.csv",
        "claim_parameter_register": "docs/cells/wildfire_wind/proposed/CLAIM_PARAMETER_REGISTER_wildfire_wind__model_v1_0__docs_r1.csv",
        "value_crosswalk": "docs/cells/wildfire_wind/proposed/VALUE_CROSSWALK_wildfire_wind__model_v1_0__docs_r1.csv",
        "failure_units": failure_units(),
        "pathways": [{
            "pathway_id": PATHWAY,
            "hazard_scope": {
                "included": ["screening-only exogenous wildfire thermal attack represented by the exact FSim conditional flame-length class at the subject cell"],
                "excluded": ["firebrand attack", "destructive residue alone", "internal turbine or electrical fire", "direct local heat-flux inference", "business interruption", "frequency and annual or tail metrics"],
                "compound_event_rule": "Preserve event_family_id and prevent duplicate value charges across thermal attack, dependent internal spread, firebrand, residue, and post-fire hazards.",
            },
            "hazard_axis": {
                "id": "FSIM_CONDITIONAL_FLAME_LENGTH_CLASS_STATE_SCREENING",
                "preferred_input_field": "conditional_flame_length_class_state",
                "unit": "source_native_categorical_class",
                "valid_range": [0, 6],
                "extrapolation_policy": "exact integers only; 0 is no-event control; reject unknown noninteger and out-of-range values",
                "source_product": FSIM_PRODUCT,
                "native_class_map": {"lt_2_ft": 1, "gte_2_lt_4_ft": 2, "gte_4_lt_6_ft": 3, "gte_6_lt_8_ft": 4, "gte_8_lt_12_ft": 5, "gte_12_ft": 6},
                "guardrail": "The categorical state is not a conversion to equipment heat flux, duration, or ignition probability.",
            },
            "selector_logic": [
                {"field": "source_wildfire_product_id", "required": True, "allowed": [FSIM_PRODUCT], "effect": "pins source semantics and spatial support"},
                {"field": "screening_assumption_set_id", "required": True, "allowed": [ASSUMPTION_SET], "effect": "acknowledges Tier-4 cell-local ordinates and limitations"},
            ],
            "conditioner_logic": [
                {"field": "mitigation_or_clearance_credit", "required": False, "numeric_effect": "none", "reason": "no protection efficacy is calibrated"},
                {"field": "energization_or_shutdown_state", "required": "capture_if_known", "numeric_effect": "none_in_model_v1_0", "flag_when_missing": "OPERATING_STATE_NOT_MODELED"},
            ],
            "exposure_contract": {
                "required_subject_grain": "one physical pad-electrical unit or one physical shared GSU protection-control-DC package",
                "whole_wind_farm_default": "prohibited",
                "spatial_rule": "couple each point or small footprint independently to its intersecting FSim grid cell; never spread one class over the lease polygon by default",
                "value_rule": "no implicit value; same-unit site SOV and exposure fraction are required for future scenario loss",
            },
            "failure_unit_coverage": coverage,
            "curve_records": records,
        }],
        "value_linkage": {
            "curve_denominator": "same named failure-unit direct replacement value",
            "scenario_loss_formula": "loss_u = DR_u * same_unit_direct_replacement_value_usd * exposure_fraction_u",
            "scenario_loss_status": "withheld_missing_site_split_and_noncanonical_proposal",
            "required_fields": ["same_unit_direct_replacement_value_usd", "exposure_fraction", "owner_entity_id", "project_owned", "value_basis_id"],
            "implicit_default_profile": None,
            "mixed_72_usd_per_kw_electrical_row_allowed": False,
            "full_project_tiv_allowed": False,
            "assembly_component_double_count_rule": "Use each physical pad and the one shared GSU package once; do not also apply an aggregate electrical or substation DR.",
        },
        "parameter_tier_table": tiers,
        "evaluation_contract": {
            "supported_pathway_id": PATHWAY,
            "supported_failure_unit_ids": list(CURVES),
            "required_selector_fields": ["source_wildfire_product_id", "screening_assumption_set_id"],
            "accepted_axis_payloads": [{"mode": "exact_fsim_class_state", "required_fields": ["conditional_flame_length_class_state"]}],
            "metadata_flags_always": FLAGS,
            "out_of_range_behavior": "reject without numeric fallback",
            "unsupported_unit_behavior": "return null DR with NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT",
            "scenario_loss_behavior": "evaluator emits DR only; no implicit site value or whole-plant aggregation",
        },
        "emit_contract": {
            "schema_version": "damage_emit.v2",
            "populated_emit_modes_for_this_cell": ["scalar_mean"],
            "supported_output": "conditional scalar mean DR for two named electrical failure units only",
            "scenario_loss": "withheld",
            "annual_and_tail_metrics": "withheld",
            "withheld_units_emit": "Return explicit null DR plus reason codes; never convert missing curves to zero.",
        },
        "capability_declaration": cap,
    }


def kats() -> dict:
    tests = []
    for unit, spec in CURVES.items():
        for state, dr in spec["points"]:
            tests.append({
                "test_id": f"{unit.lower()}_state_{state}",
                "input": {
                    "event_id": "WW-KAT-EVENT",
                    "event_family_id": "WW-KAT-FAMILY",
                    "pathway_id": PATHWAY,
                    "failure_unit_id": unit,
                    "source_wildfire_product_id": FSIM_PRODUCT,
                    "screening_assumption_set_id": ASSUMPTION_SET,
                    "conditional_flame_length_class_state": state,
                },
                "expected": {"status": "conditional", "curve_id": spec["curve_id"], "failure_unit_damage_ratio": dr},
            })
    negative = [
        {"test_id": "reject_noninteger_class", "mutate": {"conditional_flame_length_class_state": 3.5}, "expected_error": "INVALID_CLASS_STATE"},
        {"test_id": "reject_out_of_range_class", "mutate": {"conditional_flame_length_class_state": 7}, "expected_error": "INVALID_CLASS_STATE"},
        {"test_id": "reject_wrong_fsim_product", "mutate": {"source_wildfire_product_id": "UNKNOWN"}, "expected_error": "SELECTOR_MISMATCH"},
        {"test_id": "reject_missing_assumption_acknowledgement", "remove": "screening_assumption_set_id", "expected_error": "MISSING_REQUIRED_FIELD"},
        {"test_id": "reject_firebrand_fallback", "mutate": {"pathway_id": "wildfire_firebrand_ignition"}, "expected_error": "UNSUPPORTED_PATHWAY"},
        {"test_id": "withhold_main_transformer", "mutate": {"failure_unit_id": "WT_GSU_MAIN_TRANSFORMER"}, "expected_status": "withheld", "expected_reason": "NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT"},
    ]
    return {
        "schema_version": "known_answer_tests.v3",
        "cell_id": "wildfire_wind",
        "semantic_damage_model_version": MODEL,
        "documentation_revision": DOCS,
        "absolute_tolerance": 1e-12,
        "formula_known_answer_tests": tests,
        "negative_contract_tests": negative,
    }


def build_registers() -> None:
    source_fields = ["source_id", "citation", "url", "accessed_on", "exact_locator", "source_type", "source_role", "evidence_tier", "permitted_inference", "prohibited_inference", "decision", "notes"]
    sources = [
        {"source_id": "WW1-S001", "citation": "Dillon et al. Spatial datasets of probabilistic wildfire risk components for the United States (270 m), 3rd ed., USFS RDS-2016-0034-3 (2023).", "url": "https://www.fs.usda.gov/rds/archive/products/RDS-2016-0034-3/_metadata_RDS-2016-0034-3.html", "accessed_on": "2026-08-08", "exact_locator": "conditional flame-length probability class definitions and 270 m support", "source_type": "official hazard dataset metadata", "source_role": "axis_semantics", "evidence_tier": "T2_public_lab_standard_or_physics", "permitted_inference": "Preserve exact product and six source-native conditional classes.", "prohibited_inference": "Infer local equipment heat flux, duration, ignition probability, or economic DR.", "decision": "adopt_with_limits", "notes": "Burn probability remains a Hazard frequency input."},
        {"source_id": "WW1-S002", "citation": "Severino et al. Assessing wildfire risk to critical infrastructure in central Chile: application to an electrical substation. International Journal of Wildland Fire (2024). DOI 10.1071/WF22113.", "url": "https://doi.org/10.1071/WF22113", "accessed_on": "2026-08-08", "exact_locator": "substation target inventory; PMMA proxy-fuel radiant-flux probit; case heat-flux results", "source_type": "peer-reviewed primary modeling study", "source_role": "mechanism_and_ordering", "evidence_tier": "T3_engineering_proxy_or_adjacent_empirical", "permitted_inference": "Support nonzero vulnerability of polymeric substation contents and nonlinear ignition sensitivity to local radiant heat.", "prohibited_inference": "Transfer site frequency, PMMA ignition probability, local heat flux, or economic DR to a wind farm or FSim class.", "decision": "adopt_with_limits", "notes": "Pressure-test anchor, not numerical calibration."},
        {"source_id": "WW1-S003", "citation": "NEMA GD 2-2016, Evaluating Fire- and Heat-Damaged Electrical Equipment.", "url": "https://www.nema.org/docs/default-source/standards-document-library/nema-gd-2-2016-evaluating-fire-and-heat-damaged-electrical-equipment-guide.pdf", "accessed_on": "2026-08-08", "exact_locator": "sections on switchgear, transformers, wire/cable, protective relays/meters, communications, UPS and residue", "source_type": "industry consensus guidance", "source_role": "post_fire_disposition", "evidence_tier": "T2_public_lab_standard_or_physics", "permitted_inference": "Support replacement-prone endpoints for damaged electronic, relay, communications, UPS, and cable items and qualified review of apparatus.", "prohibited_inference": "Infer pre-event wildfire fragility, FSim-class DR, or generic smoke loss.", "decision": "adopt_with_limits", "notes": "Disposition evidence only."},
        {"source_id": "WW1-S004", "citation": "Butler, Wallace, and Hogge. Wildfire vulnerability of power transmission and distribution infrastructure (USFS, 2015).", "url": "https://research.fs.usda.gov/treesearch/49452", "accessed_on": "2026-08-08", "exact_locator": "preliminary steel-tower and transformer/junction-enclosure simulations", "source_type": "official preliminary engineering study", "source_role": "negative_and_ordering_evidence", "evidence_tier": "T3_engineering_proxy_or_adjacent_empirical", "permitted_inference": "Keep common steel-enclosed apparatus below exposed polymeric/electronic contents in a screening ordering.", "prohibited_inference": "Assign immunity, a universal zero, or a numeric wind-farm DR.", "decision": "adopt_with_limits", "notes": "Preliminary and not wind-farm calibrated."},
        {"source_id": "WW1-S005", "citation": "Damage Modeling wildfire_solar model v1.0/docs r3.", "url": "docs/cells/wildfire_solar/current/wildfire_solar__model_v1_0__docs_r3__curve_artifact.json", "accessed_on": "2026-08-08", "exact_locator": "WSV1_MV_EQUIPMENT_THERMAL and WSV1_SCADA_THERMAL point arrays", "source_type": "neighboring governed model", "source_role": "audit_comparison_only", "evidence_tier": "T4_placeholder_or_expert_judgment", "permitted_inference": "Record numerical identity after independent cell-local adoption as an audit fingerprint.", "prohibited_inference": "Claim evidence transfer, shared runtime dependency, or solar-to-wind calibration.", "decision": "audit_only", "notes": "Not included in source_parameter_refs as scientific evidence."},
        {"source_id": "WW1-A001", "citation": "Owner direction in Damage Modeling task, 2026-08-08: represent wildfire-wind risk even if only one or two subsystems can be supported.", "url": "conversation_record", "accessed_on": "2026-08-08", "exact_locator": "explicit coverage-first Tier-4 business decision", "source_type": "owner assumption authorization", "source_role": "ordinate_adoption_authority", "evidence_tier": "T4_placeholder_or_expert_judgment", "permitted_inference": "Create a visibly labeled noncanonical two-unit screening proposal.", "prohibited_inference": "Relabel assumptions as empirical calibration or authorize canonical cutover.", "decision": "adopt", "notes": f"Assumption set {ASSUMPTION_SET}."},
        {"source_id": "GOVERNANCE_CONTRACT", "citation": "Damage Modeling governed damage-curve standards and skill workflow.", "url": "docs/method/standards/00_index.md", "accessed_on": "2026-08-08", "exact_locator": "cell package, provenance, pathway, capability, versioning, and fail-closed standards", "source_type": "internal governance control", "source_role": "governance", "evidence_tier": "T4_placeholder_or_expert_judgment", "permitted_inference": "Control package structure, labels, tests, and release state.", "prohibited_inference": "Serve as scientific calibration.", "decision": "adopt", "notes": "Governance only."},
    ]
    write_csv("SOURCE_REGISTER_wildfire_wind__model_v1_0__docs_r1.csv", source_fields, sources)

    claim_fields = ["claim_id", "claim_text", "claim_type", "source_ids", "evidence_tier", "parameter_or_rule", "adoption_status", "permitted_inference", "prohibited_inference", "reasoning", "update_trigger"]
    claims = [
        {"claim_id": "WW1-C001", "claim_text": "FSim conditional flame-length class is a screening ordinal and not a local equipment heat-flux history.", "claim_type": "source_semantics", "source_ids": "WW1-S001", "evidence_tier": "T2_public_lab_standard_or_physics", "parameter_or_rule": "axis_guardrail", "adoption_status": "adopt", "permitted_inference": "Select an exact categorical screening state.", "prohibited_inference": "Convert class to kW/m2, duration, or ignition probability.", "reasoning": "The source product does not carry the component-demand bridge.", "update_trigger": "Validated local-attack adapter."},
        {"claim_id": "WW1-C002", "claim_text": "Two named electrical failure units may emit conditional screening DR; all other wind-farm units remain withheld-not-zero.", "claim_type": "governance_decision", "source_ids": "WW1-A001;GOVERNANCE_CONTRACT", "evidence_tier": "T4_placeholder_or_expert_judgment", "parameter_or_rule": "partial_coverage", "adoption_status": "adopt", "permitted_inference": "Report risk for the two named units only.", "prohibited_inference": "Report whole-farm DR or zero for missing units.", "reasoning": "Partial honest coverage is more useful than an unqualified aggregate.", "update_trigger": "Additional failure-unit calibration or owner decision."},
        {"claim_id": "WW1-C003", "claim_text": "Pad electrical response is screened below protection-control-DC response.", "claim_type": "engineering_ordering", "source_ids": "WW1-S002;WW1-S003;WW1-S004", "evidence_tier": "T3_engineering_proxy_or_adjacent_empirical", "parameter_or_rule": "relative_vulnerability_order", "adoption_status": "adopt_with_limits", "permitted_inference": "Pressure-test the ordinal ordering.", "prohibited_inference": "Claim the numeric gap is measured.", "reasoning": "Steel enclosure offers relative resistance while polymeric/electronic/cable contents are replacement-prone after fire/heat damage.", "update_trigger": "Wind-site inspection and claims evidence."},
        {"claim_id": "WW1-C004", "claim_text": "Both point arrays are Tier-4 cell-local assumptions.", "claim_type": "assumption", "source_ids": "WW1-A001", "evidence_tier": "T4_placeholder_or_expert_judgment", "parameter_or_rule": "curve_ordinates", "adoption_status": "adopt_for_noncanonical_screening", "permitted_inference": "Use for sensitivity and interface testing with exact acknowledgement.", "prohibited_inference": "Call the arrays wildfire-wind calibration, claims experience, or standard curves.", "reasoning": "No public matched demand-to-disposition-to-cost dataset was found.", "update_trigger": "Structured elicitation or qualified empirical calibration."},
        {"claim_id": "WW1-C005", "claim_text": "Numerical identity to two wildfire-solar profiles is an audit fingerprint only.", "claim_type": "cross_cell_boundary", "source_ids": "WW1-S005;WW1-A001", "evidence_tier": "T4_placeholder_or_expert_judgment", "parameter_or_rule": "no_shared_runtime_dependency", "adoption_status": "adopt", "permitted_inference": "Detect drift and explain lineage.", "prohibited_inference": "Use the solar package as evidence or runtime fallback.", "reasoning": "The wind cell independently owns the assumptions and failure-unit binding.", "update_trigger": "Independent wind calibration replaces the arrays."},
        {"claim_id": "WW1-C006", "claim_text": "Scenario dollars, whole-plant DR, EAL, PML, VaR, and TVaR remain withheld.", "claim_type": "capability_boundary", "source_ids": "GOVERNANCE_CONTRACT", "evidence_tier": "T4_placeholder_or_expert_judgment", "parameter_or_rule": "capability", "adoption_status": "adopt", "permitted_inference": "Emit unit DR only.", "prohibited_inference": "Bind mixed electrical USD/kW or full TIV by default.", "reasoning": "Site splits, exposure, frequency, aggregation, and canonical promotion are absent.", "update_trigger": "All recorded gates pass."},
    ]
    write_csv("CLAIM_PARAMETER_REGISTER_wildfire_wind__model_v1_0__docs_r1.csv", claim_fields, claims)

    tier_fields = ["parameter_id", "curve_id", "parameter", "value", "role", "tier", "source_ids", "status", "update_trigger"]
    tiers = [
        {"parameter_id": "WW1-P001", "curve_id": "all", "parameter": "FSim class semantics", "value": "states 1..6; state 0 no-event control", "role": "axis", "tier": "T2_public_lab_standard_or_physics", "source_ids": "WW1-S001", "status": "adopt_with_limits", "update_trigger": "source product change"},
        {"parameter_id": "WW1-P002", "curve_id": CURVES["WT_PAD_ELECTRICAL"]["curve_id"], "parameter": "point array", "value": json.dumps(CURVES["WT_PAD_ELECTRICAL"]["points"]), "role": "shape_and_level", "tier": "T4_placeholder_or_expert_judgment", "source_ids": "WW1-A001", "status": "adopt_noncanonical_screening", "update_trigger": "qualified calibration or elicitation"},
        {"parameter_id": "WW1-P003", "curve_id": CURVES["WT_GSU_PROTECTION_CONTROL_DC"]["curve_id"], "parameter": "point array", "value": json.dumps(CURVES["WT_GSU_PROTECTION_CONTROL_DC"]["points"]), "role": "shape_and_level", "tier": "T4_placeholder_or_expert_judgment", "source_ids": "WW1-A001", "status": "adopt_noncanonical_screening", "update_trigger": "qualified calibration or elicitation"},
        {"parameter_id": "WW1-P004", "curve_id": "all", "parameter": "exact integer handling", "value": "reject noninteger/unknown/out-of-range", "role": "boundary", "tier": "T4_placeholder_or_expert_judgment", "source_ids": "WW1-A001;GOVERNANCE_CONTRACT", "status": "adopt", "update_trigger": "continuous local-demand bridge"},
        {"parameter_id": "WW1-P005", "curve_id": "all", "parameter": "intrinsic spread", "value": "not_carried", "role": "uncertainty", "tier": "T4_placeholder_or_expert_judgment", "source_ids": "WW1-A001", "status": "withheld", "update_trigger": "structured uncertainty elicitation or empirical residuals"},
    ]
    write_csv("PARAMETER_TIER_TABLE_wildfire_wind__model_v1_0__docs_r1.csv", tier_fields, tiers)

    value_fields = ["failure_unit_id", "spatial_grain", "curve_denominator", "available_reference_value", "scenario_loss_status", "required_site_fields", "double_count_guardrail", "notes"]
    values = [
        {"failure_unit_id": "WT_PAD_ELECTRICAL", "spatial_grain": "one turbine/pad point or small footprint", "curve_denominator": CURVES["WT_PAD_ELECTRICAL"]["denominator"], "available_reference_value": "none; mixed NREL 72 USD/kW row prohibited", "scenario_loss_status": "withheld", "required_site_fields": "unit count; construction; ownership; same-unit SOV; local class; exposure fraction", "double_count_guardrail": "Do not also charge collection or aggregate electrical value.", "notes": "DR is usable without dollars as a unit-level screening indicator."},
        {"failure_unit_id": "WT_GSU_PROTECTION_CONTROL_DC", "spatial_grain": "one shared GSU control building/room/cabinet package", "curve_denominator": CURVES["WT_GSU_PROTECTION_CONTROL_DC"]["denominator"], "available_reference_value": "none; mixed NREL 72 USD/kW row prohibited", "scenario_loss_status": "withheld", "required_site_fields": "exact inventory; project ownership; same-unit SOV; local class; exposure fraction", "double_count_guardrail": "Value the physical shared package once across wind, solar, or hybrid hosts.", "notes": "Main transformer, switchgear, and cable terminations remain separate and withheld."},
        {"failure_unit_id": "FULL_WIND_FARM", "spatial_grain": "lease/project polygon", "curve_denominator": "prohibited", "available_reference_value": "full TIV prohibited", "scenario_loss_status": "withheld", "required_site_fields": "not applicable", "double_count_guardrail": "No aggregate alias or default value shares.", "notes": "Two-unit coverage must not be presented as total farm damage."},
    ]
    write_csv("VALUE_CROSSWALK_wildfire_wind__model_v1_0__docs_r1.csv", value_fields, values)


def build_workbook(art: dict, kat: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    navy, teal, pale, amber = "17243B", "0F766E", "E7F3F1", "FCE8C5"

    def make_sheet(name: str, title: str, headers: list[str], rows: list[list[object]], widths: list[int]) -> None:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        ws.cell(1, 1, title)
        ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=14)
        ws.cell(1, 1).fill = PatternFill("solid", fgColor=navy)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(headers)))
        ws.cell(2, 1, "Noncanonical Tier-4 partial screening proposal • independent review and promotion blocked")
        ws.cell(2, 1).fill = PatternFill("solid", fgColor=pale)
        for col, head in enumerate(headers, 1):
            c = ws.cell(4, col, head); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=teal)
        for r, row in enumerate(rows, 5):
            for c, value in enumerate(row, 1):
                cell = ws.cell(r, c, value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i, width in enumerate(widths, 1): ws.column_dimensions[chr(64 + i)].width = width

    make_sheet("README", "Wildfire × wind — model-v1 partial electrical screening", ["Identity", "Value", "Meaning"], [
        ["Damage code", art["damage_code_id"], "two named unit DRs only"],
        ["Model", MODEL, "behavior-changing proposal"],
        ["Canonical", False, "do not cut over Hazard"],
        ["Supported units", ", ".join(CURVES), "not a whole-farm curve"],
        ["Axis", "exact FSim class state 0..6", "screening ordinal, not local heat flux"],
        ["Ordinate grade", "Tier 4", "owner-authorized assumptions"],
        ["Scenario dollars", "withheld", "site values and exposure split absent"],
        ["Annual/tail metrics", "withheld", "downstream and prerequisites unmet"],
    ], [28, 52, 70])
    curve_rows = []
    for unit, spec in CURVES.items():
        for state, dr in spec["points"]: curve_rows.append([unit, spec["curve_id"], state, dr, "T4", spec["profile_note"]])
    make_sheet("Curves", "Exact categorical screening state tables", ["Failure unit", "Curve ID", "Class state", "DR", "Tier", "Interpretation"], curve_rows, [34, 44, 14, 12, 12, 80])
    make_sheet("KATs", "Known-answer tests", ["Test ID", "Failure unit", "State", "Expected DR", "Status"], [[t["test_id"], t["input"]["failure_unit_id"], t["input"]["conditional_flame_length_class_state"], t["expected"]["failure_unit_damage_ratio"], t["expected"]["status"]] for t in kat["formula_known_answer_tests"]], [48, 34, 14, 16, 24])

    for sheet, filename in [
        ("Sources", "SOURCE_REGISTER_wildfire_wind__model_v1_0__docs_r1.csv"),
        ("Claims", "CLAIM_PARAMETER_REGISTER_wildfire_wind__model_v1_0__docs_r1.csv"),
        ("Tiers", "PARAMETER_TIER_TABLE_wildfire_wind__model_v1_0__docs_r1.csv"),
        ("Values", "VALUE_CROSSWALK_wildfire_wind__model_v1_0__docs_r1.csv"),
    ]:
        with (OUT / filename).open(newline="") as fh:
            rows = list(csv.reader(fh))
        make_sheet(sheet, filename.removesuffix(".csv"), rows[0], rows[1:], [24] * len(rows[0]))
    ws = wb["README"]
    ws.merge_cells("A16:C18"); ws["A16"] = "Guardrail: these curves make two real physical risks visible for screening. They do not prove a wind-farm loss calibration, do not value the farm, and do not turn every unmodeled subsystem into zero."
    ws["A16"].fill = PatternFill("solid", fgColor=amber); ws["A16"].font = Font(bold=True); ws["A16"].alignment = Alignment(wrap_text=True, vertical="center")
    wb.save(OUT / "damage_curve_records_wildfire_wind__model_v1_0__docs_r1.xlsx")


def main() -> None:
    build_registers()
    cap = capability()
    art = artifact(cap)
    kat = kats()
    dump_json("wildfire_wind__model_v1_0__docs_r1__capability.json", cap)
    dump_json("wildfire_wind__model_v1_0__docs_r1__curve_artifact.json", art)
    dump_json("known_answer_tests_wildfire_wind__model_v1_0__docs_r1.json", kat)
    build_workbook(art, kat)
    print("Built wildfire_wind model v1.0/docs r1 partial screening package")


if __name__ == "__main__":
    main()
