#!/usr/bin/env python3
"""Build the TC-wind × canonical Wind Farm tower-only v1.2 proposal.

The proposal does not tune the Jaimes curve.  It corrects the failure-unit and
value binding: the unchanged 3.3 MW / 100 m tower-state curve is applied only
to the canonical asset's tower share (0.16 of project TIV).  The other 0.84 is
explicitly withheld.
"""

from __future__ import annotations

import csv
import json
from copy import deepcopy
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import tropical_cyclone_wind_wind_curve_eval as evaluator


ROOT = Path(__file__).resolve().parents[2]
CELL = ROOT / "docs/cells/tropical_cyclone_wind_wind"
CURRENT = CELL / "current"
PROPOSED = CELL / "proposed"
BASELINE = CELL / "archive/model_v1_0__docs_r1"
STEM = "model_v1_2__docs_r2"
PREFIX = f"tropical_cyclone_wind_wind__{STEM}"
V1_1_ARCHIVE = CELL / "archive/model_v1_1__docs_r1"
CURRENT_ARTIFACT = V1_1_ARCHIVE / "tropical_cyclone_wind_wind__model_v1_1__docs_r1__curve_artifact.json"
CURRENT_CAPABILITY = V1_1_ARCHIVE / "tropical_cyclone_wind_wind__model_v1_1__docs_r1__capability.json"
CURRENT_KATS = V1_1_ARCHIVE / "known_answer_tests_tropical_cyclone_wind_wind__model_v1_1__docs_r1.json"
BASELINE_ARTIFACT = BASELINE / "tropical_cyclone_wind_wind__model_v1_0__docs_r1__curve_artifact.json"

ARTIFACT = PROPOSED / f"{PREFIX}__curve_artifact.json"
CAPABILITY = PROPOSED / f"{PREFIX}__capability.json"
KATS = PROPOSED / f"known_answer_tests_tropical_cyclone_wind_wind__{STEM}.json"
SOURCES = PROPOSED / f"SOURCE_REGISTER_tropical_cyclone_wind_wind__{STEM}.csv"
CLAIMS = PROPOSED / f"CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_wind__{STEM}.csv"
PARAMETERS = PROPOSED / f"PARAMETER_TIER_TABLE_tropical_cyclone_wind_wind__{STEM}.csv"
VALUES = PROPOSED / f"VALUE_CROSSWALK_tropical_cyclone_wind_wind__{STEM}.csv"
OLD_NEW = PROPOSED / f"OLD_VS_NEW_COMPARISON_tropical_cyclone_wind_wind__{STEM}.csv"
WORKBOOK = PROPOSED / f"damage_curve_records_tropical_cyclone_wind_wind__{STEM}.xlsx"
README = PROPOSED / f"README_tropical_cyclone_wind_wind__{STEM}.md"
CLASSIFICATION = PROPOSED / f"CHANGE_CLASSIFICATION_tropical_cyclone_wind_wind__{STEM}.md"
DECISIONS = PROPOSED / f"DECISION_LOG_tropical_cyclone_wind_wind__{STEM}.md"
DOSSIER = PROPOSED / f"tropical_cyclone_wind_wind_curve_derivation_dossier__{STEM}.md"
GATES = PROPOSED / f"PROMOTION_GATE_MATRIX_tropical_cyclone_wind_wind__{STEM}.md"
VALIDATION = PROPOSED / f"VALIDATION_REPORT_tropical_cyclone_wind_wind__{STEM}.md"


def load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=False) + "\n")


def write_csv(path: Path, header: list[str], rows: Iterable[Iterable[Any]]) -> None:
    with path.open("w", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(header)
        writer.writerows(rows)


def source_record(pathway: Mapping[str, Any]) -> Mapping[str, Any]:
    return next(
        record
        for record in pathway["curve_records"]
        if record["curve_id"] == "TCWW_JAIMES_3P3MW_100M_SCREENING"
    )


def tower_proxy_record(source: Mapping[str, Any]) -> dict[str, Any]:
    record = deepcopy(source)
    record.update(
        {
            "curve_id": "TCWW_JAIMES_3P3MW_AS_CANONICAL_5MW_TOWER_PROXY_V1",
            "failure_unit_id": evaluator.TOWER_PROXY_FAILURE_UNIT,
            "selector_match": {
                "turbine_archetype_id": evaluator.TOWER_PROXY_ARCHETYPE_ID,
                "rated_power_mw": 5,
                "hub_height_m": 100,
                "rotor_diameter_m": 114,
            },
        }
    )
    return record


def build_capability(current: Mapping[str, Any], baseline: Mapping[str, Any]) -> dict[str, Any]:
    capability = deepcopy(current)
    capability["canonical_runtime_artifact"] = False
    pathway = capability["pathway_capabilities"][0]
    baseline_pathway = baseline["capability_declaration"]["pathway_capabilities"][0]
    withheld = deepcopy(baseline_pathway["withheld_failure_units"])
    pathway.update(
        {
            "scenario_loss_given_value_basis": "supported_with_explicit_failure_unit_value_and_exposure_basis",
            "conditions": [
                "exact source-native Jaimes selectors retain model-v1.0 behavior",
                "canonical 5 MW use requires exact tower-proxy policy, asset-profile and value-basis IDs",
                "the proxy evaluates the unchanged Jaimes 3.3 MW / 100 m parameters",
                "the failure unit remains WT_JAIMES_TURBINE_TOWER_EXPOSURE_UNIT",
                "only tower value is covered: 0.16 of project TIV; 0.84 is withheld, not zero",
                "the named proxy alone completes the 90–108 and above-252 km/h review boundaries",
            ],
            "limitation_flags": [
                "OWNER_APPROVED_SCREENING_PROXY",
                "REQUESTED_5MW_EVALUATED_WITH_3P3MW_SOURCE_CURVE",
                "NO_CAPACITY_RATIO_SCALING",
                "NOT_TARGET_MATCHED_5MW_EVIDENCE",
                "NOT_FIELD_CALIBRATED",
                "NOT_CLAIMS_CALIBRATED",
                "PARTIAL_TOWER_VALUE_COVERAGE_16PCT",
                "UNCOVERED_PROJECT_VALUE_84PCT_WITHHELD_NOT_ZERO",
                "CURVE_INTRINSIC_SPREAD_NOT_CARRIED",
                "NO_NHC_OR_HUB_HEIGHT_BRIDGE",
                "PROXY_SCREENING_TRANSITION_BAND_ZERO_RULE",
                "PROXY_SCREENING_ABOVE_RANGE_MAX_DR_CAP",
            ],
            "withheld_failure_units": withheld,
        }
    )
    capability["consumer_annual_metrics"] = {
        "computation_owner": "downstream_consumer",
        "status_before_promotion": "withheld_noncanonical_proposal",
        "status_after_promotion": "consumer_computable_from_validated_frequency_intensity_coupling_value_and_cap_model",
        "prerequisites": [
            "released exact cell/model/docs/schema/SHA pin",
            "validated Hurricane event frequency and intensity coupling",
            "exact canonical asset, tower-proxy policy and tower-value IDs",
            "0.16 occurrence and annual caps",
            "explicit disclosure that 0.84 of project TIV is withheld",
        ],
        "limitation_flags": [
            "CURVE_INTRINSIC_SPREAD_NOT_CARRIED",
            "PARTIAL_TOWER_VALUE_COVERAGE_16PCT",
            "ANNUAL_AND_TAIL_METRICS_CONDITIONAL_ON_SCREENING_PROXY",
        ],
    }
    capability["cap_binding"] = {
        "policy": "consumer_enforced_fail_closed",
        "enforcement_owner": "downstream_consumer",
        "checks_required": [
            "exact tower-proxy policy, asset profile, value basis, model/docs/schema/SHA pin",
            "covered loss <= 0.16 * project TIV for every occurrence",
            "uncovered 0.84 remains withheld and is never emitted as zero loss",
            "no 5/3.3 multiplier and no rotor or nacelle value binding",
            "annual/TIV cap is applied inside the annual calculation",
        ],
        "action_if_fail": "withhold scenario loss and all annual or tail metrics",
    }
    capability["promotion_gate"] = {
        "status": "ready_for_review",
        "required_before_canonical_use": [
            "source-native reproduction",
            "tower proxy, negative-contract, value-share and cap known answers",
            "old-v-new consumer comparison on the governed Hurricane population",
            "owner review of partial-value reporting boundary",
        ],
    }
    return capability


def build_artifact() -> tuple[dict[str, Any], dict[str, Any]]:
    current = load(CURRENT_ARTIFACT)
    baseline = load(BASELINE_ARTIFACT)
    artifact = deepcopy(current)
    capability = build_capability(load(CURRENT_CAPABILITY), baseline)
    artifact.update(
        {
            "damage_code_id": "TROPICAL_CYCLONE_WIND_WIND_JAIMES_TOWER_SCREENING_V1_2",
            "semantic_damage_model_version": "model v1.2",
            "documentation_revision": "docs r2",
            "lifecycle_state": "proposed_v1_2",
            "promotion_status": "proposed",
            "review_status": "owner_directed_tower_scope_correction_pending_consumer_validation",
            "model_grade": "screening_owner_approved_target_mismatch_tower_only_proxy",
            "canonical_runtime_artifact": False,
            "source_dossier": DOSSIER.relative_to(ROOT).as_posix(),
            "source_workbook": WORKBOOK.relative_to(ROOT).as_posix(),
            "known_answer_tests": KATS.relative_to(ROOT).as_posix(),
            "source_register": SOURCES.relative_to(ROOT).as_posix(),
            "claim_parameter_register": CLAIMS.relative_to(ROOT).as_posix(),
            "value_crosswalk": VALUES.relative_to(ROOT).as_posix(),
        }
    )
    pathway = artifact["pathways"][0]
    source = source_record(pathway)
    pathway["curve_records"] = [
        record
        for record in pathway["curve_records"]
        if record["selector_match"]["turbine_archetype_id"] != evaluator.PROXY_ARCHETYPE_ID
    ]
    pathway["curve_records"].append(tower_proxy_record(source))
    selector = next(item for item in pathway["selector_logic"] if item["field"] == "turbine_archetype_id")
    selector["allowed"] = [
        value for value in selector["allowed"] if value != evaluator.PROXY_ARCHETYPE_ID
    ] + [evaluator.TOWER_PROXY_ARCHETYPE_ID]
    selector.update(
        {
            "routing": "exact selector match; the canonical 5 MW route also requires exact tower-proxy IDs",
            "nearest_neighbor": "prohibited; the one named bridge is explicit and never inferred",
            "modern_fleet_transfer": "one explicit canonical 5 MW tower-only screening bridge",
        }
    )
    pathway["exposure_contract"].update(
        {
            "whole_farm_default": "prohibited; tower-only use requires the exact proxy contract",
            "value_rule": "Proxy DR applies only to tower = 0.16 of project TIV; the other 0.84 is withheld, not zero.",
        }
    )
    baseline_units = {item["id"]: item for item in baseline["failure_units"]}
    for index, unit in enumerate(artifact["failure_units"]):
        if unit["id"] == evaluator.PROXY_FAILURE_UNIT:
            artifact["failure_units"][index] = deepcopy(baseline_units[evaluator.PROXY_FAILURE_UNIT])
        elif unit["id"] == evaluator.TOWER_PROXY_FAILURE_UNIT:
            unit["proxy_value_binding"] = {
                "covered_value_basis_id": evaluator.TOWER_PROXY_VALUE_BASIS_ID,
                "covered_value_share_of_project_tiv": evaluator.TOWER_PROXY_COVERED_VALUE_SHARE,
                "covered_subsystems": ["tower"],
            }
    baseline_coverage = {
        item["failure_unit_id"]: item for item in baseline["pathways"][0]["failure_unit_coverage"]
    }
    pathway["failure_unit_coverage"] = [
        deepcopy(baseline_coverage[item["failure_unit_id"]])
        if item["failure_unit_id"] == evaluator.PROXY_FAILURE_UNIT
        else item
        for item in pathway["failure_unit_coverage"]
    ]
    artifact["derivation_rationale"].update(
        {
            "v1_2_decision": "replace the 0.63 equipment-assembly value bridge with a 0.16 tower-only bridge",
            "why_minor_version": "curve parameters and source-native behavior are unchanged; supported target value binding changes",
            "why_not_curve_tuning": "the diagnosed defect is failure-unit/value scope, not the fitted Jaimes equation",
            "owner_approved_proxy_contract": {
                "status": "owner_approved_screening_proxy",
                "proxy_policy_id": evaluator.TOWER_PROXY_POLICY_ID,
                "canonical_asset_profile_id": evaluator.TOWER_PROXY_ASSET_PROFILE_ID,
                "source_curve_id": source["curve_id"],
                "source_evidence_identity": "Jaimes 3.3 MW / 100 m / 114 m tower-state record",
                "target_identity": "canonical 5 MW / 100 m turbine tower",
                "numeric_rule": "evaluate unchanged source parameters; no 5/3.3 scaling",
                "screening_completion_rule": {
                    "transition_band_kmh": [90, 108],
                    "transition_treatment": "zero_with_explicit_flag",
                    "source_ceiling_kmh": 252,
                    "above_ceiling_treatment": "cap_at_max_dr_with_explicit_flag",
                },
                "failure_unit_id": evaluator.TOWER_PROXY_FAILURE_UNIT,
                "covered_value_basis_id": evaluator.TOWER_PROXY_VALUE_BASIS_ID,
                "covered_value_share_of_project_tiv": 0.16,
                "uncovered_value_share_of_project_tiv": 0.84,
            },
        }
    )
    artifact["evaluation_contract"].update(
        {
            "selector_behavior": "exact source selectors plus one exact owner-approved 5 MW tower bridge",
            "proxy_screening_completion": {
                "applies_only_when_proxy_policy_id": evaluator.TOWER_PROXY_POLICY_ID,
                "transition_band_kmh": [90, 108],
                "transition_treatment": "return DR=0 with SCREENING_TRANSITION_BAND_ASSIGNED_ZERO",
                "source_ceiling_kmh": 252,
                "above_ceiling_treatment": "return max_dr with SCREENING_ABOVE_SOURCE_RANGE_CAPPED_AT_MAX_DR",
                "source_native_selectors_unchanged": True,
            },
        }
    )
    artifact["value_linkage"].update(
        {
            "scenario_loss_status": "consumer_computable_only_for_explicit_tower_proxy_value",
            "owner_approved_proxy_value_basis": {
                "id": evaluator.TOWER_PROXY_VALUE_BASIS_ID,
                "canonical_asset_profile_id": evaluator.TOWER_PROXY_ASSET_PROFILE_ID,
                "project_tiv_usd_at_activation": 140000000,
                "covered_subsystem_shares": {"tower": 0.16},
                "covered_value_share": 0.16,
                "covered_value_usd_at_activation": 22400000,
                "uncovered_value_share": 0.84,
                "uncovered_value_usd_at_activation": 117600000,
                "uncovered_subsystems": ["rotor", "nacelle", "foundation", "substation", "electrical", "civil"],
                "uncovered_treatment": "withheld_not_zero",
            },
            "approval_required": "exact Hazard consumer pin, cap validation and owner review before promotion",
        }
    )
    artifact["emit_contract"].update(
        {
            "supported_output": "conditional scalar mean DR for exact source records or the named canonical-5-MW tower proxy",
            "scenario_loss": "consumer-computable only with the exact 0.16 tower-value basis and caps",
            "annual_and_tail_metrics": "downstream consumer only after frequency/coupling/cap validation",
        }
    )
    artifact["capability_declaration"] = capability
    return artifact, capability


def build_kats(artifact: Mapping[str, Any]) -> dict[str, Any]:
    current = load(CURRENT_KATS)
    base = {
        "pathway_id": evaluator.SUPPORTED_PATHWAY,
        "failure_unit_id": evaluator.TOWER_PROXY_FAILURE_UNIT,
        "turbine_archetype_id": evaluator.TOWER_PROXY_ARCHETYPE_ID,
        "source_model_assumption_set_id": evaluator.SUPPORTED_ASSUMPTION_SET,
        "proxy_policy_id": evaluator.TOWER_PROXY_POLICY_ID,
        "canonical_asset_profile_id": evaluator.TOWER_PROXY_ASSET_PROFILE_ID,
        "covered_value_basis_id": evaluator.TOWER_PROXY_VALUE_BASIS_ID,
    }
    record = next(
        item for item in artifact["pathways"][0]["curve_records"]
        if item["selector_match"]["turbine_archetype_id"] == evaluator.TOWER_PROXY_ARCHETYPE_ID
    )
    proxy_tests = []
    for speed in (90.0, 100.0, 108.0, 160.0, 163.3, 180.0, 200.0, 252.0, 300.0):
        expected_dr, boundary_flag = evaluator.evaluate_proxy_screening_completion(record, speed)
        proxy_tests.append(
            {
                "test_id": f"TCWW12_TOWER_PROXY_V{str(speed).replace('.', 'P')}",
                "input": {**base, evaluator.AXIS_FIELD: speed},
                "expected": {
                    "status": "supported",
                    "curve_id": record["curve_id"],
                    "failure_unit_damage_ratio": expected_dr,
                    "boundary_flag": boundary_flag,
                    "covered_value_share_of_project_tiv": 0.16,
                },
            }
        )
    negative = [
        {
            "test_id": "TCWW12_PROXY_OPT_IN_REQUIRED",
            "input": {key: value for key, value in {**base, evaluator.AXIS_FIELD: 180}.items() if key not in {"proxy_policy_id", "canonical_asset_profile_id", "covered_value_basis_id"}},
            "expected_error_code": "OWNER_APPROVED_PROXY_OPT_IN_REQUIRED",
        },
        {
            "test_id": "TCWW12_OLD_63PCT_PROXY_REJECTED",
            "input": {**base, "turbine_archetype_id": evaluator.PROXY_ARCHETYPE_ID, evaluator.AXIS_FIELD: 180},
            "expected_error_code": "TURBINE_ARCHETYPE_UNSUPPORTED",
        },
        {
            "test_id": "TCWW12_OLD_VALUE_BASIS_REJECTED",
            "input": {**base, "covered_value_basis_id": evaluator.PROXY_VALUE_BASIS_ID, evaluator.AXIS_FIELD: 180},
            "expected_error_code": "OWNER_APPROVED_PROXY_IDENTITY_MISMATCH",
        },
        {
            "test_id": "TCWW12_OLD_POLICY_REJECTED",
            "input": {**base, "proxy_policy_id": evaluator.PROXY_POLICY_ID, evaluator.AXIS_FIELD: 180},
            "expected_error_code": "OWNER_APPROVED_PROXY_IDENTITY_MISMATCH",
        },
    ]
    reproduction = deepcopy(current["v1_0_reproduction_tests"])
    return {
        "schema_version": "known_answer_tests.tcww_v1_2.v1",
        "cell_id": artifact["cell_id"],
        "semantic_damage_model_version": "model v1.2",
        "documentation_revision": "docs r2",
        "artifact_schema_version": artifact["schema_version"],
        "absolute_tolerance": 1e-12,
        "v1_0_reproduction_tests": reproduction,
        "proxy_known_answer_tests": proxy_tests,
        "negative_contract_tests": negative,
        "value_crosswalk_tests": [
            {"test_id": "TCWW12_CANONICAL_TOWER_CAP", "project_tiv_usd": 140000000, "covered_value_share": 0.16, "expected_covered_value_usd": 22400000, "expected_uncovered_value_usd": 117600000},
            {"test_id": "TCWW12_UNIT_TIV_CAP", "project_tiv_usd": 1, "covered_value_share": 0.16, "expected_covered_value_usd": 0.16, "expected_uncovered_value_usd": 0.84},
        ],
    }


def write_registers() -> None:
    write_csv(SOURCES, ["source_id", "role", "use", "limits"], [
        ["JAIMES_2020_TC_WIND_TURBINE", "numeric source", "unchanged 3.3 MW/100 m curve and source-native axis", "tower-state model; not target-matched to 5 MW"],
        ["OWNER_SCOPE_CORRECTION_2026_08_14", "governance", "authorizes a tower-only 0.16 value binding", "not empirical curve evidence"],
        ["CONUS_WIND_FARM_REFERENCE_V1", "consumer asset profile", "20×5 MW, 100 MW, $140M; tower share 0.16", "screening value grade"],
        ["HAZARD_HURRICANE_WIND_FARM_EAL_REVIEW_2026_08_14", "consumer measurement", "diagnoses 0.63 failure-unit/value mismatch", "not a financial benchmark"],
    ])
    write_csv(CLAIMS, ["claim_id", "claim", "status", "basis", "revisit_trigger"], [
        ["TCWW12-C01", "The canonical proxy uses the unchanged Jaimes 3.3 MW numerical curve.", "owner-approved screening assumption", "JAIMES_2020 + owner scope decision", "target-matched 5 MW evidence"],
        ["TCWW12-C02", "The proxy remains on the source-defined tower exposure unit.", "hard contract", "Jaimes damage states are tower states", "qualified component-complete model"],
        ["TCWW12-C03", "Only tower value, 0.16 of project TIV, is covered.", "owner-approved value crosswalk", "CONUS_WIND_FARM_REFERENCE_V1", "improved tower valuation"],
        ["TCWW12-C04", "The other 0.84 is withheld, not zero.", "hard reporting contract", "partial coverage discipline", "additional governed curves"],
        ["TCWW12-C05", "The prior 0.63 equipment-assembly route is rejected.", "corrective contract", "failure-unit/value review", "never; superseded route"],
    ])
    write_csv(PARAMETERS, ["parameter", "value", "tier", "meaning"], [
        ["V_zero_kmh", 90, "source-derived", "unchanged Jaimes 3.3 MW record"],
        ["delta_V50_kmh", 73.3, "source-derived", "unchanged Jaimes 3.3 MW record"],
        ["rho", 4.99, "source-derived", "unchanged Jaimes 3.3 MW record"],
        ["target_rated_power_mw", 5, "owner-approved identity", "no numeric scaling effect"],
        ["covered_value_share", 0.16, "owner-approved value crosswalk", "tower only"],
    ])
    write_csv(VALUES, ["value_basis_id", "subsystem", "share_of_project_tiv", "treatment", "activation_value_usd", "notes"], [
        [evaluator.TOWER_PROXY_VALUE_BASIS_ID, "tower", 0.16, "covered_screening_proxy", 22400000, "Jaimes tower-state DR applies"],
        [evaluator.TOWER_PROXY_VALUE_BASIS_ID, "covered_total", 0.16, "aggregate_cap", 22400000, "maximum covered occurrence loss"],
        [evaluator.TOWER_PROXY_VALUE_BASIS_ID, "all_non_tower_value", 0.84, "withheld_not_zero", 117600000, "outside this model"],
    ])
    write_csv(OLD_NEW, ["dimension", "model_v1_1", "model_v1_2", "expected_change"], [
        ["Jaimes curve parameters", "3.3 MW source values", "identical", "none"],
        ["source-native selectors", "supported", "identical", "none"],
        ["canonical failure unit", "equipment assembly", "Jaimes tower exposure unit", "corrected"],
        ["covered value", "rotor+nacelle+tower = 0.63", "tower = 0.16", "reduced to evidence-aligned scope"],
        ["uncovered value", "0.37 withheld", "0.84 withheld", "explicit"],
        ["old proxy identities", "accepted", "rejected", "fail-closed migration"],
        ["governed Hurricane analytical max EAL", "about 7.78% TIV/year", "about 1.98% TIV/year", "expected consumer consequence; not a benchmark claim"],
    ])


def build_workbook(artifact: Mapping[str, Any], kats: Mapping[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    def sheet(name: str, title: str, header: list[str], rows: list[list[Any]]) -> None:
        ws = workbook.create_sheet(name)
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(header)))
        ws["A1"].fill = fill
        ws["A1"].font = font
        ws.append(header)
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A3"
        for index, column in enumerate(ws.columns, start=1):
            ws.column_dimensions[get_column_letter(index)].width = min(70, max(12, max(len(str(cell.value or "")) for cell in column) + 2))

    sheet("README", "TC-wind × Wind Farm model v1.2 — tower-only proxy", ["Field", "Value"], [
        ["Source curve", "Jaimes 3.3 MW / 100 m, unchanged"],
        ["Failure unit", evaluator.TOWER_PROXY_FAILURE_UNIT],
        ["Covered value", "tower = 0.16 of project TIV"],
        ["Uncovered value", "0.84 withheld, not zero"],
    ])
    sheet("Curve records", "Source records plus one exact tower bridge", ["Curve ID", "Failure unit", "Archetype", "V0", "delta V50", "rho"], [
        [record["curve_id"], record["failure_unit_id"], record["selector_match"]["turbine_archetype_id"], record["parameters"]["V_zero_kmh"], record["parameters"]["delta_V50_kmh"], record["parameters"]["rho"]]
        for record in artifact["pathways"][0]["curve_records"]
    ])
    sheet("Proxy KATs", "Tower-proxy known answers", ["Test", "Speed km/h", "DR", "Covered share"], [
        [test["test_id"], test["input"][evaluator.AXIS_FIELD], test["expected"]["failure_unit_damage_ratio"], 0.16]
        for test in kats["proxy_known_answer_tests"]
    ])
    sheet("Value crosswalk", "Canonical project value scope", ["Subsystem", "Share", "Treatment", "USD at $140M TIV"], [
        ["tower", 0.16, "covered", 22400000], ["all non-tower value", 0.84, "withheld", 117600000]
    ])
    workbook.save(WORKBOOK)


def write_docs() -> None:
    README.write_text(f"""# TC-wind × Wind Farm model v1.2/docs r2 proposal

This proposal corrects the prior value scope without changing the Jaimes curve.

```text
Jaimes 3.3 MW / 100 m tower-state curve (unchanged)
                         ↓
canonical 5 MW tower-only screening proxy
                         ↓
tower = 0.16 of project TIV covered
all other value = 0.84 withheld, not zero
```

The old 0.63 rotor+nacelle+tower route is not retained. Requests must carry the new exact archetype,
proxy-policy and tower-value-basis IDs. This is still target-mismatched screening evidence, not a
component-complete or financially calibrated wind-farm model.

## Package

- [{ARTIFACT.name}]({ARTIFACT.name})
- [{CAPABILITY.name}]({CAPABILITY.name})
- [{KATS.name}]({KATS.name})
- [{VALUES.name}]({VALUES.name})
- [{OLD_NEW.name}]({OLD_NEW.name})
- [{WORKBOOK.name}]({WORKBOOK.name})
- [{DOSSIER.name}]({DOSSIER.name})
- [{GATES.name}]({GATES.name})
- [{VALIDATION.name}]({VALIDATION.name})
""")
    CLASSIFICATION.write_text("""# Change classification — model v1.2/docs r2

```yaml
change_class: MODEL_BEHAVIOR_CHANGE
cell_id: tropical_cyclone_wind_wind
current: model v1.1/docs r1
proposed: model v1.2/docs r2
outputs_can_change_for_same_inputs: true
schema_change: false
```

The numerical curve is unchanged. The supported failure-unit/value binding changes from a 0.63
equipment-assembly proxy to a 0.16 tower-only proxy, so a model minor and documentation revision are required.
""")
    DECISIONS.write_text("""# Decision log — model v1.2/docs r2

## TCWW12-D01 · Correct the failure unit before tuning the curve

The Jaimes ordinate is built from tower damage states. Keep its numerical parameters unchanged and bind the
canonical proxy to the source-defined tower exposure unit.

## TCWW12-D02 · Cover only tower value

Apply the conditional damage ratio to 0.16 of project TIV. Explicitly withhold the other 0.84.

## TCWW12-D03 · Fail closed on the old route

The old 0.63 value-basis and equipment-assembly identities are rejected. They are preserved only in the
model-v1.1 archive for reproduction.
""")
    DOSSIER.write_text("""# Derivation dossier — TC-wind × Wind Farm model v1.2/docs r2

## What changed

The curve did not change. The denominator did. Model v1.1 applied a tower-state expected-damage curve to
rotor+nacelle+tower value (0.63 of project TIV). The Hurricane consumer review found that this produced a
maximum analytical EAL near 7.78% of project TIV/year and, more importantly, exceeded the failure scope
supported by the Jaimes evidence.

## Corrected bridge

Model v1.2 keeps `V_zero=90 km/h`, `delta_V50=73.3 km/h`, `rho=4.99`, and `max_dr=1`. It uses the same exact
canonical 5 MW target mismatch disclosure but binds the result only to the source-defined tower exposure
unit and the canonical tower share of 0.16. The other 0.84 remains unknown/unsupported.

On the governed Hurricane population, changing only that value scope reduces the analytical maximum from
about 7.78% to about 1.98% of project TIV/year. This is a consequence check, not external financial-range
validation. The model remains screening-grade until target-matched evidence and a governed financial range
exist.
""")
    GATES.write_text("""# Promotion gates — model v1.2/docs r2

| Gate | Required result |
|---|---|
| source-native reproduction | model-v1.0 source selectors unchanged |
| numerical proxy | unchanged 3.3 MW curve KATs pass |
| identity | old 0.63 route and wrong units fail closed |
| value | 0.16 covered; 0.84 withheld; occurrence/annual cap enforced |
| consumer | governed Hurricane M2–M4 rebuild and old-v-new comparison pass |
| reporting | tower-only, target-mismatch and unvalidated-range caveats visible |
""")
    VALIDATION.write_text("""# Validation report — model v1.2/docs r2

Status: **pending independent validator and Hazard consumer rebuild**.

This file is replaced with measured results before promotion.
""")


def main() -> None:
    PROPOSED.mkdir(parents=True, exist_ok=True)
    artifact, capability = build_artifact()
    kats = build_kats(artifact)
    write_registers()
    build_workbook(artifact, kats)
    write_docs()
    write_json(CAPABILITY, capability)
    write_json(ARTIFACT, artifact)
    write_json(KATS, kats)
    print(f"built={ARTIFACT.relative_to(ROOT)}")
    print("curve_parameters=unchanged")
    print("covered_value_share=0.16")
    print("uncovered_value_share=0.84")


if __name__ == "__main__":
    main()
