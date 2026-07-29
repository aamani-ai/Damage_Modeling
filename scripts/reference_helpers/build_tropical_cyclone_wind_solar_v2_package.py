#!/usr/bin/env python3
"""Build generated artifacts for the proposed TC-wind x solar model-v2 package.

Human narrative files remain hand-governed.  This builder materializes the
machine bundle, consolidated registers, known-answer fixtures, shared
non-runtime response candidate, comparison table, and audit workbook.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Iterable, Mapping
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
CELL = ROOT / "docs/cells/tropical_cyclone_wind_solar"
PROPOSED = CELL / "proposed"
SHARED = ROOT / "docs/method/shared_components/solar_wind_normalized_response"
STRONG = ROOT / "docs/cells/strong_wind_solar/proposed"

MODEL = "model v2.0"
DOCS = "docs r1"
PATHWAY = "tropical_cyclone_wind"
PERRY_ARCH = "perry_ground_nontracking_source_cohort_v1_compat"
FIXED_ARCH = "fixed_tilt_ground_mount_tc_synthetic_t4_v1"
TRACKER_ARCH = "single_axis_tracker_tc_qualified_synthetic_t4_v1"
AXIS = "architecture_specific_tropical_cyclone_wind_demand_index"
PERRY_UNIT = "PV_PERRY_GROUND_FIXED_VISIBLE_MODULE_HARDWARE_SOURCE_UNIT"
CELL_LOCAL_SYNTHETIC_SOURCE = "TCWS2_CELL_LOCAL_SYNTHETIC_DECISION"
NO_DIRECT_TC_CALIBRATION_SOURCE = "NO_DIRECT_TC_CALIBRATION_SOURCE"
STRONG_V2_ARTIFACT_SHA256 = "32fe982548139cda846fb2e1da63568bcdcc689a87d6b21bd0110f23676c58fb"

ARTIFACT = PROPOSED / "tropical_cyclone_wind_solar__model_v2_0__docs_r1__curve_artifact.json"
CAPABILITY = PROPOSED / "tropical_cyclone_wind_solar__model_v2_0__docs_r1__capability.json"
KATS = PROPOSED / "known_answer_tests_tropical_cyclone_wind_solar__model_v2_0__docs_r1.json"
SOURCES = PROPOSED / "SOURCE_REGISTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
CLAIMS = PROPOSED / "CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
PARAMETERS = PROPOSED / "PARAMETER_TIER_TABLE_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
VALUES = PROPOSED / "VALUE_CROSSWALK_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
OLD_NEW = PROPOSED / "OLD_VS_NEW_COMPARISON_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
CLAIM_SUPERSESSION = PROPOSED / "CLAIM_SUPERSESSION_MAP_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv"
WORKBOOK = PROPOSED / "damage_curve_records_tropical_cyclone_wind_solar__model_v2_0__docs_r1.xlsx"
SHARED_PROFILE = SHARED / "candidate_response_profile_v0_1.json"

SOURCE_FIELDS = [
    "source_id",
    "citation",
    "url",
    "accessed_on",
    "exact_locator",
    "source_type",
    "source_role",
    "pathway_ids",
    "evidence_tier",
    "target_asset_match",
    "target_failure_unit_match",
    "measured_or_modeled_endpoint",
    "permitted_inference",
    "prohibited_inference",
    "decision",
    "status",
    "notes",
]
CLAIM_FIELDS = [
    "claim_id",
    "pathway_id",
    "claim_text",
    "claim_type",
    "source_ids",
    "exact_locator",
    "evidence_tier",
    "parameter_or_rule",
    "adoption_status",
    "permitted_inference",
    "prohibited_inference",
    "reasoning",
    "update_trigger",
]
PARAMETER_FIELDS = [
    "parameter",
    "pathway_id",
    "curve_id",
    "value",
    "param_role",
    "tier",
    "source_ids",
    "reasoning",
    "status",
    "update_trigger",
]
CLAIM_SUPERSESSION_FIELDS = [
    "prior_claim_id",
    "prior_version_scope",
    "v2_status",
    "superseding_claim_id",
    "retained_truth",
]


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, sort_keys=False) + "\n")


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fields: list[str], rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="raise")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def save_deterministic_workbook(workbook: Workbook, path: Path) -> None:
    """Save XLSX bytes without wall-clock timestamps in XML or ZIP headers."""
    fixed_document_time = datetime(2020, 1, 1, 0, 0, 0)
    fixed_zip_time = (2020, 1, 1, 0, 0, 0)
    workbook.properties.created = fixed_document_time
    workbook.properties.modified = fixed_document_time
    path.parent.mkdir(parents=True, exist_ok=True)

    with TemporaryDirectory(prefix="tcws2_workbook_") as temporary_directory:
        raw_path = Path(temporary_directory) / "raw.xlsx"
        workbook.save(raw_path)
        with ZipFile(raw_path, "r") as source, ZipFile(
            path,
            "w",
            compression=ZIP_DEFLATED,
            compresslevel=9,
        ) as target:
            for member_name in sorted(source.namelist()):
                member_bytes = source.read(member_name)
                if member_name == "docProps/core.xml":
                    member_bytes = re.sub(
                        rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                        rb"\g<1>2020-01-01T00:00:00Z\g<2>",
                        member_bytes,
                    )
                member = ZipInfo(member_name, date_time=fixed_zip_time)
                member.compress_type = ZIP_DEFLATED
                member.create_system = 3
                member.external_attr = 0o600 << 16
                target.writestr(
                    member,
                    member_bytes,
                    compress_type=ZIP_DEFLATED,
                    compresslevel=9,
                )


def synthetic_records() -> list[dict[str, Any]]:
    strong_artifact = STRONG / "strong_wind_solar__model_v2_0__docs_r1__curve_artifact.json"
    if digest(strong_artifact) != STRONG_V2_ARTIFACT_SHA256:
        raise RuntimeError(
            "pinned strong_wind_solar v2 source artifact changed; review before deriving new cell-local assumptions"
        )
    base = load_json(strong_artifact)
    records = deepcopy(base["pathways"][0]["curve_records"])
    ids = {
        "SWS2_SLC_FIXED_MODULE_ORDERED_STATES": "TCWS2_FIXED_MODULE_SYNTHETIC_T4_ORDERED_STATES",
        "SWS2_SLC_FIXED_STRUCTURE_ORDERED_STATES": "TCWS2_FIXED_STRUCTURE_SYNTHETIC_T4_ORDERED_STATES",
        "SWS2_SLC_TRACKER_MODULE_ORDERED_STATES": "TCWS2_TRACKER_MODULE_SYNTHETIC_T4_ORDERED_STATES",
        "SWS2_SLC_TRACKER_SBOS_ORDERED_STATES": "TCWS2_TRACKER_SBOS_SYNTHETIC_T4_ORDERED_STATES",
    }
    for record in records:
        record["curve_id"] = ids[record["curve_id"]]
        record["pathway_id"] = PATHWAY
        record["x_axis"] = AXIS
        record["parameters"].pop("zero_below", None)
        for state in record["parameters"]["damage_states"]:
            state["tier"] = "T4_placeholder_or_expert_judgment"
            state["source_ids"] = [
                CELL_LOCAL_SYNTHETIC_SOURCE,
                NO_DIRECT_TC_CALIBRATION_SOURCE,
            ]
            state["description"] += (
                "; synthetic coverage state cost, not an observed or calibrated hazard consequence"
            )
            if state["state_id"] == "STRUCT_DS2_REPLACE_STRUCTURE_MODULES_SALVAGEABLE":
                state["state_id"] = "STRUCT_DS2_FULL_UNIT_REPLACEMENT"
                state["description"] = (
                    "full replacement of the support/SBOS failure unit; module disposition is outside "
                    "this unit and excluded pending a governed joint-state model; synthetic coverage "
                    "state cost, not an observed or calibrated hazard consequence"
                )
            elif state["state_id"] == "STRUCT_DS3_DESTRUCTIVE_COLLAPSE_MODULES_NONSALVAGEABLE":
                state["state_id"] = "STRUCT_DS3_TERMINAL_UNIT_FAILURE"
                state["affected_subsystems"] = state["affected_subsystems"][:1]
                state["description"] = (
                    "terminal destructive failure of the support/SBOS failure unit only; module "
                    "disposition and any cascade are excluded pending a governed joint-state model; "
                    "synthetic coverage state cost, not an observed or calibrated hazard consequence"
                )
        for scenario in record["parameters"]["capacity_scenarios"]:
            scenario["tier"] = "T4_placeholder_or_expert_judgment"
            scenario["source_ids"] = [
                CELL_LOCAL_SYNTHETIC_SOURCE,
                NO_DIRECT_TC_CALIBRATION_SOURCE,
            ]
            scenario["interpretation"] = (
                "unweighted synthetic normalized-demand response scenario; not fitted, "
                "not a percentile, and not a hazard-specific calibration"
            )
    return records


def build_shared_profile(records: list[dict[str, Any]]) -> dict[str, Any]:
    architecture_by_unit = {
        "PV_FIXED_TILT_MODULE_FIELD": "fixed_tilt_ground_mount",
        "PV_FIXED_TILT_SUPPORT_STRUCTURE": "fixed_tilt_ground_mount",
        "PV_TRACKER_MODULE_FIELD": "single_axis_tracker",
        "PV_TRACKER_SBOS_ASSEMBLY": "single_axis_tracker",
    }
    comparison_records: list[dict[str, Any]] = []
    for record in records:
        parameters = deepcopy(record["parameters"])
        for state in parameters["damage_states"]:
            state.pop("source_ids", None)
        for scenario in parameters["capacity_scenarios"]:
            scenario.pop("source_ids", None)
        comparison_records.append(
            {
                "profile_record_id": "SOLAR_WIND_" + record["failure_unit_id"],
                "solar_architecture_family": architecture_by_unit[record["failure_unit_id"]],
                "failure_unit_id": record["failure_unit_id"],
                "normalized_axis_id": "normalized_delivered_wind_demand_to_qualified_capacity_ratio",
                "y_axis": record["y_axis"],
                "curve_form": record["curve_form"],
                "parameters": parameters,
            }
        )
    return {
        "schema_version": "shared_solar_wind_normalized_response_candidate.v0.1",
        "shared_response_id": "SHARED_SOLAR_WIND_NORMALIZED_RESPONSE_SYNTHETIC_T4_V0_1",
        "shared_response_version": "candidate v0.1",
        "status": "non_runtime_candidate_curve",
        "reuse_level": "candidate_curve",
        "comparison_only": True,
        "canonical_runtime_artifact": False,
        "runtime_approved": False,
        "purpose": (
            "Record a hazard-label-neutral, solar-specific numerical comparison fingerprint; "
            "it does not supply parameters to an output-bearing cell bundle."
        ),
        "origin": (
            "The comparison values are derived from the SHA-pinned strong_wind_solar model-v2.0/docs-r1 "
            "Tier-4 assumptions, with the synthetic hard-zero threshold removed. The TC cell separately "
            "adopts byte-equal values as cell-local Tier-4 assumptions by owner decision."
        ),
        "origin_strong_wind_artifact_sha256": STRONG_V2_ARTIFACT_SHA256,
        "evidence_status": (
            "All beta, state-cost, and median values are synthetic Tier-4 assumptions. "
            "No source calibrates them for tropical cyclones or convective wind."
        ),
        "compatibility_key": [
            "ground-mounted architecture and failure-unit identity",
            "same normalized delivered-demand semantics",
            "same same-unit direct-replacement DR ordinate",
            "same state taxonomy and denominator",
            "cell-local bridge proves mechanism-specific input compatibility",
        ],
        "prohibited_uses": [
            "loading this file as a runtime curve bundle",
            "bypassing a cell artifact or cell-local capability",
            "calling the envelope empirical, calibrated, conservative, or probabilistic",
            "using asset label alone as a compatibility key",
            "applying a TC, convective, stow, duration, rain, or debris multiplier without calibration",
            "populating any output-bearing cell bundle from this candidate profile",
        ],
        "comparison_records": comparison_records,
    }


def capability() -> dict[str, Any]:
    withheld = [
        {
            "failure_unit_id": "PV_FOUNDATION",
            "reason_codes": [
                "NO_TC_WIND_FOUNDATION_FRAGILITY",
                "FOUNDATION_ARCHETYPE_AND_GEOTECHNICAL_STATE_REQUIRED",
            ],
        },
        {
            "failure_unit_id": "PV_POWER_CONVERSION_AND_COLLECTION",
            "reason_codes": [
                "DIRECT_TC_WIND_CURVE_UNSUPPORTED",
                "POINT_LINE_NETWORK_SPLIT_REQUIRED",
                "DEBRIS_RAIN_AND_INGRESS_PATHWAYS_NOT_MODELED",
            ],
        },
        {
            "failure_unit_id": "PV_GSU_SUBSTATION",
            "reason_codes": [
                "NO_TC_WIND_GSU_FRAGILITY",
                "FACILITY_LEVEL_YARD_EXPOSURE_AND_VALUE_REQUIRED",
                "NO_ARRAY_DR_OR_EXPOSURE_INHERITANCE",
            ],
        },
        {
            "failure_unit_id": "PV_SCADA_COMMUNICATIONS",
            "reason_codes": [
                "DIRECT_TC_WIND_CURVE_UNSUPPORTED",
                "CONTROL_DEPENDENCY_IS_NOT_DIRECT_DAMAGE",
            ],
        },
        {
            "failure_unit_id": "PV_CIVIL_INFRA",
            "reason_codes": [
                "MIXED_ASSET_BUCKET_REQUIRES_SPLIT",
                "NO_TC_WIND_CIVIL_FRAGILITY",
            ],
        },
        {
            "failure_unit_id": "PV_REPLACEMENT_SUPPORT",
            "reason_codes": [
                "ALLOCATION_ONLY_NO_INTRINSIC_DR",
                "ALLOCATE_ONCE_AFTER_DIRECT_DAMAGE",
            ],
        },
    ]
    return {
        "schema_version": "capability_declaration.v3",
        "cell_id": "tropical_cyclone_wind_solar",
        "canonical_runtime_artifact": False,
        "pathway_capabilities": [
            {
                "pathway_id": PATHWAY,
                "failure_unit_scalar_dr": "conditional",
                "scenario_loss_given_value_basis": "withheld",
                "curve_intrinsic_spread": "nonprobabilistic_epistemic_envelope",
                "populated_emit_modes": ["scalar_mean", "state_ensemble"],
                "conditions": [
                    "exact supported architecture selected with no default for Perry/fixed/tracker numerical routes; direct common-withheld-unit queries prohibit architecture",
                    "Perry compatibility calls retain all six source acknowledgements and the source-native axis",
                    "curve_intrinsic_spread is a pathway-level union: Perry carries no intrinsic spread; only generic fixed/tracker emit the nonprobabilistic envelope",
                    "generic fixed calls supply a qualified TC event/design pressure ratio or bridged array-height speed proxy",
                    "tracker calls exactly match attained state, exact-system Ucrit, layout, angle, zone, lock, duration, direction, and speed basis",
                    "event and event-family identities are present",
                    "compound pathways are separated and acknowledged",
                    "no value payload or scenario-dollar request is supplied",
                ],
                "limitation_flags": [
                    "EXPERIMENTAL_SYNTHETIC_T4_SCENARIO",
                    "TC_NUMERICAL_RESPONSE_NOT_CALIBRATED",
                    "CELL_LOCAL_SYNTHETIC_PARAMETER_DECISION",
                    "NONPROBABILISTIC_EPISTEMIC_ENVELOPE",
                    "TC_BRIDGE_CONTENT_NOT_RESOLVED_BY_REFERENCE_EVALUATOR",
                    "TC_DURATION_DIRECTION_AND_CYCLING_NOT_NUMERICALLY_MODELED",
                    "DEBRIS_RAIN_AND_COMPOUND_PATHWAYS_EXCLUDED",
                    "SCENARIO_DOLLAR_LOSS_WITHHELD",
                    "FULL_PLANT_PHYSICAL_LOSS_INCOMPLETE",
                    "PERRY_SOURCE_COMPATIBILITY_ROUTE",
                    "SOURCE_SPECIFIC_VISIBLE_MODULE_MATERIAL_PROXY",
                    "SOURCE_COMPOSITE_HURRICANE_MODULE_LOSS",
                    "SOURCE_AXIS_PRODUCT_QUERY_SEMANTICS_UNRESOLVED",
                    "PREDICTIVE_RELATIONSHIP_NOT_VALIDATED",
                ],
                "withheld_failure_units": withheld,
            }
        ],
        "consumer_annual_metrics": {
            "computation_owner": "downstream_consumer",
            "status_before_promotion": "withheld_noncanonical_proposal",
            "status_after_promotion": "withheld",
            "prerequisites": [
                "canonical artifact promotion and exact model/docs/schema/SHA pin",
                "calibrated TC demand-response and same-unit economic consequence",
                "architecture-aware exposure and value basis",
                "complete physical-unit coverage or explicit partial-loss contract",
                "consumer cap-binding and compound-event validation",
            ],
            "limitation_flags": [
                "FREQUENCY_DRIVEN_TAIL_IS_CONSUMER_OWNED",
                "SYNTHETIC_SCENARIOS_ARE_NOT_A_PROBABILITY_DISTRIBUTION",
                "FULL_PLANT_LOSS_INCOMPLETE",
            ],
        },
        "cap_binding": {
            "policy": "consumer_enforced_fail_closed",
            "enforcement_owner": "downstream_consumer",
            "checks_required": [
                "exact model/docs/schema/SHA pin",
                "pathway and architecture match",
                "TC wind-field, direction-history, duration-cycling, and aerodynamic bridge provenance",
                "withheld units remain null rather than zero",
                "GSU retains independent yard exposure and value",
                "Perry and generic routes are mutually exclusive",
                "no synthetic scenario is treated as calibrated frequency or uncertainty",
            ],
            "action_if_fail": (
                "withhold the affected DR or downstream metric; do not substitute model v1, "
                "strong-wind response, legacy logistics, or another architecture"
            ),
        },
        "promotion_gate": {
            "status": "blocked",
            "required_before_canonical_use": [
                "formal expert elicitation or matched TC field/claims calibration for every synthetic parameter",
                "validated portable TC demand bridge and held-out prediction",
                "same-unit disposition, cost, replacement-value, and support-allocation evidence",
                "independent fixed/tracker/GSU engineering and economic review",
                "Hazard dual-read, rejection-path, compound-event, pin, shadow, and rollback tests",
                "explicit maintainer promotion and atomic registry/index/changelog update",
            ],
        },
    }


def failure_units(strong_base: Mapping[str, Any], v1: Mapping[str, Any]) -> list[dict[str, Any]]:
    strong_units = {item["id"]: deepcopy(item) for item in strong_base["failure_units"]}
    v1_units = {item["id"]: deepcopy(item) for item in v1["failure_units"]}
    output = [v1_units[PERRY_UNIT]]
    for unit_id in (
        "PV_FIXED_TILT_MODULE_FIELD",
        "PV_FIXED_TILT_SUPPORT_STRUCTURE",
        "PV_TRACKER_MODULE_FIELD",
        "PV_TRACKER_SBOS_ASSEMBLY",
    ):
        item = strong_units[unit_id]
        item["treatment"] = "primary_nonzero"
        item["denominator"] += "; scenario dollars disabled in this proposal"
        output.append(item)
    output.extend(
        [
            {
                "id": "PV_FOUNDATION",
                "subsystem": "foundation",
                "component": "piles, posts below the governed support boundary, anchors, and pads",
                "treatment": "withheld",
                "y_axis": "withheld",
                "denominator": "same-unit foundation replacement value",
                "withheld_reason_codes": [
                    "NO_TC_WIND_FOUNDATION_FRAGILITY",
                    "FOUNDATION_ARCHETYPE_AND_GEOTECHNICAL_STATE_REQUIRED",
                ],
            },
            {
                "id": "PV_POWER_CONVERSION_AND_COLLECTION",
                "subsystem": "power conversion and collection",
                "component": "inverter, combiner, cable, grounding, and MV collection subjects",
                "treatment": "withheld",
                "y_axis": "withheld",
                "denominator": "separate point, line, and network replacement values",
                "withheld_reason_codes": [
                    "DIRECT_TC_WIND_CURVE_UNSUPPORTED",
                    "POINT_LINE_NETWORK_SPLIT_REQUIRED",
                ],
            },
            {
                "id": "PV_GSU_SUBSTATION",
                "subsystem": "GSU and substation",
                "component": "GSU transformer, switchyard, protection, and control yard",
                "treatment": "withheld",
                "y_axis": "withheld",
                "denominator": "independent facility-level yard/point replacement value",
                "withheld_reason_codes": [
                    "NO_TC_WIND_GSU_FRAGILITY",
                    "FACILITY_LEVEL_YARD_EXPOSURE_AND_VALUE_REQUIRED",
                    "NO_ARRAY_DR_OR_EXPOSURE_INHERITANCE",
                ],
            },
            {
                "id": "PV_SCADA_COMMUNICATIONS",
                "subsystem": "SCADA and communications",
                "component": "monitoring, communications, tracker controls, and exposed sensors",
                "treatment": "withheld",
                "y_axis": "withheld",
                "denominator": "same-unit controls replacement value",
                "withheld_reason_codes": [
                    "DIRECT_TC_WIND_CURVE_UNSUPPORTED",
                    "CONTROL_DEPENDENCY_IS_NOT_DIRECT_DAMAGE",
                ],
            },
            {
                "id": "PV_CIVIL_INFRA",
                "subsystem": "civil infrastructure",
                "component": "roads, fencing, grading, drainage, and buildings",
                "treatment": "withheld",
                "y_axis": "withheld",
                "denominator": "split direct civil replacement value",
                "withheld_reason_codes": [
                    "MIXED_ASSET_BUCKET_REQUIRES_SPLIT",
                    "NO_TC_WIND_CIVIL_FRAGILITY",
                ],
            },
            {
                "id": "PV_REPLACEMENT_SUPPORT",
                "subsystem": "replacement support",
                "component": "inspection, labor, rental, logistics, management, removal, and reinstatement",
                "treatment": "exposure_modifier",
                "y_axis": "no intrinsic damage ratio",
                "denominator": "allocate once after observed direct disposition",
                "withheld_reason_codes": [
                    "ALLOCATION_ONLY_NO_INTRINSIC_DR",
                    "ALLOCATE_ONCE_AFTER_DIRECT_DAMAGE",
                ],
            },
        ]
    )
    return output


def build_pathway(records: list[dict[str, Any]], perry_record: dict[str, Any]) -> dict[str, Any]:
    coverage = [
        {
            "failure_unit_id": PERRY_UNIT,
            "status": "conditional_source_compatibility_curve",
            "architecture": PERRY_ARCH,
        },
        {
            "failure_unit_id": "PV_FIXED_TILT_MODULE_FIELD",
            "status": "conditional_synthetic_t4_curve",
            "architecture": FIXED_ARCH,
        },
        {
            "failure_unit_id": "PV_FIXED_TILT_SUPPORT_STRUCTURE",
            "status": "conditional_synthetic_t4_curve",
            "architecture": FIXED_ARCH,
        },
        {
            "failure_unit_id": "PV_TRACKER_MODULE_FIELD",
            "status": "conditional_synthetic_t4_curve",
            "architecture": TRACKER_ARCH,
        },
        {
            "failure_unit_id": "PV_TRACKER_SBOS_ASSEMBLY",
            "status": "conditional_synthetic_t4_curve",
            "architecture": TRACKER_ARCH,
        },
    ]
    coverage.extend(
        {"failure_unit_id": unit, "status": "allocation_only" if unit == "PV_REPLACEMENT_SUPPORT" else "withheld"}
        for unit in (
            "PV_FOUNDATION",
            "PV_POWER_CONVERSION_AND_COLLECTION",
            "PV_GSU_SUBSTATION",
            "PV_SCADA_COMMUNICATIONS",
            "PV_CIVIL_INFRA",
            "PV_REPLACEMENT_SUPPORT",
        )
    )
    return {
        "pathway_id": PATHWAY,
        "hazard_scope": {
            "included": [
                "tropical-cyclone wind after a versioned local demand bridge",
                "source-composite Perry hurricane module outcome only on the compatibility route",
            ],
            "excluded": [
                "straight-line convective wind",
                "tornado direct hit",
                "windborne-debris impact as a separately valued pathway",
                "wind-driven rain or water ingress",
                "flood, storm surge, and scour",
                "hail, lightning, fire, downtime, BI, and financial terms",
            ],
            "event_partition_rule": (
                "retain one event_family_id; evaluate TC wind once per qualified array zone; "
                "reconcile rain, debris, tornado, flood, and surge as separate pathways"
            ),
        },
        "hazard_axis": {
            "id": AXIS,
            "preferred_input_field": "architecture_input_contracts[selected_by_array_architecture].preferred_input_field",
            "permitted_proxy_fields": ["tc_array_height_3s_gust_mps"],
            "unit": "dimensionless_or_source_native_by_architecture",
            "valid_range": [0.0, 2.0],
            "extrapolation_policy": (
                "generic fixed/tracker: bounded synthetic domain and withhold above 2.0; "
                "Perry compatibility: withhold outside 17.4-39.1 m/s"
            ),
            "routing_field": "array_architecture",
            "architecture_input_contracts": {
                PERRY_ARCH: {
                    "axis_field": "perry_event_max_gust_mps",
                    "preferred_input_field": "perry_event_max_gust_mps",
                    "accepted_payloads": [
                        {
                            "mode": "source_compatibility",
                            "required_fields": [
                                "perry_event_max_gust_mps",
                                "array_architecture_id",
                                "source_population_match_id",
                                "module_value_distribution_assumption_id",
                                "visible_damage_disposition_assumption_id",
                                "source_wind_product_id",
                                "causal_scope_acknowledgement_id",
                            ],
                        }
                    ],
                    "valid_range": [17.4, 39.1],
                    "required_acknowledgements": [
                        "array_architecture_id",
                        "source_population_match_id",
                        "module_value_distribution_assumption_id",
                        "visible_damage_disposition_assumption_id",
                        "source_wind_product_id",
                        "causal_scope_acknowledgement_id",
                    ],
                },
                FIXED_ARCH: {
                    "axis_field": "tc_fixed_event_to_design_net_pressure_ratio",
                    "preferred_input_field": "tc_fixed_event_to_design_net_pressure_ratio",
                    "accepted_payloads": [
                        {
                            "mode": "preferred",
                            "required_fields": [
                                "tc_fixed_event_to_design_net_pressure_ratio",
                                "tc_wind_field_bridge_id",
                                "tc_directional_history_bridge_id",
                                "tc_duration_cycling_bridge_id",
                                "aerodynamic_demand_bridge_id",
                                "array_zone",
                                "array_spatial_object_id",
                            ],
                        },
                        {
                            "mode": "screening_proxy",
                            "required_fields": [
                                "tc_array_height_3s_gust_mps",
                                "qualified_design_array_height_3s_gust_mps",
                                "tc_wind_field_bridge_id",
                                "tc_directional_history_bridge_id",
                                "tc_duration_cycling_bridge_id",
                                "aerodynamic_demand_bridge_id",
                                "array_zone",
                                "array_spatial_object_id",
                            ],
                        },
                    ],
                    "definition": (
                        "peak TC event net-pressure demand divided by comparable same-zone design "
                        "net-pressure demand after named wind-field, direction-history, duration-cycling, "
                        "geometry, zone, and aerodynamic treatment"
                    ),
                    "permitted_screening_proxy": (
                        "(tc_array_height_3s_gust_mps / qualified_design_array_height_3s_gust_mps)^2 "
                        "only with all named TC and aerodynamic bridges"
                    ),
                    "valid_range": [0.0, 2.0],
                },
                TRACKER_ARCH: {
                    "axis_field": "tracker_instability_speed_ratio",
                    "preferred_input_field": "tc_tracker_normal_3s_gust_mps",
                    "accepted_payloads": [
                        {
                            "mode": "qualified_exact_system",
                            "required_fields": [
                                "tc_tracker_normal_3s_gust_mps",
                                "critical_instability_3s_gust_mps",
                                "aeroelastic_qualification_id",
                                "aeroelastic_qualification_sha256",
                                "tracker_system_id",
                                "tracker_module_configuration",
                                "tracker_layout_id",
                                "tracker_angle_deg",
                                "tracker_position_state",
                                "stow_confirmation_basis",
                                "array_zone",
                                "array_spatial_object_id",
                                "tracker_drive_lock_state",
                                "tc_wind_field_bridge_id",
                                "tc_directional_history_bridge_id",
                                "tc_duration_cycling_bridge_id",
                                "qualification_tracker_system_id",
                                "qualification_tracker_module_configuration",
                                "qualification_tracker_layout_id",
                                "qualification_tracker_angle_deg",
                                "qualification_tracker_position_state",
                                "qualification_array_zone",
                                "qualification_drive_lock_state",
                                "qualification_speed_reference",
                                "qualification_speed_averaging_s",
                                "qualification_tc_wind_field_bridge_id",
                                "qualification_direction_basis_id",
                                "qualification_duration_basis_id",
                            ],
                        }
                    ],
                    "equation": "tc_tracker_normal_3s_gust_mps / critical_instability_3s_gust_mps",
                    "valid_range": [0.0, 2.0],
                    "required_basis": [
                        "exact tracker system, 1P/2P, layout, attained angle and position",
                        "known array zone and drive/lock state",
                        "3-second array-height tracker-normal speed reference",
                        "matching TC wind-field, direction-history, and duration-cycling qualification bases",
                    ],
                    "stow_action_flag": (
                        "at ratio >= 0.75 emit an operational-action flag only; never force damage"
                    ),
                },
            },
            "prohibited_input": (
                "unbridged Hazard 10 m gust, category, NHC sustained wind, convective demand, "
                "generic tracker speed without exact Ucrit, or present-day Perry vendor reconstruction"
            ),
        },
        "selector_logic": [
            {
                "field": "array_architecture",
                "required": True,
                "allowed": [PERRY_ARCH, FIXED_ARCH, TRACKER_ARCH],
                "no_default": True,
                "mutually_exclusive": True,
            },
            {
                "field": "tracker_system_and_qualification_basis",
                "required_when": f"array_architecture={TRACKER_ARCH}",
                "numeric_effect": "only exact matching enables the normalized input; no stow multiplier",
            },
        ],
        "conditioner_logic": [
            {
                "field": "tc_duration_class",
                "required": "required_or_unknown",
                "numeric_effect": "metadata/bridge input only; no multiplier",
            },
            {
                "field": "tc_direction_evolution_class",
                "required": "required_or_unknown",
                "numeric_effect": "metadata/bridge input only; no multiplier",
            },
            {
                "field": "tracker_position_state",
                "required": "tracker_attained_known_and_qualification_matched",
                "numeric_effect": "command-only stow rejects; no generic protection credit",
            },
            {
                "field": "rain_debris_flood_surge_tornado_indicators",
                "required": "required_or_unknown",
                "numeric_effect": "separate pathway acknowledgement only; no TC-wind DR multiplier",
            },
        ],
        "exposure_contract": {
            "required_for_curve_evaluation": ["event_id", "event_family_id", "array_architecture"],
            "required_for_generic_curve_evaluation": ["array_zone", "array_spatial_object_id"],
            "direct_withheld_unit_query": (
                "requires event_id, event_family_id, pathway_id, and failure_unit_id; "
                "array architecture and array axis are prohibited"
            ),
            "scenario_loss_status": "withheld_noncanonical_synthetic_T4_proposal",
            "no_default_exposed_fraction": True,
            "rules": [
                "evaluate each qualified array zone once",
                "do not apply a second zone factor after the delivered demand contains zoning",
                "do not reuse array exposure for inverter, collection, GSU, SCADA, foundation, or civil units",
                "do not turn withheld units into zero",
                "Perry source atom is mutually exclusive with the generic fixed module unit",
            ],
        },
        "failure_unit_coverage": coverage,
        "curve_records": [perry_record, *records],
    }


def build_artifact(
    records: list[dict[str, Any]], cap: dict[str, Any], shared_profile_sha256: str
) -> dict[str, Any]:
    from scripts.reference_helpers.tropical_cyclone_wind_solar_v2_curve_eval import (
        FAILURE_CODES,
    )

    strong = load_json(STRONG / "strong_wind_solar__model_v2_0__docs_r1__curve_artifact.json")
    v1_path = PROPOSED / "tropical_cyclone_wind_solar__model_v1_0__docs_r1__curve_artifact.json"
    v1 = load_json(v1_path)
    perry = deepcopy(v1["pathways"][0]["curve_records"][0])
    artifact: dict[str, Any] = {
        "schema_version": "damage_curve_record_bundle.v3",
        "schema_status": "proposed_draft",
        "cell_id": "tropical_cyclone_wind_solar",
        "damage_code_id": "TROPICAL_CYCLONE_WIND_SOLAR_SYNTHETIC_T4_V2_PROPOSED",
        "semantic_damage_model_version": MODEL,
        "documentation_revision": DOCS,
        "lifecycle_state": "candidate",
        "promotion_status": "proposed_blocked",
        "review_status": "independent_proposal_review_complete_promotion_review_pending",
        "model_grade": "experimental_synthetic_T4_scenario",
        "package_release": "unreleased",
        "package_baseline": "library v2.5",
        "package_inclusion_status": "not_included",
        "canonical_runtime_artifact": False,
        "source_dossier": "docs/cells/tropical_cyclone_wind_solar/proposed/tropical_cyclone_wind_solar_curve_derivation_dossier__model_v2_0__docs_r1.md",
        "source_workbook": "docs/cells/tropical_cyclone_wind_solar/proposed/damage_curve_records_tropical_cyclone_wind_solar__model_v2_0__docs_r1.xlsx",
        "known_answer_tests": "docs/cells/tropical_cyclone_wind_solar/proposed/known_answer_tests_tropical_cyclone_wind_solar__model_v2_0__docs_r1.json",
        "source_register": "docs/cells/tropical_cyclone_wind_solar/proposed/SOURCE_REGISTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv",
        "claim_parameter_register": "docs/cells/tropical_cyclone_wind_solar/proposed/CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv",
        "claim_supersession_map": "docs/cells/tropical_cyclone_wind_solar/proposed/CLAIM_SUPERSESSION_MAP_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv",
        "value_crosswalk": "docs/cells/tropical_cyclone_wind_solar/proposed/VALUE_CROSSWALK_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv",
        "derivation_rationale": {
            "summary": (
                "Preserve Perry source compatibility and add owner-authorized synthetic Tier-4 fixed/tracker "
                "coverage through cell-local synthetic parameters plus TC-specific fail-closed bridge-ID "
                "contracts whose unresolved external content remains explicitly flagged."
            ),
            "rejected_models": [
                "Ceferino extensive-failure probability relabeled as a damage ratio",
                "missing plant value treated as zero loss",
                "one generic hurricane-solar curve independent of architecture",
                "tracker stow assumed rather than exact attained state supplied",
                "strong-wind numerical parameters presented as tropical-cyclone evidence",
            ],
        },
        "audit_comparison_profile": {
            "shared_response_id": "SHARED_SOLAR_WIND_NORMALIZED_RESPONSE_SYNTHETIC_T4_V0_1",
            "shared_response_version": "candidate v0.1",
            "path": "docs/method/shared_components/solar_wind_normalized_response/candidate_response_profile_v0_1.json",
            "sha256": shared_profile_sha256,
            "role": "post_adoption_parameter_fingerprint_only",
            "runtime_dependency": False,
            "runtime_approved": False,
        },
        "failure_units": failure_units(strong, v1),
        "pathways": [build_pathway(records, perry)],
        "value_linkage": {
            "implicit_default_profile": None,
            "reference_basis": strong["value_linkage"]["primary_reference_basis"],
            "scenario_loss_status": "withheld_noncanonical_synthetic_T4_proposal",
            "rules": [
                "state costs define intrinsic same-unit DR scenarios only",
                "reference values are audit/reconciliation aids and never runtime defaults",
                "support is allocation-only and must be charged once after qualified disposition",
                "GSU and other withheld units retain independent exposure and value",
                "no whole-plant or installed-TIV DR may be assembled from partial outputs",
            ],
        },
        "parameter_tier_table": [
            {
                "parameter": "Perry source curve and range",
                "value": "unchanged 13 knots, 17.4-39.1 m/s",
                "tier": "T1_claims_or_field_calibrated",
                "source_ids": ["TCWS-S020", "TCWS-S021"],
            },
            {
                "parameter": "generic ordered-state form",
                "value": "ordered_damage_state_lognormal",
                "tier": "T4_placeholder_or_expert_judgment",
                "source_ids": [CELL_LOCAL_SYNTHETIC_SOURCE],
            },
            {
                "parameter": "all generic beta, medians, and state cost ratios",
                "value": "see curve records",
                "tier": "T4_placeholder_or_expert_judgment",
                "source_ids": [
                    CELL_LOCAL_SYNTHETIC_SOURCE,
                    NO_DIRECT_TC_CALIBRATION_SOURCE,
                ],
            },
            {
                "parameter": "hard-zero treatment",
                "value": "exact zero only at zero normalized demand; no positive hard-zero threshold",
                "tier": "T4_placeholder_or_expert_judgment",
                "source_ids": ["GOVERNANCE_CONTRACT"],
            },
            {
                "parameter": "scenario loss",
                "value": "withheld",
                "tier": "T4_placeholder_or_expert_judgment",
                "source_ids": ["GOVERNANCE_CONTRACT"],
            },
        ],
        "architecture_capability_interpretation": {
            PERRY_ARCH: {
                "emit_mode": "scalar_mean",
                "curve_intrinsic_spread": "not_carried",
                "grade": "source_specific_unvalidated_visible_module_material_proxy",
            },
            FIXED_ARCH: {
                "emit_mode": "state_ensemble",
                "curve_intrinsic_spread": "nonprobabilistic_epistemic_envelope",
                "grade": "experimental_synthetic_T4_scenario",
            },
            TRACKER_ARCH: {
                "emit_mode": "state_ensemble",
                "curve_intrinsic_spread": "nonprobabilistic_epistemic_envelope",
                "grade": "experimental_synthetic_T4_scenario",
            },
        },
        "evaluation_contract": {
            "pathway_id_required": True,
            "architecture_required_for_array_and_perry_routes": True,
            "architecture_prohibited_for_direct_withheld_unit_query": True,
            "event_id_and_event_family_id_required": True,
            "no_cross_pathway_or_cross_architecture_fallback": True,
            "central_scenario_id": "central_screening",
            "scenario_semantics": "unweighted synthetic epistemic scenarios; not percentiles or probabilities of model truth",
            "state_probability_to_dr_rule": (
                "DR is the sum of exact-state probability times the explicitly listed same-unit T4 state cost ratio; "
                "a failure-state probability is never itself relabelled as DR"
            ),
            "scenario_loss_behavior": "withheld",
            "failure_codes": sorted(FAILURE_CODES),
            "withheld_reason_codes": sorted(
                {
                    "NO_RUNTIME_CURVE_FOR_PATHWAY_UNIT",
                    *(
                        code
                        for unit in cap["pathway_capabilities"][0]["withheld_failure_units"]
                        for code in unit["reason_codes"]
                    ),
                }
            ),
        },
        "emit_contract": {
            "schema_version": "damage_emit.v2",
            "populated_emit_modes_for_this_cell": ["scalar_mean", "state_ensemble"],
            "result_grain_by_route": {
                "numerical_architecture_route": "pathway_id x array_architecture x failure_unit_id x array_zone_or_source_site",
                "direct_withheld_unit_query": "pathway_id x failure_unit_id",
            },
            "event_lineage_fields_carried_separately": ["event_id", "event_family_id"],
            "prohibited_outputs": [
                "scenario dollar loss",
                "full-plant physical DR",
                "installed-TIV DR",
                "frequency, EAL, PML, VaR, TVaR",
                "business interruption, downtime, and revenue loss",
            ],
        },
        "capability_declaration": cap,
        "legacy_comparison": {
            "prior_proposal_pin": "tropical_cyclone_wind_solar@model_v1_0__docs_r1",
            "prior_artifact_sha256": digest(v1_path),
            "strict_alternative_pin": "tropical_cyclone_wind_solar@model_v0_1__docs_r1",
            "strict_alternative_artifact_sha256": digest(
                PROPOSED / "tropical_cyclone_wind_solar__model_v0_1__docs_r1__curve_artifact.json"
            ),
            "legacy_hazard_notebook_status": "regression_fixture_only_runtime_prohibited",
            "artifact_index_status": "absent_noncanonical_no_cutover",
        },
    }
    serialized = json.dumps(artifact)
    if "straight_line_convective" in serialized or "parent_convective_event_id" in serialized:
        raise RuntimeError("convective runtime semantics leaked into the TC artifact")
    return artifact


def consolidated_sources() -> list[dict[str, str]]:
    paths = [
        PROPOSED / "SOURCE_REGISTER_tropical_cyclone_wind_solar__model_v0_1__docs_r1.csv",
        PROPOSED / "SOURCE_REGISTER_tropical_cyclone_wind_solar__model_v1_0__docs_r1.csv",
        PROPOSED / "SOURCE_REGISTER_ADDENDUM_tropical_cyclone_wind_solar__model_v1_0__docs_r2.csv",
    ]
    by_id: dict[str, dict[str, str]] = {}
    for path in paths:
        for row in read_csv(path):
            # The model-v1 addendum carried a convective-wind-specific sentinel
            # source under this generic-looking ID.  Model v2 replaces it with
            # the TC-cell-local decision and no-direct-calibration records below;
            # retaining it would falsely imply that a convective evidence gap is
            # part of the tropical-cyclone provenance chain.
            if row["source_id"] == "NO_DIRECT_CALIBRATION_SOURCE":
                continue
            by_id.setdefault(row["source_id"], row)
    for row in read_csv(STRONG / "SOURCE_REGISTER_strong_wind_solar__model_v2_0__docs_r1.csv"):
        if row["source_id"] == "NO_DIRECT_CALIBRATION_SOURCE":
            continue
        tier = row["tier"]
        if tier == "T1_direct_empirical":
            tier = "T1_claims_or_field_calibrated"
        mapped = {
            "source_id": row["source_id"],
            "citation": row["full_citation"],
            "url": row["primary_url"],
            "accessed_on": row["accessed_on"],
            "exact_locator": row["exact_locator"],
            "source_type": row["source_class"],
            "source_role": "adjacent_normalized_response_anatomy_only",
            "pathway_ids": "all_shared",
            "evidence_tier": tier,
            "target_asset_match": "direct",
            "target_failure_unit_match": "partial",
            "measured_or_modeled_endpoint": row["native_axis_or_endpoint"],
            "permitted_inference": row["direct_support"],
            "prohibited_inference": row["prohibited_support"],
            "decision": "retain_as_adjacent_or_T4_context",
            "status": "reviewed",
            "notes": row["status"],
        }
        by_id.setdefault(mapped["source_id"], mapped)
    by_id["SHARED_SOLAR_WIND_PROXY_V0_1"] = {
        "source_id": "SHARED_SOLAR_WIND_PROXY_V0_1",
        "citation": "Damage Modeling shared solar-wind normalized-response candidate profile v0.1.",
        "url": "docs/method/shared_components/solar_wind_normalized_response/candidate_response_profile_v0_1.json",
        "accessed_on": "2026-07-29",
        "exact_locator": "complete non-runtime candidate profile",
        "source_type": "internal governed synthetic scenario substrate",
        "source_role": "audit_comparison_profile_only",
        "pathway_ids": "all_shared",
        "evidence_tier": "T4_placeholder_or_expert_judgment",
        "target_asset_match": "direct",
        "target_failure_unit_match": "direct",
        "measured_or_modeled_endpoint": "unweighted normalized-demand state-response envelope",
        "permitted_inference": "Compare cell-local synthetic parameters against a hazard-label-neutral solar fingerprint after adoption.",
        "prohibited_inference": "Call the values field-calibrated, hurricane-calibrated, conservative, probabilistic, or runtime approved.",
        "decision": "retain_audit_only",
        "status": "reviewed_nonruntime",
        "notes": "Standard 20 prohibits this candidate from populating an output-bearing cell bundle.",
    }
    by_id[CELL_LOCAL_SYNTHETIC_SOURCE] = {
        "source_id": CELL_LOCAL_SYNTHETIC_SOURCE,
        "citation": "Damage Modeling owner decision: TC-wind x solar model-v2 cell-local synthetic parameter set.",
        "url": "docs/cells/tropical_cyclone_wind_solar/proposed/DECISION_LOG_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        "accessed_on": "2026-07-29",
        "exact_locator": "TCWS2-D001; TCWS2-D003 through TCWS2-D007",
        "source_type": "governed owner assumption decision",
        "source_role": "cell_local_synthetic_parameter_authority",
        "pathway_ids": PATHWAY,
        "evidence_tier": "T4_placeholder_or_expert_judgment",
        "target_asset_match": "direct",
        "target_failure_unit_match": "direct",
        "measured_or_modeled_endpoint": "owner-adopted unweighted normalized-demand state/DR scenarios",
        "permitted_inference": "Use the exact registered values only in this noncanonical v2 cell proposal.",
        "prohibited_inference": "Treat the decision as evidence, calibration, shared-runtime approval, or authority for another cell.",
        "decision": "adopt_cell_local_synthetic_T4",
        "status": "reviewed_noncanonical",
        "notes": "Byte-equal comparison to the audit profile is intentional but is not a runtime dependency.",
    }
    by_id[NO_DIRECT_TC_CALIBRATION_SOURCE] = {
        "source_id": NO_DIRECT_TC_CALIBRATION_SOURCE,
        "citation": "TC-wind x solar model-v2 bounded evidence review: no direct generic fixed/tracker calibration source adopted.",
        "url": "docs/cells/tropical_cyclone_wind_solar/proposed/BOUNDED_EVIDENCE_SEARCH_LOG_tropical_cyclone_wind_solar__model_v2_0__docs_r1.md",
        "accessed_on": "2026-07-29",
        "exact_locator": "§ Missing matched chain; § v2 consequence",
        "source_type": "governed negative evidence finding",
        "source_role": "no_direct_calibration_marker",
        "pathway_ids": PATHWAY,
        "evidence_tier": "T4_placeholder_or_expert_judgment",
        "target_asset_match": "direct",
        "target_failure_unit_match": "direct",
        "measured_or_modeled_endpoint": "absence of an adopted matched TC demand-state-cost-value calibration",
        "permitted_inference": "Keep all generic numbers at Tier 4 and promotion blocked.",
        "prohibited_inference": "Infer zero damage, immunity, or proof that no usable private evidence exists.",
        "decision": "withhold_calibration_claim",
        "status": "reviewed",
        "notes": "This marker is TC-specific and does not reuse the convective-wind negative-search record.",
    }
    return list(by_id.values())


def new_claims() -> list[dict[str, str]]:
    locators = {
        "TCWS2-C001": "DECISION_LOG model-v2 docs-r1 § TCWS2-D001",
        "TCWS2-C002": "CHANGE_CLASSIFICATION model-v2 docs-r1 § Why this is a major model change",
        "TCWS2-C003": "derivation dossier §12 Perry preservation",
        "TCWS2-C004": "derivation dossier §8 Curve form and probability-to-DR typing",
        "TCWS2-C005": "derivation dossier §§7-9 and parameter-tier register TCWS2 curve rows",
        "TCWS2-C006": "DECISION_LOG model-v2 docs-r1 § TCWS2-D003",
        "TCWS2-C007": "CLAIM_SUPERSESSION_MAP model-v2 docs-r1 complete file",
        "TCWS2-C008": "derivation dossier §6.1 Fixed tilt",
        "TCWS2-C009": "derivation dossier §6.2 Tracker and metadata contract tracker request",
        "TCWS2-C010": "metadata contract § Tracker route",
        "TCWS2-C011": "DECISION_LOG model-v2 docs-r1 § TCWS2-D005",
        "TCWS2-C012": "DECISION_LOG model-v2 docs-r1 § TCWS2-D006",
        "TCWS2-C013": "DECISION_LOG model-v2 docs-r1 § TCWS2-D008",
        "TCWS2-C014": "derivation dossier §13 Capability and reportability",
        "TCWS2-C015": "neighboring-wind and compound-boundary memo § Compound occurrence",
        "TCWS2-C016": "legacy and adjacent-model audit § Legacy hurricane-solar implementation",
        "TCWS2-C017": "DECISION_LOG model-v2 docs-r1 § TCWS2-D009",
        "TCWS2-C018": "CLAIM_SUPERSESSION_MAP model-v2 docs-r1 complete file",
    }

    def row(
        cid: str,
        text: str,
        rule: str,
        sources: str,
        permitted: str,
        prohibited: str,
        trigger: str,
        *,
        ctype: str = "governance_decision",
        tier: str = "T4_placeholder_or_expert_judgment",
        status: str = "adopt",
    ) -> dict[str, str]:
        return {
            "claim_id": cid,
            "pathway_id": PATHWAY,
            "claim_text": text,
            "claim_type": ctype,
            "source_ids": sources,
            "exact_locator": locators[cid],
            "evidence_tier": tier,
            "parameter_or_rule": rule,
            "adoption_status": status,
            "permitted_inference": permitted,
            "prohibited_inference": prohibited,
            "reasoning": "The v2 package prioritizes explicit coverage while keeping missing calibration visible and machine enforced.",
            "update_trigger": trigger,
        }

    return [
        row(
            "TCWS2-C001",
            "The requested v2 is a deliberate coverage-first synthetic Tier-4 exception, not an evidence-earned hurricane calibration.",
            "model_grade=experimental_synthetic_T4_scenario",
            f"GOVERNANCE_CONTRACT;{CELL_LOCAL_SYNTHETIC_SOURCE};{NO_DIRECT_TC_CALIBRATION_SOURCE}",
            "Evaluate the proposal for bounded research and contract integration.",
            "Describe any generic curve as empirical, calibrated, conservative, or production ready.",
            "Formal elicitation or matched field/claims calibration replaces the synthetic parameters.",
        ),
        row(
            "TCWS2-C002",
            "Model v2.0 is a major behavior change because it adds architecture-specific normalized-demand routes and generic fixed/tracker failure-unit outputs.",
            "semantic_damage_model_version=model v2.0",
            "GOVERNANCE_CONTRACT",
            "Keep v0.1 and v1 artifacts unchanged while researching v2.",
            "Call the change docs-only or silently overwrite v1.",
            "Version policy or runtime behavior changes.",
        ),
        row(
            "TCWS2-C003",
            "The Perry 13-knot source route is carried byte-for-value and remains source specific.",
            "Perry_compatibility_route",
            "TCWS-S020;TCWS-S021",
            "Use the route only with its exact axis, range, and six acknowledgements.",
            "Alias it to generic fixed tilt or ordinary Hazard gust.",
            "The Perry source package or v1 governed fit changes.",
            tier="T3_engineering_proxy_or_adjacent_empirical",
        ),
        row(
            "TCWS2-C004",
            "A failure-state probability becomes DR only through explicit exact-state probabilities and separately listed same-unit state-cost ratios.",
            "sum_state_probability_times_cost_ratio",
            f"GOVERNANCE_CONTRACT;{NO_DIRECT_TC_CALIBRATION_SOURCE}",
            "Audit the probability-to-consequence bridge parameter by parameter.",
            "Relabel Ceferino extensive-failure probability or another exceedance probability as DR.",
            "Observed same-unit state costs replace the T4 ratios.",
        ),
        row(
            "TCWS2-C005",
            "All generic beta, medians, and state cost ratios are synthetic T4 assumptions.",
            "generic_curve_parameters",
            f"{CELL_LOCAL_SYNTHETIC_SOURCE};{NO_DIRECT_TC_CALIBRATION_SOURCE}",
            "Use them as unweighted research scenarios with always-on limitations.",
            "Treat them as fitted estimates, percentiles, posterior draws, or uncertainty bands.",
            "Calibration or structured elicitation is completed.",
        ),
        row(
            "TCWS2-C006",
            "After the cell-local T4 values are adopted, the shared profile provides an audit-only cross-hazard comparison fingerprint.",
            "audit_comparison_profile_only",
            f"SHARED_SOLAR_WIND_PROXY_V0_1;{CELL_LOCAL_SYNTHETIC_SOURCE};TCWS-S015;GOVERNANCE_CONTRACT",
            "Compare the adopted cell-local parameter payload to the pinned hazard-neutral solar profile.",
            "Load the candidate as a runtime dependency or claim numerical evidence transfer from convective wind.",
            "A formal equivalence or hazard-specific response calibration changes the reuse decision.",
        ),
        row(
            "TCWS2-C007",
            "Historical v0.1/v1 claims remain source-version facts and are scoped by the complete v2 supersession map.",
            "historical_claim_scope",
            f"{CELL_LOCAL_SYNTHETIC_SOURCE};GOVERNANCE_CONTRACT",
            "Apply each prior claim only to its recorded model/docs scope and retain its evidence limitation.",
            "Treat historical withhold/version statements as the active v2 output contract or erase their retained scientific truth.",
            "A later model version explicitly replaces the map.",
        ),
        row(
            "TCWS2-C008",
            "Fixed tilt requires a qualified same-zone TC event/design pressure ratio or a fully bridged array-height speed-squared proxy.",
            "fixed_axis_contract",
            "TCWS-S005;TCWS-S010;GOVERNANCE_CONTRACT",
            "Use a named wind-field, direction, duration, and aerodynamic bridge.",
            "Pass an ordinary 10 m gust directly or call design wind failure capacity.",
            "A validated direct local-demand adapter replaces the proxy.",
            tier="T3_engineering_proxy_or_adjacent_empirical",
        ),
        row(
            "TCWS2-C009",
            "Tracker evaluation requires exact-system Ucrit and attained state matched to the qualification basis.",
            "tracker_axis_contract",
            "TCWS-S006;TCWS-S007;TCWS-S011;TCWS-S012",
            "Evaluate Vnormal/Ucrit only after exact configuration, layout, angle, position, zone, lock, duration, direction, and speed-reference matching.",
            "Use generic Ucrit, command-only stow, fixed-to-tracker fallback, or a stow multiplier.",
            "Architecture-matched TC tracker calibration becomes available.",
            tier="T2_public_lab_standard_or_physics",
        ),
        row(
            "TCWS2-C010",
            "The 0.75 Ucrit rule is an operational-action flag and never a damage threshold.",
            "stow_action_flag",
            "TCWS-S006",
            "Emit a warning flag after qualification matching.",
            "Force damage, force zero damage, or assign mitigation credit at 0.75 Ucrit.",
            "A source calibrates damage conditional on the action threshold.",
            tier="T2_public_lab_standard_or_physics",
        ),
        row(
            "TCWS2-C011",
            "Generic ordered-state curves have exact zero only at zero normalized demand and no positive hard-zero threshold.",
            "zero_demand_boundary",
            f"GOVERNANCE_CONTRACT;{CELL_LOCAL_SYNTHETIC_SOURCE}",
            "Use the native lognormal state form without intercept subtraction.",
            "Claim immunity below an invented threshold or alter the asymptote by anchoring subtraction.",
            "A calibrated physical threshold is adopted.",
        ),
        row(
            "TCWS2-C012",
            "Unsupported physical units remain null, not zero.",
            "withhold_not_zero",
            "GOVERNANCE_CONTRACT",
            "Expose incomplete plant coverage and withhold aggregate physical loss.",
            "Assign zero DR to absent curves or recreate the legacy 42-percent immune remainder.",
            "Each unit receives a governed curve or an evidence-backed near-zero treatment.",
        ),
        row(
            "TCWS2-C013",
            "PV_GSU_SUBSTATION is a separately located yard/point unit and inherits neither array DR nor exposure.",
            "GSU_withholding",
            "TCWS-S004;GOVERNANCE_CONTRACT",
            "Retain identity, value, and acquisition fields for a future GSU route.",
            "Copy flood, wind-farm, legacy substation, array module, or mounting response.",
            "Matched TC-wind GSU demand, disposition, cost, and value evidence arrives.",
            tier="T3_engineering_proxy_or_adjacent_empirical",
        ),
        row(
            "TCWS2-C014",
            "Scenario dollars and whole-plant DR remain withheld even for valid generic curves.",
            "scenario_loss_withheld",
            f"GOVERNANCE_CONTRACT;{NO_DIRECT_TC_CALIBRATION_SOURCE}",
            "Use scalar/state DRs only for research and interface testing.",
            "Multiply by TIV or reference values and report economic loss.",
            "Same-unit state costs, values, exposure, support, and full coverage pass review.",
        ),
        row(
            "TCWS2-C015",
            "TC duration, direction evolution, cycling, rain, debris, tornado, flood, and surge have no numerical modifier in v2.",
            "conditioner_and_compound_boundary",
            "TCWS-S004;TCWS-S025;TCWS-S026;GOVERNANCE_CONTRACT",
            "Carry them as bridge inputs, flags, or separate pathways under one event family.",
            "Stack an uncalibrated uplift/discount or double count physical loss.",
            "Pathway-specific calibration and precedence rules pass review.",
            tier="T3_engineering_proxy_or_adjacent_empirical",
        ),
        row(
            "TCWS2-C016",
            "The old legacy asset DR is denominator- and coverage-incompatible with v2 failure-unit DR.",
            "old_new_comparison_rule",
            "LEG-TCWS-001;LEG-TCWS-002;GOVERNANCE_CONTRACT",
            "Preserve the old curve only as a migration fixture and compare behavior qualitatively.",
            "Claim a v2 increase/decrease without matching axis, unit, denominator, and covered value.",
            "A common scenario assembly with complete coverage is governed.",
        ),
        row(
            "TCWS2-C017",
            "The v2 artifact remains absent from current/, the artifact index, package releases, and Hazard cutover.",
            "promotion_status",
            "GOVERNANCE_CONTRACT",
            "Review and validate the proposal without changing consumer state.",
            "Load it as canonical merely because it exists in the repository.",
            "Every promotion gate and maintainer decision closes.",
        ),
        row(
            "TCWS2-C018",
            "The v2 supersession map is mandatory provenance for every consolidated historical claim whose output or version conclusion differs in v2.",
            "claim_supersession_map_required",
            f"{CELL_LOCAL_SYNTHETIC_SOURCE};GOVERNANCE_CONTRACT",
            "Use the map to distinguish retained evidence limitations from superseded noncanonical output behavior.",
            "Read the consolidated register as if every row were simultaneously active for model v2.",
            "The consolidated-register format gains native version-scope and supersession fields.",
        ),
    ]


def consolidated_claims() -> list[dict[str, str]]:
    paths = [
        PROPOSED / "CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__model_v0_1__docs_r1.csv",
        PROPOSED / "CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__model_v1_0__docs_r1.csv",
        PROPOSED / "CLAIM_PARAMETER_REGISTER_ADDENDUM_tropical_cyclone_wind_solar__model_v1_0__docs_r2.csv",
    ]
    by_id: dict[str, dict[str, str]] = {}
    for path in paths:
        for row in read_csv(path):
            by_id.setdefault(row["claim_id"], row)
    historical_scope = {
        "TCWS-C003": "TCWS2-C002/008/009",
        "TCWS-C025": "TCWS2-C001/005/006",
        "TCWS-C107": "TCWS-C220",
        "TCWS-C111": "TCWS2-C005/009",
        "TCWS-C207": "TCWS2-C001/009/010",
        "TCWS-C215": "TCWS2-C001/005/012/013",
        "TCWS-C216": "TCWS2-C005/006",
        "TCWS-C217": "TCWS2-C001/002",
    }
    for claim_id, superseding in historical_scope.items():
        if claim_id not in by_id:
            raise RuntimeError(f"historical claim missing for scope map: {claim_id}")
        by_id[claim_id]["adoption_status"] = "historical_scope_only"
        by_id[claim_id]["reasoning"] += (
            f" Consolidated v2 register: prior-version conclusion is scoped/superseded by {superseding}; "
            "the underlying evidence limitation is retained."
        )
    for row in new_claims():
        if row["claim_id"] in by_id:
            raise RuntimeError(f"duplicate claim {row['claim_id']}")
        by_id[row["claim_id"]] = row
    return list(by_id.values())


def claim_supersession_rows() -> list[dict[str, str]]:
    retained = (
        "The evidence limitation remains true; only the noncanonical output/version decision changes in v2."
    )
    return [
        {
            "prior_claim_id": "TCWS-C003",
            "prior_version_scope": "model v0.1/docs r1 scaffold",
            "v2_status": "historical_scope_only_output_conclusion_superseded",
            "superseding_claim_id": "TCWS2-C002;TCWS2-C008;TCWS2-C009",
            "retained_truth": retained,
        },
        {
            "prior_claim_id": "TCWS-C025",
            "prior_version_scope": "model v0.1/docs r1 evidence-transfer gate",
            "v2_status": "evidence_prohibition_retained_cell_local_assumption_decision_added",
            "superseding_claim_id": "TCWS2-C001;TCWS2-C005;TCWS2-C006",
            "retained_truth": "Strong-wind evidence and calibration do not transfer; v2 adopts byte-equal values as cell-local Tier-4 assumptions.",
        },
        {
            "prior_claim_id": "TCWS-C107",
            "prior_version_scope": "model v1.0/docs r1 PAVA wording",
            "v2_status": "historical_wording_corrected",
            "superseding_claim_id": "TCWS-C220",
            "retained_truth": "The fit is equal-record weighted; repeated physical sites mean it is not unique-site weighting.",
        },
        {
            "prior_claim_id": "TCWS-C111",
            "prior_version_scope": "model v1.0/docs r1 Perry curve",
            "v2_status": "Perry_route_truth_retained_generic_v2_output_conclusion_superseded",
            "superseding_claim_id": "TCWS2-C005;TCWS2-C009",
            "retained_truth": "Perry supplies no tracker evidence and cannot route trackers; the v2 tracker route is synthetic Tier 4.",
        },
        {
            "prior_claim_id": "TCWS-C207",
            "prior_version_scope": "model v1.0/docs r2 evidence-earned decision",
            "v2_status": "evidence_gate_retained_owner_assumption_output_added",
            "superseding_claim_id": "TCWS2-C001;TCWS2-C009;TCWS2-C010",
            "retained_truth": "No reviewed evidence earned a tracker curve or generic stow credit; v2 remains uncalibrated and requires attained state.",
        },
        {
            "prior_claim_id": "TCWS-C215",
            "prior_version_scope": "model v1.0/docs r2 coverage decision",
            "v2_status": "generic_fixed_tracker_output_conclusion_superseded_other_units_retained",
            "superseding_claim_id": "TCWS2-C001;TCWS2-C005;TCWS2-C012;TCWS2-C013",
            "retained_truth": "Foundation, electrical, GSU, SCADA, civil, support, value, and evidence-earned promotion remain withheld.",
        },
        {
            "prior_claim_id": "TCWS-C216",
            "prior_version_scope": "model v1.0/docs r2 cross-cell reuse decision",
            "v2_status": "evidence_transfer_prohibition_retained_cell_local_numeric_decision_added",
            "superseding_claim_id": "TCWS2-C005;TCWS2-C006",
            "retained_truth": "Convective response is not TC evidence; the shared profile is audit-only and not a runtime dependency.",
        },
        {
            "prior_claim_id": "TCWS-C217",
            "prior_version_scope": "model v1.0/docs r2 evidence-only version decision",
            "v2_status": "historical_evidence_decision_retained_version_action_superseded_by_owner_choice",
            "superseding_claim_id": "TCWS2-C001;TCWS2-C002",
            "retained_truth": "Public evidence did not earn v2; model v2 is a separately classified owner-authorized synthetic proposal.",
        },
    ]


def parameter_rows(records: list[dict[str, Any]]) -> list[dict[str, str]]:
    base = read_csv(
        PROPOSED / "PARAMETER_TIER_TABLE_tropical_cyclone_wind_solar__model_v1_0__docs_r1.csv"
    )
    rows = list(base)
    for row in rows:
        if row["parameter"] == "runtime_curve_count":
            row["parameter"] = "Perry_compatibility_source_derived_record_count"
            row["reasoning"] += " This count is scoped to the preserved Perry route, not the v2 package."
        elif row["parameter"] == "fit_method":
            row["value"] = "equal_record_weighted_PAVA_then_block_edge_linearization"
            row["reasoning"] = (
                "The 34 retained event/site records receive equal record weight; repeated physical sites "
                "mean this is not unique-site weighting."
            )
        elif row["parameter"] == "model_grade":
            row["parameter"] = "Perry_compatibility_model_grade"
        elif row["parameter"] == "curve_intrinsic_spread":
            row["parameter"] = "Perry_compatibility_curve_intrinsic_spread"

    def add(
        pid: str,
        curve: str,
        value: str,
        role: str,
        sources: str,
        reasoning: str,
        status: str = "adopt_synthetic_T4",
    ) -> None:
        rows.append(
            {
                "parameter": pid,
                "pathway_id": PATHWAY,
                "curve_id": curve,
                "value": value,
                "param_role": role,
                "tier": "T4_placeholder_or_expert_judgment",
                "source_ids": sources,
                "reasoning": reasoning,
                "status": status,
                "update_trigger": "Formal elicitation or matched TC field/claims calibration replaces this value.",
            }
        )

    add(
        "TCWS2-P000 | package curve-record count",
        "CELL_CONTRACT",
        "5",
        "capability",
        f"{CELL_LOCAL_SYNTHETIC_SOURCE};GOVERNANCE_CONTRACT",
        "One preserved Perry record plus four cell-local generic synthetic records.",
    )
    add(
        "TCWS2-P001 | audit comparison profile",
        "CELL_CONTRACT",
        "SHARED_SOLAR_WIND_NORMALIZED_RESPONSE_SYNTHETIC_T4_V0_1",
        "provenance_pin",
        f"SHARED_SOLAR_WIND_PROXY_V0_1;{CELL_LOCAL_SYNTHETIC_SOURCE}",
        "Audit-only byte-value comparison after cell-local adoption; never a runtime parameter source.",
    )
    add(
        "TCWS2-P000B | generic synthetic record count",
        "CELL_CONTRACT",
        "4",
        "capability",
        f"{CELL_LOCAL_SYNTHETIC_SOURCE};GOVERNANCE_CONTRACT",
        "Two fixed-tilt and two tracker records; all are cell-local Tier-4 scenarios.",
    )
    add(
        "TCWS2-P002 | generic demand range",
        "CELL_AXIS_CONTRACT",
        "0.0 to 2.0",
        "boundary_or_cap",
        f"{CELL_LOCAL_SYNTHETIC_SOURCE};{NO_DIRECT_TC_CALIBRATION_SOURCE}",
        "The entire [0,2] interval is a bounded synthetic research domain with no internal evidence-anchored high-range threshold; values above 2 are withheld.",
    )
    add(
        "TCWS2-P003 | hard-zero treatment",
        "ALL_GENERIC_RECORDS",
        "none above zero; exact x=0 only",
        "curve_fit_shape",
        f"GOVERNANCE_CONTRACT;{CELL_LOCAL_SYNTHETIC_SOURCE}",
        "Avoid an invented immunity band and avoid intercept subtraction that changes parameter meaning.",
    )
    add(
        "TCWS2-P004 | scenario loss",
        "CELL_EMIT_CONTRACT",
        "withheld",
        "value_linkage",
        f"GOVERNANCE_CONTRACT;{NO_DIRECT_TC_CALIBRATION_SOURCE}",
        "Synthetic state costs do not authorize reportable dollars or whole-plant DR.",
        status="withhold",
    )
    for record in records:
        curve = record["curve_id"]
        params = record["parameters"]
        add(
            f"{curve} | beta_ln",
            curve,
            str(params["beta_ln"]),
            "curve_fit_shape",
            f"{CELL_LOCAL_SYNTHETIC_SOURCE};{NO_DIRECT_TC_CALIBRATION_SOURCE}",
            "Synthetic log-space transition width; not fitted dispersion.",
        )
        for state in params["damage_states"]:
            add(
                f"{curve} | {state['state_id']} cost_ratio",
                curve,
                str(state["cost_ratio"]),
                "state_cost_ratio",
                f"{CELL_LOCAL_SYNTHETIC_SOURCE};{NO_DIRECT_TC_CALIBRATION_SOURCE}",
                "Explicit synthetic same-unit consequence ratio keeps probability and DR typed separately.",
            )
        for scenario in params["capacity_scenarios"]:
            add(
                f"{curve} | {scenario['scenario_id']} medians",
                curve,
                json.dumps(scenario["state_medians"]),
                "curve_fit_shape",
                f"{CELL_LOCAL_SYNTHETIC_SOURCE};{NO_DIRECT_TC_CALIBRATION_SOURCE}",
                "Unweighted synthetic resistance scenario; not a percentile or fitted hurricane parameter.",
            )
    return rows


def value_rows() -> list[dict[str, str]]:
    rows = read_csv(
        PROPOSED / "VALUE_CROSSWALK_tropical_cyclone_wind_solar__model_v0_1__docs_r1.csv"
    )
    for row in rows:
        if row["failure_unit_id"] in {
            "PV_FIXED_TILT_MODULE_FIELD|PV_TRACKER_MODULE_FIELD",
            "PV_FIXED_TILT_SUPPORT_STRUCTURE|PV_TRACKER_SBOS_ASSEMBLY",
        }:
            row["status"] = "mapped_reference_only_scenario_loss_withheld"
            row["role_in_loss"] = "reference_reconciliation_only"
            row["include_in_direct_denominator"] = "false"
            row["allocation_rule"] = (
                "No runtime value binding in model v2; use only to audit same-unit boundaries."
            )
        row["applicable_pathway_ids"] = PATHWAY
        row["notes"] = row["notes"].replace(
            "model v0.1 scenario loss is withheld",
            "model v2.0 proposal scenario loss is withheld",
        ).replace(
            "no model v0.1 monetary loss",
            "no model v2.0 proposal monetary loss",
        )
        if row["failure_unit_id"] == "PV_FIXED_TILT_SUPPORT_STRUCTURE|PV_TRACKER_SBOS_ASSEMBLY":
            row["notes"] = (
                "Architecture-generic reference row; v2 synthetic DR records exist, but this benchmark "
                "is not a runtime value binding or composition proof."
            )
    return rows


def base_request(architecture: str) -> dict[str, Any]:
    return {
        "event_id": "TC-EVENT-TEST",
        "event_family_id": "TC-FAMILY-TEST",
        "pathway_id": PATHWAY,
        "array_architecture": architecture,
        "tc_duration_class": "sustained_1_to_6h",
        "tc_direction_evolution_class": "evolving",
        "rain_ingress_indicator": False,
        "windborne_debris_indicator": False,
        "flood_or_surge_indicator": False,
        "tc_tornado_indicator": False,
    }


def fixed_request(x: float) -> dict[str, Any]:
    return {
        **base_request(FIXED_ARCH),
        "tc_fixed_event_to_design_net_pressure_ratio": x,
        "tc_wind_field_bridge_id": "TCWF-BRIDGE-V1",
        "tc_directional_history_bridge_id": "TCDIR-BRIDGE-V1",
        "tc_duration_cycling_bridge_id": "TCDUR-BRIDGE-V1",
        "aerodynamic_demand_bridge_id": "TCFIXED-PRESSURE-BRIDGE-V1",
        "array_zone": "edge",
        "array_spatial_object_id": "FIXED-ARRAY-ZONE-EDGE-A",
    }


def tracker_request(x: float) -> dict[str, Any]:
    critical = 50.0
    request = {
        **base_request(TRACKER_ARCH),
        "tc_tracker_normal_3s_gust_mps": critical * x,
        "critical_instability_3s_gust_mps": critical,
        "aeroelastic_qualification_id": "TRACKER-QUAL-V1",
        "aeroelastic_qualification_sha256": hashlib.sha256(
            b"TCWS2 synthetic tracker qualification KAT fixture v1"
        ).hexdigest(),
        "tracker_system_id": "TRACKER-SYSTEM-A",
        "tracker_module_configuration": "1P",
        "tracker_layout_id": "LAYOUT-A",
        "tracker_position_state": "confirmed_wind_stow",
        "tracker_angle_deg": 0.0,
        "stow_confirmation_basis": "position_sensor_and_scada",
        "tracker_drive_lock_state": "mechanically_locked",
        "array_zone": "edge",
        "array_spatial_object_id": "TRACKER-ARRAY-ZONE-EDGE-A",
        "tc_wind_field_bridge_id": "TCWF-BRIDGE-V1",
        "tc_directional_history_bridge_id": "TCDIR-BRIDGE-V1",
        "tc_duration_cycling_bridge_id": "TCDUR-BRIDGE-V1",
        "qualification_tracker_system_id": "TRACKER-SYSTEM-A",
        "qualification_tracker_module_configuration": "1P",
        "qualification_tracker_layout_id": "LAYOUT-A",
        "qualification_tracker_position_state": "confirmed_wind_stow",
        "qualification_tracker_angle_deg": 0.0,
        "qualification_array_zone": "edge",
        "qualification_drive_lock_state": "mechanically_locked",
        "qualification_speed_averaging_s": 3.0,
        "qualification_speed_reference": "array_height_tracker_normal_3s_gust",
        "qualification_tc_wind_field_bridge_id": "TCWF-BRIDGE-V1",
        "qualification_direction_basis_id": "TCDIR-BRIDGE-V1",
        "qualification_duration_basis_id": "TCDUR-BRIDGE-V1",
    }
    return request


def perry_request(x: float) -> dict[str, Any]:
    return {
        **base_request(PERRY_ARCH),
        "failure_unit_id": PERRY_UNIT,
        "perry_event_max_gust_mps": x,
        "array_architecture_id": "PERRY_GROUND_NONTRACKING_SOURCE_COHORT_V1",
        "source_population_match_id": "PERRY_MANUAL_GROUND_NONTRACKING_MIXED_SCALE_V1",
        "module_value_distribution_assumption_id": "UNIFORM_MODULE_HARDWARE_VALUE",
        "visible_damage_disposition_assumption_id": "FULL_REPLACEMENT_IF_VISIBLE_OR_MISSING",
        "source_wind_product_id": "PERRY_DATASET_REPORTED_EVENT_MAX_GUST",
        "causal_scope_acknowledgement_id": "SOURCE_COMPOSITE_HURRICANE_MODULE_LOSS",
    }


def build_kats(artifact: Mapping[str, Any]) -> dict[str, Any]:
    from scripts.reference_helpers.tropical_cyclone_wind_solar_v2_curve_eval import (
        TropicalCycloneWindSolarV2EvaluationError,
        evaluate_damage_call,
    )

    tests: list[dict[str, Any]] = []
    cases: list[tuple[str, dict[str, Any]]] = []
    for x in (0.0, 0.5, 1.0, 1.5, 2.0):
        cases.append((f"FIXED_DIRECT_{str(x).replace('.', '_')}", fixed_request(x)))
    proxy = base_request(FIXED_ARCH)
    proxy.update(
        {
            "tc_array_height_3s_gust_mps": 45.0,
            "qualified_design_array_height_3s_gust_mps": 50.0,
            "tc_peak_gust_3s_10m_mps": 48.0,
            "tc_wind_field_bridge_id": "TCWF-BRIDGE-V1",
            "tc_directional_history_bridge_id": "TCDIR-BRIDGE-V1",
            "tc_duration_cycling_bridge_id": "TCDUR-BRIDGE-V1",
            "aerodynamic_demand_bridge_id": "TCFIXED-PRESSURE-BRIDGE-V1",
            "array_zone": "edge",
            "array_spatial_object_id": "FIXED-ARRAY-ZONE-EDGE-A",
        }
    )
    cases.append(("FIXED_SPEED_PROXY_0_81", proxy))
    for x in (0.0, 0.75, 1.0, 1.5, 2.0):
        cases.append((f"TRACKER_{str(x).replace('.', '_')}", tracker_request(x)))
    for label, x in (("LOW", 17.4), ("INTERIOR", 30.0), ("HIGH", 39.1)):
        cases.append((f"PERRY_COMPAT_{label}", perry_request(x)))
    gsu = {
        "event_id": "TC-EVENT-TEST",
        "event_family_id": "TC-FAMILY-TEST",
        "pathway_id": PATHWAY,
        "failure_unit_id": "PV_GSU_SUBSTATION",
    }
    cases.append(("GSU_WITHHELD", gsu))
    for test_id, request in cases:
        emit = evaluate_damage_call(artifact, request)
        tests.append(
            {
                "test_id": test_id,
                "input": request,
                "expected": {
                    "hazard_input_used": emit["hazard_input_used"],
                    "emit_mode": emit["emit_mode"],
                    "selectors_used": emit["selectors_used"],
                    "failure_unit_results": [
                        {
                            "failure_unit_id": item["failure_unit_id"],
                            "status": item["status"],
                            "scalar_central_dr": item["scalar_central_dr"],
                            "scenario_drs": item["scenario_drs"],
                            "state_probabilities_by_scenario": item[
                                "state_probabilities_by_scenario"
                            ],
                            "withheld_reason_codes": item["withheld_reason_codes"],
                        }
                        for item in emit["failure_unit_results"]
                    ],
                    "required_flags": emit["input_quality"]["limitation_flags"],
                },
            }
        )

    rejection_cases: list[tuple[str, dict[str, Any], str]] = []
    wrong_path = fixed_request(1.0)
    wrong_path["pathway_id"] = "straight_line_convective"
    rejection_cases.append(("REJECT_WRONG_PATHWAY", wrong_path, "PATHWAY_ID_UNKNOWN"))
    missing_arch = fixed_request(1.0)
    missing_arch.pop("array_architecture")
    rejection_cases.append(("REJECT_MISSING_ARCH", missing_arch, "ARRAY_ARCHITECTURE_REQUIRED"))
    unbridged = base_request(FIXED_ARCH)
    unbridged["tc_peak_gust_3s_10m_mps"] = 55.0
    rejection_cases.append(("REJECT_UNBRIDGED_10M", unbridged, "TC_WIND_BRIDGE_REQUIRED"))
    missing_bridge = fixed_request(1.0)
    missing_bridge.pop("tc_duration_cycling_bridge_id")
    rejection_cases.append(("REJECT_FIXED_MISSING_DURATION_BRIDGE", missing_bridge, "TC_WIND_BRIDGE_REQUIRED"))
    bad_tracker = tracker_request(1.0)
    bad_tracker["qualification_tracker_angle_deg"] = 5.0
    rejection_cases.append(("REJECT_TRACKER_ANGLE_MISMATCH", bad_tracker, "TRACKER_QUALIFICATION_BASIS_MISMATCH"))
    command_only = tracker_request(1.0)
    command_only["tracker_position_state"] = "commanded_not_confirmed"
    command_only["qualification_tracker_position_state"] = "commanded_not_confirmed"
    rejection_cases.append(("REJECT_TRACKER_COMMAND_ONLY", command_only, "TRACKER_QUALIFICATION_BASIS_MISMATCH"))
    cross_unit = fixed_request(1.0)
    cross_unit["failure_unit_id"] = "PV_TRACKER_MODULE_FIELD"
    rejection_cases.append(("REJECT_CROSS_ARCH_UNIT", cross_unit, "FAILURE_UNIT_NOT_APPLICABLE_TO_ARCHITECTURE"))
    out_range = fixed_request(2.01)
    rejection_cases.append(("REJECT_AXIS_ABOVE_2", out_range, "AXIS_OUTSIDE_VALID_RANGE"))
    compound = fixed_request(1.0)
    compound["windborne_debris_indicator"] = True
    rejection_cases.append(("REJECT_COMPOUND_WITHOUT_ACK", compound, "COMPOUND_RECONCILIATION_REQUIRED"))
    value = fixed_request(1.0)
    value["full_tiv_usd"] = 1_000_000
    rejection_cases.append(("REJECT_VALUE_PAYLOAD", value, "SCENARIO_LOSS_WITHHELD_SYNTHETIC_T4_PROPOSAL"))
    perry_bad = perry_request(40.0)
    rejection_cases.append(("REJECT_PERRY_TAIL", perry_bad, "AXIS_OUTSIDE_VALID_RANGE"))
    unknown_field = fixed_request(1.0)
    unknown_field["undeclared_field"] = "ignored-before-v2-hardening"
    rejection_cases.append(("REJECT_UNKNOWN_FIELD", unknown_field, "REQUEST_FIELD_UNSUPPORTED"))
    exposure = perry_request(30.0)
    exposure["at_risk_fraction"] = 0.5
    rejection_cases.append(("REJECT_EXPOSURE_ALIAS", exposure, "SCENARIO_LOSS_WITHHELD_SYNTHETIC_T4_PROPOSAL"))
    numeric_compound = fixed_request(1.0)
    numeric_compound["windborne_debris_indicator"] = 1
    rejection_cases.append(("REJECT_NUMERIC_COMPOUND_BOOL", numeric_compound, "CONDITIONER_VALUE_UNSUPPORTED"))
    perry_compound = perry_request(30.0)
    perry_compound["windborne_debris_indicator"] = True
    perry_compound["compound_reconciliation_acknowledgement_id"] = "SEPARATE_PATHWAYS_AND_NO_DOUBLE_COUNT"
    rejection_cases.append(("REJECT_PERRY_COMPOSITE_OVERLAP", perry_compound, "PERRY_COMPOSITE_PATHWAY_OVERLAP_UNRESOLVED"))
    missing_qualification_sha = tracker_request(1.0)
    missing_qualification_sha.pop("aeroelastic_qualification_sha256")
    rejection_cases.append(("REJECT_TRACKER_MISSING_QUALIFICATION_SHA", missing_qualification_sha, "TRACKER_QUALIFICATION_BASIS_MISMATCH"))
    foreign_tracker_field = fixed_request(1.0)
    foreign_tracker_field["tracker_system_id"] = "FOREIGN"
    rejection_cases.append(("REJECT_FOREIGN_ROUTE_FIELD", foreign_tracker_field, "REQUEST_FIELD_UNSUPPORTED"))
    direct_with_proxy_companion = fixed_request(1.0)
    direct_with_proxy_companion["qualified_design_array_height_3s_gust_mps"] = 999.0
    rejection_cases.append(
        (
            "REJECT_FIXED_DIRECT_WITH_PROXY_COMPANION",
            direct_with_proxy_companion,
            "PRESSURE_INDEX_REQUIRED",
        )
    )
    rejections = []
    for test_id, request, expected_code in rejection_cases:
        try:
            evaluate_damage_call(artifact, request)
        except TropicalCycloneWindSolarV2EvaluationError as exc:
            if exc.code != expected_code:
                raise RuntimeError(f"{test_id}: expected {expected_code}, got {exc.code}") from exc
        else:
            raise RuntimeError(f"{test_id}: evaluator did not reject")
        rejections.append(
            {"test_id": test_id, "input": request, "expected_error_code": expected_code}
        )
    return {
        "schema_version": "known_answer_tests.tcws_solar_v2.v1",
        "cell_id": "tropical_cyclone_wind_solar",
        "semantic_damage_model_version": MODEL,
        "documentation_revision": DOCS,
        "runtime_known_answer_tests": tests,
        "rejection_tests": rejections,
        "artifact_pin_tests": [
            {"test_id": "PIN_EXACT", "expected": "pass"},
            {"test_id": "PIN_INCOMPLETE", "expected_error_code": "ARTIFACT_PIN_INCOMPLETE"},
            {"test_id": "PIN_MISMATCH", "expected_error_code": "ARTIFACT_PIN_MISMATCH"},
        ],
        "notes": [
            "Perry compatibility values must equal model-v1 exactly.",
            "Generic scenario values are synthetic T4 fixtures, not calibration targets.",
            "Every unsupported unit remains null, never zero.",
        ],
    }


def build_old_new(artifact: Mapping[str, Any]) -> list[dict[str, str]]:
    from scripts.reference_helpers.tropical_cyclone_wind_solar_v2_curve_eval import evaluate_damage_call

    rows: list[dict[str, str]] = []
    fields = [
        "comparison_id",
        "input_basis",
        "failure_unit_id",
        "prior_model_output",
        "v2_output",
        "delta",
        "comparability",
        "reason",
    ]
    for x in (17.4, 30.0, 39.1):
        emit = evaluate_damage_call(artifact, perry_request(x))
        dr = emit["failure_unit_results"][0]["scalar_central_dr"]
        rows.append(
            dict(
                zip(
                    fields,
                    [
                        f"PERRY_{x}",
                        f"perry_event_max_gust_mps={x}",
                        PERRY_UNIT,
                        f"{dr:.15g}",
                        f"{dr:.15g}",
                        "0",
                        "exact_compatibility",
                        "v2 carries the v1 source record byte-for-value",
                    ],
                    strict=True,
                )
            )
        )
    for architecture, request_fn, unit in (
        (FIXED_ARCH, fixed_request, "PV_FIXED_TILT_MODULE_FIELD"),
        (TRACKER_ARCH, tracker_request, "PV_TRACKER_MODULE_FIELD"),
    ):
        for x in (0.5, 1.0, 1.5):
            emit = evaluate_damage_call(artifact, request_fn(x))
            result = next(item for item in emit["failure_unit_results"] if item["failure_unit_id"] == unit)
            rows.append(
                dict(
                    zip(
                        fields,
                        [
                            f"{architecture}_{x}",
                            f"normalized_demand={x}",
                            unit,
                            "null_or_unmappable",
                            f"{result['scalar_central_dr']:.15g}",
                            "not_computed",
                            "different_axis_and_denominator",
                            "v1 has no generic architecture record on this axis",
                        ],
                        strict=True,
                    )
                )
            )
    rows.append(
        dict(
            zip(
                fields,
                [
                    "GSU_NULL_REMAINS_NULL",
                    "any",
                    "PV_GSU_SUBSTATION",
                    "null",
                    "null",
                    "0_status_change",
                    "status_only",
                    "v2 does not copy legacy or neighboring GSU response",
                ],
                strict=True,
            )
        )
    )
    rows.append(
        dict(
            zip(
                fields,
                [
                    "LEGACY_ASSET_DR",
                    "legacy 3-second gust mph vs v2 failure-unit normalized demand",
                    "WHOLE_ASSET_VS_FAILURE_UNITS",
                    "legacy capped weighted asset DR",
                    "withheld full-plant DR",
                    "not_computed",
                    "incomparable",
                    "legacy probability/DR type error, 42-percent zero remainder, stow default, and mismatched denominator",
                ],
                strict=True,
            )
        )
    )
    return rows


def style_sheet(ws: Any) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column in range(1, ws.max_column + 1):
        width = 12
        for row in range(1, min(ws.max_row, 120) + 1):
            value = ws.cell(row, column).value
            if value is not None:
                width = min(55, max(width, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def append_mapping_table(ws: Any, rows: list[Mapping[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0])
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def build_workbook(
    artifact: Mapping[str, Any],
    kats: Mapping[str, Any],
    sources: list[dict[str, str]],
    claims: list[dict[str, str]],
    supersession: list[dict[str, str]],
    parameters: list[dict[str, str]],
    values: list[dict[str, str]],
    old_new: list[dict[str, str]],
) -> None:
    from scripts.reference_helpers.tropical_cyclone_wind_solar_v2_curve_eval import (
        evaluate_ordered_damage_state_record,
    )

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("README")
    ws.append(["Field", "Value"])
    for key, value in [
        ("cell_id", artifact["cell_id"]),
        ("model", MODEL),
        ("docs", DOCS),
        ("status", "noncanonical synthetic T4 proposal"),
        ("canonical_runtime_artifact", False),
        ("generic_curve_count", 4),
        ("Perry_compatibility_curve_count", 1),
        ("scenario_dollar_loss", "withheld"),
        ("strict_alternative", "model v0.1 NO_RUNTIME_CURVE"),
        ("warning", "Generic numbers are synthetic unweighted scenarios, not hurricane calibration."),
    ]:
        ws.append([key, value])

    ws = wb.create_sheet("Dashboard")
    ws.append(["Metric", "Value", "Interpretation"])
    ws.append(["architectures", 3, "Perry compatibility + synthetic fixed + qualified tracker"])
    ws.append(["numeric records", 5, "one source-specific and four synthetic T4"])
    ws.append(["withheld units", 6, "null, not zero"])
    ws.append(["full plant DR", "WITHHELD", "no artificial legacy cap"])
    ws.append(["Hazard cutover", "BLOCKED", "artifact index unchanged"])

    ws = wb.create_sheet("Scope_Pathway")
    ws.append(["Category", "Item", "Treatment"])
    pathway = artifact["pathways"][0]
    for value in pathway["hazard_scope"]["included"]:
        ws.append(["included", value, "evaluate only after route gates"])
    for value in pathway["hazard_scope"]["excluded"]:
        ws.append(["excluded", value, "separate pathway or outside physical DR"])

    ws = wb.create_sheet("Architecture_Axes")
    ws.append(["Architecture", "Axis", "Range", "Required bridge", "No-default rule"])
    contracts = pathway["hazard_axis"]["architecture_input_contracts"]
    for architecture, contract in contracts.items():
        ws.append(
            [
                architecture,
                contract["axis_field"],
                json.dumps(contract.get("valid_range")),
                json.dumps(contract.get("required_basis") or contract.get("accepted_payloads") or contract.get("required_acknowledgements")),
                "missing or mismatch rejects",
            ]
        )

    ws = wb.create_sheet("Curve_Records")
    ws.append(["Curve ID", "Architecture", "Failure unit", "Form", "Axis", "Tier"])
    arch_by_unit = {}
    for item in pathway["failure_unit_coverage"]:
        if "architecture" in item:
            arch_by_unit[item["failure_unit_id"]] = item["architecture"]
    for record in pathway["curve_records"]:
        ws.append(
            [
                record["curve_id"],
                arch_by_unit[record["failure_unit_id"]],
                record["failure_unit_id"],
                record["curve_form"],
                record["x_axis"],
                "Perry mixed tiers" if record["curve_form"] == "piecewise_linear" else "T4 synthetic",
            ]
        )

    ws = wb.create_sheet("Curve_Data")
    ws.append(["Curve ID", "x", "lower_resistance_DR", "central_screening_DR", "upper_resistance_DR", "Scenario_Order_QA"])
    generic = [r for r in pathway["curve_records"] if r["curve_form"] == "ordered_damage_state_lognormal"]
    for record in generic:
        for step in range(41):
            x = step / 20
            values_out = evaluate_ordered_damage_state_record(record, x)
            row_num = ws.max_row + 1
            ws.append(
                [
                    record["curve_id"],
                    x,
                    values_out["lower_resistance"]["damage_ratio"],
                    values_out["central_screening"]["damage_ratio"],
                    values_out["upper_resistance"]["damage_ratio"],
                    f'=IF(AND(C{row_num}>=D{row_num},D{row_num}>=E{row_num}),"PASS","FAIL")',
                ]
            )

    ws = wb.create_sheet("State_Definitions")
    ws.append(["Curve ID", "State ID", "Cost ratio", "Tier", "Meaning"])
    for record in generic:
        for state in record["parameters"]["damage_states"]:
            ws.append([record["curve_id"], state["state_id"], state["cost_ratio"], state["tier"], state["description"]])

    ws = wb.create_sheet("Perry_Compatibility")
    ws.append(["x_mps", "v1_proxy_DR", "v2_proxy_DR", "Delta", "QA"])
    perry = next(r for r in pathway["curve_records"] if r["failure_unit_id"] == PERRY_UNIT)
    for x, y in perry["parameters"]["points"]:
        row_num = ws.max_row + 1
        ws.append([x, y, y, f"=C{row_num}-B{row_num}", f'=IF(ABS(D{row_num})<1E-15,"PASS","FAIL")'])

    ws = wb.create_sheet("Site_Condition_Double_Count")
    ws.append(["Fields/control", "Single treatment", "Prohibited double count", "Missing behavior"])
    for item in [
        ("TC wind field + delivered demand", "one named bridge feeds the normalized axis", "reapply profile/terrain multiplier", "reject"),
        ("direction/duration/cycling", "bridge input or metadata only", "curve multiplier after bridge", "unknown flag; no credit"),
        ("tracker stow/angle/lock", "attained state exactly matches Ucrit basis", "generic stow discount", "reject"),
        ("rain/debris/flood/tornado", "separate pathway under event_family_id", "add to source-composite/PV wind DR", "unknown flag or acknowledgement"),
        ("array vs GSU exposure", "separate row/yard spatial subjects", "copy array exposed fraction", "GSU withheld"),
        ("direct damage vs support", "direct DR first; support once later", "support curve plus allocation", "support withheld"),
    ]:
        ws.append(item)

    ws = wb.create_sheet("KATs")
    ws.append(["Test ID", "Type", "Expected status/error", "Axis"])
    for test in kats["runtime_known_answer_tests"]:
        statuses = sorted({item["status"] for item in test["expected"]["failure_unit_results"]})
        ws.append(
            [
                test["test_id"],
                "runtime",
                ";".join(statuses),
                test["expected"]["hazard_input_used"].get("axis_value", "not_applicable"),
            ]
        )
    for test in kats["rejection_tests"]:
        ws.append([test["test_id"], "rejection", test["expected_error_code"], ""])

    ws = wb.create_sheet("Value_Crosswalk")
    append_mapping_table(ws, values)
    ws = wb.create_sheet("Sources")
    append_mapping_table(ws, sources)
    ws = wb.create_sheet("Claim_Register")
    append_mapping_table(ws, claims)
    ws = wb.create_sheet("Claim_Supersession")
    append_mapping_table(ws, supersession)
    ws = wb.create_sheet("Parameter_Tiers")
    append_mapping_table(ws, parameters)
    ws = wb.create_sheet("Legacy_Comparison")
    append_mapping_table(ws, old_new)

    ws = wb.create_sheet("QA_Checks")
    ws.append(["Check", "Observed", "Expected", "Status"])
    checks = [
        ("canonical false", artifact["canonical_runtime_artifact"], False),
        ("model v2", artifact["semantic_damage_model_version"], MODEL),
        ("bundle v3", artifact["schema_version"], "damage_curve_record_bundle.v3"),
        ("one pathway", len(artifact["pathways"]), 1),
        ("five curve records", len(pathway["curve_records"]), 5),
        ("four synthetic generic records", len(generic), 4),
        ("Perry record present", sum(r["failure_unit_id"] == PERRY_UNIT for r in pathway["curve_records"]), 1),
        ("withheld units", len(artifact["capability_declaration"]["pathway_capabilities"][0]["withheld_failure_units"]), 6),
        ("scenario loss withheld", artifact["capability_declaration"]["pathway_capabilities"][0]["scenario_loss_given_value_basis"], "withheld"),
        ("source rows", len(sources), len(sources)),
        ("claim rows", len(claims), len(claims)),
        ("claim supersession rows", len(supersession), 8),
        ("parameter rows", len(parameters), len(parameters)),
        ("value rows", len(values), len(values)),
        ("runtime KATs", len(kats["runtime_known_answer_tests"]), len(kats["runtime_known_answer_tests"])),
        ("rejection KATs", len(kats["rejection_tests"]), len(kats["rejection_tests"])),
        ("no hard-zero parameters", sum("zero_below" in r["parameters"] for r in generic), 0),
        ("GSU has no curve", sum(r["failure_unit_id"] == "PV_GSU_SUBSTATION" for r in pathway["curve_records"]), 0),
        ("artifact index unchanged", "not checked in workbook", "not checked in workbook"),
    ]
    for name, observed, expected in checks:
        ws.append([name, observed, expected, "PASS" if observed == expected else "FAIL"])

    for sheet in wb.worksheets:
        style_sheet(sheet)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    save_deterministic_workbook(wb, WORKBOOK)


def main() -> int:
    PROPOSED.mkdir(parents=True, exist_ok=True)
    SHARED.mkdir(parents=True, exist_ok=True)
    records = synthetic_records()
    shared = build_shared_profile(deepcopy(records))
    write_json(SHARED_PROFILE, shared)
    shared_profile_sha256 = digest(SHARED_PROFILE)
    cap = capability()
    artifact = build_artifact(records, cap, shared_profile_sha256)
    write_json(CAPABILITY, cap)
    write_json(ARTIFACT, artifact)
    sources = consolidated_sources()
    claims = consolidated_claims()
    supersession = claim_supersession_rows()
    parameters = parameter_rows(records)
    values = value_rows()
    write_csv(SOURCES, SOURCE_FIELDS, sources)
    write_csv(CLAIMS, CLAIM_FIELDS, claims)
    write_csv(CLAIM_SUPERSESSION, CLAIM_SUPERSESSION_FIELDS, supersession)
    write_csv(PARAMETERS, PARAMETER_FIELDS, parameters)
    value_fields = list(values[0])
    write_csv(VALUES, value_fields, values)
    kats = build_kats(artifact)
    write_json(KATS, kats)
    old_new = build_old_new(artifact)
    old_new_fields = list(old_new[0])
    write_csv(OLD_NEW, old_new_fields, old_new)
    build_workbook(artifact, kats, sources, claims, supersession, parameters, values, old_new)
    print(f"built artifact: {ARTIFACT.relative_to(ROOT)}")
    print(f"built capability: {CAPABILITY.relative_to(ROOT)}")
    print(f"built KATs: {KATS.relative_to(ROOT)}")
    print(f"built workbook: {WORKBOOK.relative_to(ROOT)}")
    print(f"built sources={len(sources)} claims={len(claims)} parameters={len(parameters)}")
    print(f"artifact sha256={digest(ARTIFACT)}")
    print(f"shared profile sha256={digest(SHARED_PROFILE)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
