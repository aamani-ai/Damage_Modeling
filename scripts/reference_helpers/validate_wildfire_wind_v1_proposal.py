#!/usr/bin/env python3
"""Validate wildfire_wind model v1.0/docs r1 and its fail-closed evaluator."""

from __future__ import annotations

import csv
import json
from copy import deepcopy
from pathlib import Path

from jsonschema import Draft202012Validator
from openpyxl import load_workbook
from referencing import Registry, Resource

from wildfire_wind_v1_curve_eval import WildfireWindEvaluationError, evaluate_damage_call


REPO = Path(__file__).resolve().parents[2]
PROPOSED = REPO / "docs/cells/wildfire_wind/proposed"
ARTIFACT = PROPOSED / "wildfire_wind__model_v1_0__docs_r1__curve_artifact.json"
CAPABILITY = PROPOSED / "wildfire_wind__model_v1_0__docs_r1__capability.json"
KATS = PROPOSED / "known_answer_tests_wildfire_wind__model_v1_0__docs_r1.json"
WORKBOOK = PROPOSED / "damage_curve_records_wildfire_wind__model_v1_0__docs_r1.xlsx"
EXPECTED = {
    "WT_PAD_ELECTRICAL": [0.0, 0.001, 0.006, 0.03, 0.12, 0.35, 0.70],
    "WT_GSU_PROTECTION_CONTROL_DC": [0.0, 0.004, 0.02, 0.08, 0.25, 0.60, 0.90],
}
EXPECTED_DOCS = [
    "README_wildfire_wind__model_v1_0__docs_r1.md",
    "CHANGE_CLASSIFICATION_wildfire_wind__model_v1_0__docs_r1.md",
    "DEEP_RESEARCH_AND_DECISION_MEMO_wildfire_wind__model_v1_0__docs_r1.md",
    "SEVEN_STEP_AUDIT_wildfire_wind__model_v1_0__docs_r1.md",
    "wildfire_wind_curve_derivation_dossier__model_v1_0__docs_r1.md",
    "wildfire_wind_damage_code_metadata_spec__model_v1_0__docs_r1.md",
    "PRESSURE_TEST_wildfire_wind__model_v1_0__docs_r1.md",
    "PROMOTION_GATE_MATRIX_wildfire_wind__model_v1_0__docs_r1.md",
    "WORKBOOK_SHEET_MANIFEST_wildfire_wind__model_v1_0__docs_r1.md",
    "VALIDATION_REPORT_wildfire_wind__model_v1_0__docs_r1.md",
]


class ValidationFailure(AssertionError):
    pass


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ValidationFailure(message)


def load(path: Path):
    return json.loads(path.read_text())


def base_request() -> dict:
    return {
        "event_id": "WW-VALIDATION-EVENT",
        "event_family_id": "WW-VALIDATION-FAMILY",
        "pathway_id": "wildfire_thermal_attack",
        "failure_unit_id": "WT_PAD_ELECTRICAL",
        "source_wildfire_product_id": "USFS_RDS_2016_0034_3_270M",
        "screening_assumption_set_id": "WW_T4_PARTIAL_ELECTRICAL_SCREENING_2026_08_08",
        "conditional_flame_length_class_state": 3,
    }


def main() -> None:
    artifact, capability, kats = load(ARTIFACT), load(CAPABILITY), load(KATS)
    for filename in EXPECTED_DOCS:
        require((PROPOSED / filename).is_file(), f"missing governed document {filename}")
    bundle_schema = load(REPO / "docs/contracts/schemas/curve_artifact_bundle.v3.schema.json")
    capability_schema = load(REPO / "docs/contracts/schemas/capability_declaration.v3.schema.json")
    emit_schema = load(REPO / "docs/contracts/schemas/damage_emit.v2.schema.json")
    registry = Registry().with_resources([
        (bundle_schema["$id"], Resource.from_contents(bundle_schema)),
        (capability_schema["$id"], Resource.from_contents(capability_schema)),
        (emit_schema["$id"], Resource.from_contents(emit_schema)),
    ])
    Draft202012Validator(bundle_schema, registry=registry).validate(artifact)
    Draft202012Validator(capability_schema, registry=registry).validate(capability)
    Draft202012Validator(emit_schema, registry=registry).validate(evaluate_damage_call(artifact, base_request()))
    require(artifact["capability_declaration"] == capability, "embedded capability drift")
    require(artifact["canonical_runtime_artifact"] is False, "proposal became canonical")
    require(artifact["package_inclusion_status"] == "not_included", "proposal entered package")
    require(artifact["semantic_damage_model_version"] == "model v1.0", "model identity")
    require(len(artifact["pathways"]) == 1, "only thermal screening pathway may carry curves")
    pathway = artifact["pathways"][0]
    require(pathway["pathway_id"] == "wildfire_thermal_attack", "pathway identity")
    records = {r["failure_unit_id"]: r for r in pathway["curve_records"]}
    require(set(records) == set(EXPECTED), "supported unit set changed")
    for unit, expected in EXPECTED.items():
        record = records[unit]
        actual = [float(dr) for _, dr in record["parameters"]["points"]]
        require(actual == expected, f"{unit}: Tier-4 ordinates changed")
        require([int(state) for state, _ in record["parameters"]["points"]] == list(range(7)), f"{unit}: state map")
        require(all(b >= a for a, b in zip(actual, actual[1:])), f"{unit}: nonmonotone")
        require("CELL_LOCAL_T4_ORDINATES" in record["metadata_flags"], f"{unit}: assumption flag")

    tolerance = float(kats["absolute_tolerance"])
    for test in kats["formula_known_answer_tests"]:
        emit = evaluate_damage_call(artifact, test["input"])
        result = emit["failure_unit_results"][0]
        require(result["status"] == test["expected"]["status"], f"{test['test_id']}: status")
        require(result["curve_id"] == test["expected"]["curve_id"], f"{test['test_id']}: curve")
        require(abs(float(result["scalar_central_dr"]) - float(test["expected"]["failure_unit_damage_ratio"])) <= tolerance, f"{test['test_id']}: DR")

    for test in kats["negative_contract_tests"]:
        req = base_request()
        if "mutate" in test: req.update(test["mutate"])
        if "remove" in test: req.pop(test["remove"], None)
        if "expected_error" in test:
            try:
                evaluate_damage_call(artifact, req)
            except WildfireWindEvaluationError as exc:
                require(exc.code == test["expected_error"], f"{test['test_id']}: {exc.code}")
            else:
                raise ValidationFailure(f"{test['test_id']}: expected error")
        else:
            result = evaluate_damage_call(artifact, req)["failure_unit_results"][0]
            require(result["status"] == test["expected_status"], f"{test['test_id']}: status")
            require(test["expected_reason"] in result["withheld_reason_codes"], f"{test['test_id']}: reason")

    # The evaluator must reject semantic corruption even if the JSON remains shape-valid.
    for field, value in [("x_axis", "wrong"), ("valid_range", [0, 7]), ("selector_match", {})]:
        broken = deepcopy(artifact)
        broken["pathways"][0]["curve_records"][0][field] = value
        try:
            evaluate_damage_call(broken, base_request())
        except WildfireWindEvaluationError as exc:
            require(exc.code == "CURVE_PAYLOAD_INVALID", f"corrupt {field}: wrong error")
        else:
            raise ValidationFailure(f"evaluator accepted corrupt {field}")

    for filename, minimum_rows in [
        ("SOURCE_REGISTER_wildfire_wind__model_v1_0__docs_r1.csv", 7),
        ("CLAIM_PARAMETER_REGISTER_wildfire_wind__model_v1_0__docs_r1.csv", 6),
        ("PARAMETER_TIER_TABLE_wildfire_wind__model_v1_0__docs_r1.csv", 5),
        ("VALUE_CROSSWALK_wildfire_wind__model_v1_0__docs_r1.csv", 3),
    ]:
        with (PROPOSED / filename).open(newline="") as fh: rows = list(csv.DictReader(fh))
        require(len(rows) >= minimum_rows, f"{filename}: incomplete")

    wb = load_workbook(WORKBOOK, read_only=False, data_only=False)
    require(wb.sheetnames == ["README", "Curves", "KATs", "Sources", "Claims", "Tiers", "Values"], "workbook sheets")
    require(wb["Curves"].max_row == 18, "workbook curve rows")
    require(wb["README"]["B7"].value is False, "workbook canonical status")

    print(f"PASS wildfire_wind model v1.0/docs r1 proposal: {len(kats['formula_known_answer_tests'])} formula KATs, {len(kats['negative_contract_tests'])} negative KATs")


if __name__ == "__main__":
    main()
