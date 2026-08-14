#!/usr/bin/env python3
"""Validate the TC-wind × canonical Wind Farm model-v1.1 proxy proposal."""

from __future__ import annotations

import csv
import json
import math
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from tropical_cyclone_wind_wind_curve_eval import (
    PROXY_ARCHETYPE_ID,
    PROXY_ASSET_PROFILE_ID,
    PROXY_COVERED_VALUE_SHARE,
    PROXY_FAILURE_UNIT,
    PROXY_POLICY_ID,
    PROXY_VALUE_BASIS_ID,
    TropicalCycloneWindEvaluationError,
    evaluate_damage_call,
)


ROOT = Path(__file__).resolve().parents[2]
CURRENT = ROOT / "docs/cells/tropical_cyclone_wind_wind/current"
PROPOSED = ROOT / "docs/cells/tropical_cyclone_wind_wind/proposed"
CURRENT_ARTIFACT = CURRENT / "tropical_cyclone_wind_wind__model_v1_0__docs_r1__curve_artifact.json"
ARTIFACT = PROPOSED / "tropical_cyclone_wind_wind__model_v1_1__docs_r1__curve_artifact.json"
CAPABILITY = PROPOSED / "tropical_cyclone_wind_wind__model_v1_1__docs_r1__capability.json"
KATS = PROPOSED / "known_answer_tests_tropical_cyclone_wind_wind__model_v1_1__docs_r1.json"
VALUES = PROPOSED / "VALUE_CROSSWALK_tropical_cyclone_wind_wind__model_v1_1__docs_r1.csv"
WORKBOOK = PROPOSED / "damage_curve_records_tropical_cyclone_wind_wind__model_v1_1__docs_r1.xlsx"
BUNDLE_SCHEMA = ROOT / "docs/contracts/schemas/curve_artifact_bundle.v3.schema.json"
CAPABILITY_SCHEMA = ROOT / "docs/contracts/schemas/capability_declaration.v3.schema.json"
EMIT_SCHEMA = ROOT / "docs/contracts/schemas/damage_emit.v2.schema.json"


def load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _result(emit: Mapping[str, Any], unit_id: str) -> Mapping[str, Any]:
    matches = [
        item for item in emit["failure_unit_results"] if item["failure_unit_id"] == unit_id
    ]
    require(len(matches) == 1, f"expected one result for {unit_id}")
    return matches[0]


def validate_structure(
    current: Mapping[str, Any],
    artifact: Mapping[str, Any],
    capability: Mapping[str, Any],
) -> None:
    require(artifact["schema_version"] == "damage_curve_record_bundle.v3", "bundle schema changed")
    require(artifact["semantic_damage_model_version"] == "model v1.1", "model version mismatch")
    require(artifact["documentation_revision"] == "docs r1", "docs revision mismatch")
    require(artifact["canonical_runtime_artifact"] is False, "proposal became current")
    require(artifact["promotion_status"] == "proposed", "proposal status mismatch")
    require(capability == artifact["capability_declaration"], "embedded/external capability drift")
    current_records = current["pathways"][0]["curve_records"]
    proposed_records = artifact["pathways"][0]["curve_records"]
    require(proposed_records[: len(current_records)] == current_records, "v1.0 records moved")
    require(len(proposed_records) == len(current_records) + 1, "expected one proxy record")
    proxy = proposed_records[-1]
    source = current_records[-1]
    require(proxy["failure_unit_id"] == PROXY_FAILURE_UNIT, "proxy unit changed")
    require(proxy["parameters"] == source["parameters"], "proxy parameters differ from 3.3 MW source")
    require(proxy["x_axis"] == source["x_axis"], "proxy axis changed")
    require(proxy["selector_match"]["turbine_archetype_id"] == PROXY_ARCHETYPE_ID, "proxy archetype changed")
    contract = artifact["derivation_rationale"]["owner_approved_proxy_contract"]
    require(contract["proxy_policy_id"] == PROXY_POLICY_ID, "proxy policy changed")
    require(contract["canonical_asset_profile_id"] == PROXY_ASSET_PROFILE_ID, "asset profile changed")
    require(contract["covered_value_basis_id"] == PROXY_VALUE_BASIS_ID, "value basis changed")
    require(contract["numeric_rule"].endswith("no 5/3.3 scaling"), "no-scaling rule missing")
    completion = contract["screening_completion_rule"]
    require(completion["transition_band_kmh"] == [90, 108], "transition band changed")
    require(completion["above_ceiling_treatment"] == "cap_at_max_dr_with_explicit_flag", "ceiling rule changed")


def validate_reproduction(
    current: Mapping[str, Any], artifact: Mapping[str, Any], kats: Mapping[str, Any]
) -> int:
    count = 0
    for test in kats["v1_0_reproduction_tests"]:
        old_emit = evaluate_damage_call(current, test["input"])
        new_emit = evaluate_damage_call(artifact, test["input"])
        unit_id = test["input"]["failure_unit_id"]
        old_result = _result(old_emit, unit_id)
        new_result = _result(new_emit, unit_id)
        require(old_result == new_result, f"v1.0 reproduction failed: {test['test_id']}")
        expected = test["expected"]
        require(new_result["status"] == expected["status"], f"status failed: {test['test_id']}")
        require(new_result["curve_id"] == expected["curve_id"], f"curve failed: {test['test_id']}")
        require(
            math.isclose(
                new_result["scalar_central_dr"],
                expected["failure_unit_damage_ratio"],
                rel_tol=0,
                abs_tol=kats["absolute_tolerance"],
            ),
            f"DR failed: {test['test_id']}",
        )
        count += 1
    return count


def validate_proxy(artifact: Mapping[str, Any], kats: Mapping[str, Any]) -> int:
    for test in kats["proxy_known_answer_tests"]:
        emit = evaluate_damage_call(artifact, test["input"])
        result = _result(emit, PROXY_FAILURE_UNIT)
        expected = test["expected"]
        require(result["status"] == "supported", f"proxy withheld: {test['test_id']}")
        require(result["curve_id"] == expected["curve_id"], f"proxy curve drift: {test['test_id']}")
        require(
            math.isclose(
                result["scalar_central_dr"],
                expected["failure_unit_damage_ratio"],
                rel_tol=0,
                abs_tol=kats["absolute_tolerance"],
            ),
            f"proxy DR drift: {test['test_id']}",
        )
        require(
            emit["exposure_used"]["covered_value_share_of_project_tiv"]
            == PROXY_COVERED_VALUE_SHARE,
            "covered share drift",
        )
        flags = set(result["metadata_flags"])
        require("NO_CAPACITY_RATIO_SCALING" in flags, "no-scaling flag missing")
        require("OWNER_APPROVED_SCREENING_PROXY" in flags, "proxy flag missing")
        boundary_flag = expected.get("boundary_flag")
        if boundary_flag:
            require(boundary_flag in flags, f"boundary flag missing: {test['test_id']}")
    return len(kats["proxy_known_answer_tests"])


def validate_negative(artifact: Mapping[str, Any], kats: Mapping[str, Any]) -> int:
    for test in kats["negative_contract_tests"]:
        try:
            evaluate_damage_call(artifact, test["input"])
        except TropicalCycloneWindEvaluationError as exc:
            require(exc.code == test["expected_error_code"], f"wrong error: {test['test_id']}")
        else:
            raise AssertionError(f"negative test passed unexpectedly: {test['test_id']}")
    return len(kats["negative_contract_tests"])


def validate_value(kats: Mapping[str, Any]) -> int:
    with VALUES.open(newline="") as handle:
        rows = list(csv.DictReader(handle))
    shares = {row["subsystem"]: float(row["share_of_project_tiv"]) for row in rows}
    require(math.isclose(shares["rotor"] + shares["nacelle"] + shares["tower"], 0.63), "covered shares do not sum")
    require(shares["foundation+substation+electrical+civil"] == 0.37, "uncovered share changed")
    for test in kats["value_crosswalk_tests"]:
        tiv = test["project_tiv_usd"]
        require(math.isclose(tiv * 0.63, test["expected_covered_value_usd"]), "covered cap KAT failed")
        require(math.isclose(tiv * 0.37, test["expected_uncovered_value_usd"]), "uncovered value KAT failed")
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    require(workbook.sheetnames == ["README", "Curve records", "Proxy KATs", "Value crosswalk"], "workbook topology changed")
    return len(kats["value_crosswalk_tests"])


def validate_schemas(artifact: Mapping[str, Any], capability: Mapping[str, Any]) -> str:
    try:
        from jsonschema import Draft202012Validator
        from referencing import Registry, Resource
    except ModuleNotFoundError:
        return "schemas skipped: jsonschema unavailable"
    schemas = [load(BUNDLE_SCHEMA), load(CAPABILITY_SCHEMA), load(EMIT_SCHEMA)]
    registry = Registry().with_resources(
        [(schema["$id"], Resource.from_contents(schema)) for schema in schemas]
    )
    Draft202012Validator(schemas[1], registry=registry).validate(capability)
    Draft202012Validator(schemas[0], registry=registry).validate(artifact)
    sample = evaluate_damage_call(artifact, load(KATS)["proxy_known_answer_tests"][0]["input"])
    Draft202012Validator(schemas[2], registry=registry).validate(sample)
    return "bundle v3, capability v3, and emit v2 passed"


def main() -> None:
    current = load(CURRENT_ARTIFACT)
    artifact = load(ARTIFACT)
    capability = load(CAPABILITY)
    kats = load(KATS)
    validate_structure(current, artifact, capability)
    reproduction = validate_reproduction(current, artifact, kats)
    proxy = validate_proxy(artifact, kats)
    negative = validate_negative(artifact, kats)
    value = validate_value(kats)
    schema = validate_schemas(artifact, capability)
    print("PASS tropical_cyclone_wind_wind model v1.1 owner-approved proxy proposal")
    print(f"v1_0_reproduction={reproduction}")
    print(f"proxy_known_answers={proxy}")
    print(f"negative_contract_tests={negative}")
    print(f"value_crosswalk_tests={value}")
    print(schema)


if __name__ == "__main__":
    main()
