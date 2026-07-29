#!/usr/bin/env python3
"""Build the coverage-complete TC-wind x solar model-v2.1 proposal."""

from __future__ import annotations

import hashlib
import json
from copy import deepcopy
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook

import build_tropical_cyclone_wind_solar_v2_package as v20_builder
import tropical_cyclone_wind_solar_v2_1_curve_eval as evaluator


ROOT = Path(__file__).resolve().parents[2]
PROPOSED = ROOT / "docs/cells/tropical_cyclone_wind_solar/proposed"
MODEL = "model v2.1"
DOCS = "docs r1"
MODEL_STEM = "model_v2_1__docs_r1"
V20_ARTIFACT = PROPOSED / "tropical_cyclone_wind_solar__model_v2_0__docs_r1__curve_artifact.json"
V20_ARTIFACT_SHA256 = "06ee048096f3a54344e18e00cb8831a7a33910e61034f23fd1f4c33415658428"
ARTIFACT = PROPOSED / f"tropical_cyclone_wind_solar__{MODEL_STEM}__curve_artifact.json"
CAPABILITY = PROPOSED / f"tropical_cyclone_wind_solar__{MODEL_STEM}__capability.json"
KATS = PROPOSED / f"known_answer_tests_tropical_cyclone_wind_solar__{MODEL_STEM}.json"
SOURCES = PROPOSED / f"SOURCE_REGISTER_tropical_cyclone_wind_solar__{MODEL_STEM}.csv"
CLAIMS = PROPOSED / f"CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__{MODEL_STEM}.csv"
PARAMETERS = PROPOSED / f"PARAMETER_TIER_TABLE_tropical_cyclone_wind_solar__{MODEL_STEM}.csv"
VALUES = PROPOSED / f"VALUE_CROSSWALK_tropical_cyclone_wind_solar__{MODEL_STEM}.csv"
OLD_NEW = PROPOSED / f"OLD_VS_NEW_COMPARISON_tropical_cyclone_wind_solar__{MODEL_STEM}.csv"
CURVE_TABLE = PROPOSED / f"FULL_PLANT_SCREENING_CURVE_TABLE_tropical_cyclone_wind_solar__{MODEL_STEM}.csv"
WORKBOOK = PROPOSED / f"damage_curve_records_tropical_cyclone_wind_solar__{MODEL_STEM}.xlsx"
T4_SOURCES = [
    v20_builder.CELL_LOCAL_SYNTHETIC_SOURCE,
    v20_builder.NO_DIRECT_TC_CALIBRATION_SOURCE,
]


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def common_record(
    failure_unit_id: str,
    prefix: str,
    component_label: str,
    cost_ratios: list[float],
    central_medians: list[float],
) -> dict[str, Any]:
    if len(cost_ratios) != 4 or len(central_medians) != 3:
        raise ValueError("common records require four states and three exceedance medians")
    states = [
        (f"{prefix}_DS0_NO_DIRECT_DAMAGE", cost_ratios[0], []),
        (f"{prefix}_DS1_LOCALIZED_REPAIR", cost_ratios[1], [component_label]),
        (f"{prefix}_DS2_MAJOR_REPAIR", cost_ratios[2], [component_label]),
        (f"{prefix}_DS3_FULL_UNIT_REPLACEMENT", cost_ratios[3], [component_label]),
    ]
    scenarios = {
        "lower_resistance": [round(value * 0.8, 12) for value in central_medians],
        "central_screening": central_medians,
        "upper_resistance": [round(value * 1.25, 12) for value in central_medians],
    }
    return {
        "curve_id": f"TCWS21_{prefix}_SCREENING_T4_ORDERED_STATES",
        "pathway_id": evaluator.PATHWAY,
        "failure_unit_id": failure_unit_id,
        "curve_form": "ordered_damage_state_lognormal",
        "x_axis": "site_facility_tropical_cyclone_wind_demand_ratio",
        "y_axis": "failure_unit_damage_ratio",
        "parameters": {
            "beta_ln": 0.35,
            "damage_states": [
                {
                    "state_id": state_id,
                    "cost_ratio": cost_ratio,
                    "affected_subsystems": affected,
                    "description": (
                        f"{component_label} screening disposition; explicit Tier-4 coverage "
                        "assumption, not an observed or calibrated TC-wind consequence"
                    ),
                    "tier": "T4_placeholder_or_expert_judgment",
                    "source_ids": T4_SOURCES,
                }
                for state_id, cost_ratio, affected in states
            ],
            "capacity_scenarios": [
                {
                    "scenario_id": scenario_id,
                    "state_medians": medians,
                    "tier": "T4_placeholder_or_expert_judgment",
                    "source_ids": T4_SOURCES,
                    "interpretation": (
                        "unweighted screening resistance scenario on a qualified site-facility "
                        "demand/design ratio; not fitted and not a probabilistic confidence bound"
                    ),
                }
                for scenario_id, medians in scenarios.items()
            ],
        },
    }


def common_records() -> list[dict[str, Any]]:
    return [
        common_record(
            "PV_FOUNDATION",
            "FOUNDATION",
            "foundation and anchorage unit",
            [0.0, 0.05, 0.35, 1.0],
            [1.05, 1.40, 1.80],
        ),
        common_record(
            "PV_POWER_CONVERSION_AND_COLLECTION",
            "POWER_COLLECTION",
            "power-conversion and collection unit",
            [0.0, 0.10, 0.40, 1.0],
            [0.90, 1.25, 1.65],
        ),
        common_record(
            "PV_GSU_SUBSTATION",
            "GSU_SUBSTATION",
            "GSU and substation yard unit",
            [0.0, 0.08, 0.35, 1.0],
            [0.95, 1.30, 1.70],
        ),
        common_record(
            "PV_SCADA_COMMUNICATIONS",
            "SCADA",
            "SCADA, communications, controls, and sensor unit",
            [0.0, 0.15, 0.55, 1.0],
            [0.75, 1.05, 1.45],
        ),
        common_record(
            "PV_CIVIL_INFRA",
            "CIVIL",
            "wind-exposed civil infrastructure bucket",
            [0.0, 0.08, 0.35, 1.0],
            [0.80, 1.15, 1.60],
        ),
    ]


def build_capability(v20: Mapping[str, Any]) -> dict[str, Any]:
    capability = deepcopy(v20)
    pathway = capability["pathway_capabilities"][0]
    pathway.update(
        {
            "failure_unit_scalar_dr": "conditional",
            "scenario_loss_given_value_basis": (
                "supported_with_explicit_failure_unit_value_and_exposure_basis"
            ),
            "populated_emit_modes": ["scalar_mean", "state_ensemble"],
            "conditions": [
                "fixed or tracker array route supplies its qualified architecture-specific demand input",
                "common facility units supply a qualified site event/design wind-pressure ratio or complete gust-squared proxy",
                "full-plant calls explicitly select the named 2024-USD reference value profile and representative-site array exposure basis",
                "tracker calls prove attained state and exact-system qualification; commanded stow is insufficient",
                "compound rain, debris, surge, and tornado pathways are excluded from the wind-only DR and separately acknowledged",
                "scenario dollars require positive capacity_kwdc; normalized physical DR and loss per kWdc do not",
            ],
            "limitation_flags": [
                "SCREENING_ENGINEERING_PROXY",
                "TC_NUMERICAL_RESPONSE_NOT_CALIBRATED",
                "CELL_LOCAL_SYNTHETIC_PARAMETER_DECISION",
                "NONPROBABILISTIC_EPISTEMIC_ENVELOPE",
                "T4_COMMON_UNIT_CURVES_NOT_CALIBRATED",
                "REPRESENTATIVE_ARRAY_ZONE_APPLIED_TO_FULL_ARRAY_VALUE",
                "NAMED_REFERENCE_VALUE_PROFILE_USED",
                "WIND_ONLY_RAIN_DEBRIS_SURGE_AND_TORNADO_EXCLUDED",
            ],
            "withheld_failure_units": [],
        }
    )
    capability["consumer_annual_metrics"] = {
        "computation_owner": "downstream_consumer",
        "status_before_promotion": "withheld_noncanonical_proposal",
        "status_after_promotion": "consumer_computable_from_validated_frequency_intensity_coupling_value_and_cap_model",
        "prerequisites": [
            "exact model/docs/schema/SHA pin",
            "Hazard frequency-intensity-event coupling",
            "consumer cap-binding and compound-event validation",
            "explicit decision whether screening-grade v2.1 is acceptable",
        ],
        "limitation_flags": [
            "FREQUENCY_DRIVEN_ANNUAL_AND_TAIL_METRICS_ARE_CONSUMER_OWNED",
            "SYNTHETIC_SCENARIOS_ARE_NOT_A_PROBABILITY_DISTRIBUTION",
            "EVENT_PHYSICAL_DAMAGE_OUTPUTS_AVAILABLE_BEFORE_PROMOTION",
        ],
    }
    capability["cap_binding"]["checks_required"] = [
        "exact model/docs/schema/SHA pin",
        "pathway and architecture match",
        "array and site-facility demand bridges are present",
        "named value profile and representative-site exposure are explicit",
        "support is allocated once",
        "no synthetic resistance scenario is treated as calibrated frequency or uncertainty",
    ]
    capability["promotion_gate"] = {
        "status": "ready_for_review",
        "required_before_canonical_use": [
            "maintainer acceptance of screening-grade Tier-4 common-unit proxies",
            "Hazard exact-pin integration and representative event checks",
            "independent engineering review of common-unit medians and state costs",
            "dual-read and rollback test before current/index cutover",
        ],
    }
    return capability


def _replace_versioned_paths(artifact: dict[str, Any]) -> None:
    replacements = {
        "tropical_cyclone_wind_solar_curve_derivation_dossier__model_v2_0__docs_r1.md":
            "tropical_cyclone_wind_solar_curve_derivation_dossier__model_v2_1__docs_r1.md",
        "damage_curve_records_tropical_cyclone_wind_solar__model_v2_0__docs_r1.xlsx":
            "damage_curve_records_tropical_cyclone_wind_solar__model_v2_1__docs_r1.xlsx",
        "known_answer_tests_tropical_cyclone_wind_solar__model_v2_0__docs_r1.json":
            "known_answer_tests_tropical_cyclone_wind_solar__model_v2_1__docs_r1.json",
        "SOURCE_REGISTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv":
            "SOURCE_REGISTER_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv",
        "CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv":
            "CLAIM_PARAMETER_REGISTER_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv",
        "VALUE_CROSSWALK_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv":
            "VALUE_CROSSWALK_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv",
        "PARAMETER_TIER_TABLE_tropical_cyclone_wind_solar__model_v2_0__docs_r1.csv":
            "PARAMETER_TIER_TABLE_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv",
    }
    for field in (
        "source_dossier",
        "source_workbook",
        "known_answer_tests",
        "source_register",
        "claim_parameter_register",
        "value_crosswalk",
    ):
        value = artifact[field]
        for old, new in replacements.items():
            value = value.replace(old, new)
        artifact[field] = value


def build_artifact() -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    if digest(V20_ARTIFACT) != V20_ARTIFACT_SHA256:
        raise RuntimeError("model-v2.0 baseline changed; review v2.1 transformation")
    v20_artifact = json.loads(V20_ARTIFACT.read_text())
    artifact = deepcopy(v20_artifact)
    _replace_versioned_paths(artifact)
    artifact.update(
        {
            "damage_code_id": "TROPICAL_CYCLONE_WIND_SOLAR_SCREENING_COMPLETE_V2_1_PROPOSED",
            "semantic_damage_model_version": MODEL,
            "documentation_revision": DOCS,
            "lifecycle_state": "candidate",
            "promotion_status": "proposed_screening_ready_not_canonical",
            "review_status": "proposal_validation_complete_promotion_review_pending",
            "model_grade": "screening_engineering_proxy_T4_coverage_complete",
        }
    )
    capability = build_capability(v20_artifact["capability_declaration"])
    artifact["capability_declaration"] = capability
    records = common_records()
    pathway = artifact["pathways"][0]
    pathway["curve_records"].extend(records)
    pathway["site_facility_hazard_axis"] = {
        "id": "site_facility_tropical_cyclone_wind_demand_ratio",
        "preferred_input_field": "tc_site_event_to_design_wind_pressure_ratio",
        "permitted_proxy_fields": [
            "tc_peak_gust_3s_10m_mps",
            "qualified_site_design_3s_gust_mps",
        ],
        "unit": "dimensionless pressure-demand/design ratio",
        "valid_range": [0.0, 2.0],
        "extrapolation_policy": "reject",
        "required_bridge_field": "site_facility_demand_bridge_id",
    }
    coverage = []
    for row in pathway["failure_unit_coverage"]:
        updated = deepcopy(row)
        if updated["failure_unit_id"] in evaluator.COMMON_CURVE_UNITS:
            updated["status"] = "conditional_screening_t4_curve"
            updated["axis"] = "site_facility_tropical_cyclone_wind_demand_ratio"
        elif updated["failure_unit_id"] == evaluator.SUPPORT_UNIT:
            updated["status"] = "supported_derived_full_plant_assembly_rule"
        coverage.append(updated)
    pathway["failure_unit_coverage"] = coverage
    for unit in artifact["failure_units"]:
        if unit["id"] in evaluator.COMMON_CURVE_UNITS:
            unit["treatment"] = "secondary_conditional"
            unit["y_axis"] = "failure_unit_damage_ratio"
            unit["denominator"] = (
                "same-unit replacement value in the named v2.1 reference profile or a later governed site profile"
            )
            unit.pop("withheld_reason_codes", None)
        elif unit["id"] == evaluator.SUPPORT_UNIT:
            unit["denominator"] = (
                "named replacement-support value allocated once in full-plant assembly"
            )
            unit["withheld_reason_codes"] = ["DERIVED_ASSEMBLY_RULE_ONLY"]
    artifact["architecture_capability_interpretation"][evaluator.FIXED]["grade"] = (
        "screening_engineering_proxy_T4_coverage_complete"
    )
    artifact["architecture_capability_interpretation"][evaluator.TRACKER]["grade"] = (
        "screening_engineering_proxy_T4_coverage_complete"
    )
    artifact["value_linkage"] = {
        "implicit_default_profile": None,
        "screening_reference_profile": {
            "value_profile_id": evaluator.REFERENCE_PROFILE_ID,
            "source": "docs/method/value_basis/solar_wind_value_breakdown.xlsx",
            "currency_basis": "2024 USD",
            "unit": "USD_per_kWdc",
            "unit_values": evaluator.REFERENCE_VALUES_PER_KWDC,
            "physical_replacement_value_usd_per_kwdc": (
                evaluator.PHYSICAL_REPLACEMENT_VALUE_PER_KWDC
            ),
            "installed_capex_usd_per_kwdc": evaluator.INSTALLED_CAPEX_PER_KWDC,
            "request_rule": "caller must name value_profile_id; it is never silently selected",
        },
        "scenario_loss_status": "supported_with_named_reference_profile",
        "assembly_rules": [
            "evaluate every architecture-active array and common direct/civil failure unit",
            "multiply each same-unit DR by its named value row",
            "set replacement-support DR to the value-weighted direct-and-civil DR and allocate once",
            "divide total physical loss by 877.7957023626668 USD/kWdc for physical replacement DR",
            "divide the same numerator by 1120 USD/kWdc for installed-capex physical loss fraction",
            "multiply loss per kWdc by positive capacity_kwdc only when scenario dollars are requested",
        ],
    }
    artifact["evaluation_contract"].update(
        {
            "scenario_loss_behavior": "supported_in_full_plant_screening_mode",
            "full_plant_output_mode": "full_plant_screening",
            "value_profile_id_required": evaluator.REFERENCE_PROFILE_ID,
            "site_facility_axis_required_for_common_units": True,
            "support_rule": "derived_once_not_intrinsic_DR",
        }
    )
    artifact["evaluation_contract"].pop(
        "architecture_prohibited_for_direct_withheld_unit_query", None
    )
    artifact["evaluation_contract"][
        "architecture_prohibited_for_direct_failure_unit_query"
    ] = True
    artifact["evaluation_contract"]["failure_codes"] = sorted(
        (
            set(artifact["evaluation_contract"]["failure_codes"])
            | {
            "ARRAY_EXPOSURE_BASIS_REQUIRED",
            "ASSEMBLY_COVERAGE_INCOMPLETE",
            "CAPACITY_VALUE_INVALID",
            "DERIVED_ASSEMBLY_RULE_ONLY",
            "FULL_PLANT_ARCHITECTURE_UNSUPPORTED",
            "OUTPUT_MODE_REQUIRED",
            "SITE_AXIS_OUTSIDE_VALID_RANGE",
            "SITE_AXIS_PAYLOAD_AMBIGUOUS",
            "SITE_AXIS_PAYLOAD_REQUIRED",
            "SITE_FACILITY_DEMAND_BRIDGE_REQUIRED",
            "VALUE_PROFILE_UNSUPPORTED",
            }
        )
        - {"SCENARIO_LOSS_WITHHELD_SYNTHETIC_T4_PROPOSAL"}
    )
    artifact["evaluation_contract"]["withheld_reason_codes"] = [
        "DERIVED_ASSEMBLY_RULE_ONLY"
    ]
    artifact["emit_contract"] = {
        "schema_version": "damage_emit.v2",
        "populated_emit_modes_for_this_cell": ["scalar_mean", "state_ensemble"],
        "companion_output_schema": "physical_damage_assembly.v1",
        "result_grain_by_mode": {
            "failure_unit": "pathway_id x failure_unit_id x governed input state",
            "full_plant_screening": "pathway_id x array_architecture x event x named value profile",
        },
        "supported_outputs": [
            "failure-unit physical DR",
            "full-plant physical replacement DR",
            "installed-capex physical loss fraction",
            "physical loss per kWdc",
            "scenario physical dollars when capacity_kwdc is supplied",
        ],
        "prohibited_outputs": [
            "frequency and EAL inside the damage-curve layer",
            "PML, VaR, and TVaR without a downstream annual loss distribution",
            "business interruption, downtime, and revenue loss",
            "rain, debris, surge, flood, and tornado loss inside the wind-only pathway",
        ],
    }
    artifact["derivation_rationale"]["model_v2_1_usability_revision"] = {
        "reason": (
            "model v2.0 withheld most physical value and plant assembly, so it could not serve "
            "the requested screening use case"
        ),
        "change": (
            "add explicit Tier-4 common-unit curves and a fully reconciled named physical-value assembly"
        ),
        "preserved_boundaries": (
            "wind-only physical damage; calibrated annual/tail metrics and BI remain outside this repo"
        ),
    }
    artifact["legacy_comparison"]["v2_0_to_v2_1"] = (
        "v2.0 component outputs remain numerically unchanged; v2.1 adds common-unit proxy curves and full-plant assembly"
    )
    artifact["screening_curve_table"] = (
        "docs/cells/tropical_cyclone_wind_solar/proposed/"
        "FULL_PLANT_SCREENING_CURVE_TABLE_tropical_cyclone_wind_solar__model_v2_1__docs_r1.csv"
    )
    return artifact, capability, records


def full_fixed_request(array_ratio: float, site_ratio: float, *, capacity: float | None = None) -> dict[str, Any]:
    request = {
        **v20_builder.fixed_request(array_ratio),
        "output_mode": "full_plant_screening",
        "tc_site_event_to_design_wind_pressure_ratio": site_ratio,
        "site_facility_demand_bridge_id": "TC-SITE-FACILITY-BRIDGE-V1",
        "array_exposure_basis": "representative_site_array_zone",
        "value_profile_id": evaluator.REFERENCE_PROFILE_ID,
    }
    if capacity is not None:
        request["capacity_kwdc"] = capacity
    return request


def full_tracker_request(array_ratio: float, site_ratio: float) -> dict[str, Any]:
    return {
        **v20_builder.tracker_request(array_ratio),
        "output_mode": "full_plant_screening",
        "tc_site_event_to_design_wind_pressure_ratio": site_ratio,
        "site_facility_demand_bridge_id": "TC-SITE-FACILITY-BRIDGE-V1",
        "array_exposure_basis": "representative_site_array_zone",
        "value_profile_id": evaluator.REFERENCE_PROFILE_ID,
    }


def direct_gsu_request(site_ratio: float) -> dict[str, Any]:
    return {
        "event_id": "TC-EVENT-GSU",
        "event_family_id": "TC-FAMILY-GSU",
        "pathway_id": evaluator.PATHWAY,
        "failure_unit_id": "PV_GSU_SUBSTATION",
        "tc_site_event_to_design_wind_pressure_ratio": site_ratio,
        "site_facility_demand_bridge_id": "TC-SITE-FACILITY-BRIDGE-V1",
        "tc_duration_class": "sustained_1_to_6h",
        "tc_direction_evolution_class": "evolving",
        "rain_ingress_indicator": False,
        "windborne_debris_indicator": False,
        "flood_or_surge_indicator": False,
        "tc_tornado_indicator": False,
    }


def build_kats(artifact: Mapping[str, Any]) -> dict[str, Any]:
    runtime = []
    requests = [
        ("TCWS21-FIXED-ZERO", full_fixed_request(0.0, 0.0)),
        ("TCWS21-FIXED-CENTRAL", full_fixed_request(1.0, 1.0, capacity=100000.0)),
        ("TCWS21-FIXED-HIGH", full_fixed_request(2.0, 2.0)),
        ("TCWS21-TRACKER-CENTRAL", full_tracker_request(1.0, 1.0)),
        ("TCWS21-GSU-DIRECT", direct_gsu_request(1.0)),
    ]
    for test_id, request in requests:
        runtime.append(
            {
                "test_id": test_id,
                "request": request,
                "expected": evaluator.evaluate_damage_call(artifact, request),
            }
        )
    rejection_specs = []
    missing_site = full_fixed_request(1.0, 1.0)
    missing_site.pop("tc_site_event_to_design_wind_pressure_ratio")
    rejection_specs.append(("TCWS21-REJECT-MISSING-SITE-AXIS", missing_site, "SITE_AXIS_PAYLOAD_REQUIRED"))
    wrong_profile = full_fixed_request(1.0, 1.0)
    wrong_profile["value_profile_id"] = "UNKNOWN"
    rejection_specs.append(("TCWS21-REJECT-VALUE-PROFILE", wrong_profile, "VALUE_PROFILE_UNSUPPORTED"))
    no_mode = v20_builder.fixed_request(1.0)
    rejection_specs.append(("TCWS21-REJECT-NO-MODE-OR-UNIT", no_mode, "OUTPUT_MODE_REQUIRED"))
    support = direct_gsu_request(1.0)
    support["failure_unit_id"] = evaluator.SUPPORT_UNIT
    rejection_specs.append(("TCWS21-REJECT-DIRECT-SUPPORT", support, "DERIVED_ASSEMBLY_RULE_ONLY"))
    out_of_range = direct_gsu_request(2.1)
    rejection_specs.append(("TCWS21-REJECT-SITE-RANGE", out_of_range, "SITE_AXIS_OUTSIDE_VALID_RANGE"))
    rejections = [
        {"test_id": test_id, "request": request, "expected_error_code": code}
        for test_id, request, code in rejection_specs
    ]
    return {
        "cell_id": artifact["cell_id"],
        "semantic_damage_model_version": MODEL,
        "documentation_revision": DOCS,
        "runtime_known_answer_tests": runtime,
        "rejection_tests": rejections,
    }


def v21_claims() -> list[dict[str, str]]:
    claims = v20_builder.consolidated_claims()
    superseded_claims = {
        "TCWS-C023",
        "TCWS-C024",
        "TCWS-C030",
        "TCWS-C112",
        "TCWS-C113",
        "TCWS2-C014",
    }
    for claim in claims:
        if claim["claim_id"] in superseded_claims:
            claim["adoption_status"] = "historical_scope_only"
            claim["reasoning"] += (
                " Superseded for model v2.1 runtime behavior by TCWS21-C001/003/004/005; "
                "retained only to preserve the prior-version decision trail."
            )
    claims.extend(
        [
            {
                "claim_id": "TCWS21-C001",
                "pathway_id": evaluator.PATHWAY,
                "claim_text": "A screening v2 must return a complete physical-damage view rather than convert unsupported value to zero or leave the requested plant output unavailable.",
                "claim_type": "governed_model_decision",
                "source_ids": v20_builder.CELL_LOCAL_SYNTHETIC_SOURCE,
                "exact_locator": "DECISION_LOG_tropical_cyclone_wind_solar__model_v2_1__docs_r1.md#decision-1",
                "evidence_tier": "T4",
                "parameter_or_rule": "coverage_complete_screening_contract",
                "adoption_status": "adopted_model_v2_1",
                "permitted_inference": "v2.1 may use explicit labeled proxy curves to complete screening coverage",
                "prohibited_inference": "proxy status does not imply calibration or canonical promotion",
                "reasoning": "The requested product is a usable screening curve; v2.0 governance prevented that use.",
                "update_trigger": "replace proxies when matched unit-level TC evidence or elicitation is available",
            },
            {
                "claim_id": "TCWS21-C002",
                "pathway_id": evaluator.PATHWAY,
                "claim_text": "The named NLR physical-value profile reconciles 100% of the 877.7957023626668 USD/kWdc physical replacement basis when direct/civil units and support-once value are assembled.",
                "claim_type": "reproducible_value_reconciliation",
                "source_ids": "TCWS2-S013",
                "exact_locator": "docs/method/value_basis/solar_wind_value_breakdown.xlsx::Solar_Map rows 2-17",
                "evidence_tier": "T2",
                "parameter_or_rule": "reference_value_profile",
                "adoption_status": "adopted_model_v2_1",
                "permitted_inference": "named reference-profile physical DR and loss per kWdc",
                "prohibited_inference": "site-specific TIV accuracy without replacement by site values",
                "reasoning": "A named profile makes the screening assembly explicit, reproducible, and usable.",
                "update_trigger": "governed site-specific value profile or updated benchmark",
            },
            {
                "claim_id": "TCWS21-C003",
                "pathway_id": evaluator.PATHWAY,
                "claim_text": "Foundation, power/collection, GSU, SCADA, and civil infrastructure each return an explicit Tier-4 same-unit screening DR on the qualified site-facility wind-demand axis.",
                "claim_type": "governed_model_decision",
                "source_ids": v20_builder.CELL_LOCAL_SYNTHETIC_SOURCE,
                "exact_locator": "DECISION_LOG_tropical_cyclone_wind_solar__model_v2_1__docs_r1.md#decision-2",
                "evidence_tier": "T4",
                "parameter_or_rule": "common_failure_unit_numeric_coverage",
                "adoption_status": "adopted_model_v2_1",
                "permitted_inference": "numeric wind-only screening DR for each named same-unit failure unit",
                "prohibited_inference": "calibrated claims response, flood response, or transfer to another asset anatomy",
                "reasoning": "Explicit proxies make the requested screening product usable while preserving unit identity and evidence grade.",
                "update_trigger": "matched unit-level TC evidence or formal elicitation replaces a proxy record",
            },
            {
                "claim_id": "TCWS21-C004",
                "pathway_id": evaluator.PATHWAY,
                "claim_text": "Replacement support is allocated exactly once using the value-weighted direct-and-civil DR; it is not an intrinsic independently damaged unit.",
                "claim_type": "governed_assembly_rule",
                "source_ids": "GOVERNANCE_CONTRACT;TCWS2-S013",
                "exact_locator": "DECISION_LOG_tropical_cyclone_wind_solar__model_v2_1__docs_r1.md#decision-3",
                "evidence_tier": "T4",
                "parameter_or_rule": "replacement_support_allocate_once",
                "adoption_status": "adopted_model_v2_1",
                "permitted_inference": "complete named-profile physical loss without double-counting replacement support",
                "prohibited_inference": "an independent support fragility curve or repeated support loading across failure units",
                "reasoning": "The assembly needs full value coverage and must avoid multiplying the same support value by several unit DRs.",
                "update_trigger": "a governed support-cost allocation model supersedes the screening rule",
            },
            {
                "claim_id": "TCWS21-C005",
                "pathway_id": evaluator.PATHWAY,
                "claim_text": "Model v2.1 supports failure-unit DR, full-plant physical replacement DR, physical loss per kWdc, installed-capex physical loss fraction, and scenario physical dollars when capacity is supplied.",
                "claim_type": "runtime_output_contract",
                "source_ids": "GOVERNANCE_CONTRACT;TCWS2_CELL_LOCAL_SYNTHETIC_DECISION",
                "exact_locator": "tropical_cyclone_wind_solar__model_v2_1__docs_r1__curve_artifact.json::emit_contract",
                "evidence_tier": "T4",
                "parameter_or_rule": "v2_1_supported_outputs",
                "adoption_status": "adopted_model_v2_1",
                "permitted_inference": "the listed event physical-damage outputs for a valid exact-pinned screening call",
                "prohibited_inference": "EAL, PML, VaR, TVaR, BI, downtime, or excluded compound-pathway loss from the Damage artifact alone",
                "reasoning": "Usability requires numeric event damage and value outputs; annual and disruption metrics remain separate typed stages.",
                "update_trigger": "a governed output-contract revision changes the supported result set",
            },
        ]
    )
    return claims


def v21_parameter_rows(
    records: list[Mapping[str, Any]],
) -> list[dict[str, str]]:
    rows = v20_builder.parameter_rows(records)
    for row in rows:
        if row["parameter"] == "TCWS2-P000 | package curve-record count":
            row.update(
                {
                    "parameter": "TCWS21-P000 | package curve-record count",
                    "value": "10",
                    "reasoning": "One preserved Perry record, four architecture-specific array records, and five site-facility screening records.",
                    "status": "adopted_model_v2_1",
                }
            )
        elif row["parameter"] == "TCWS2-P000B | generic synthetic record count":
            row.update(
                {
                    "parameter": "TCWS21-P000B | array synthetic record count",
                    "reasoning": "Two fixed-tilt and two tracker array records remain unchanged from model v2.0.",
                    "status": "adopted_model_v2_1",
                }
            )
        elif row["parameter"] == "TCWS2-P004 | scenario loss":
            row.update(
                {
                    "parameter": "TCWS21-P004 | scenario physical loss",
                    "value": "supported_with_named_reference_profile",
                    "source_ids": "GOVERNANCE_CONTRACT;TCWS2-S013;TCWS2_CELL_LOCAL_SYNTHETIC_DECISION",
                    "reasoning": "Complete failure-unit screening coverage and the reconciled named value profile authorize explicit event physical loss; calibration grade remains Tier 4.",
                    "status": "adopted_model_v2_1",
                    "update_trigger": "A governed value profile, calibrated unit curve, or output-contract revision changes this behavior.",
                }
            )
    return rows


def v21_value_rows() -> list[dict[str, str]]:
    rows = v20_builder.value_rows()
    for row in rows:
        failure_units = row["failure_unit_id"].split("|")
        if any(unit in evaluator.COMMON_CURVE_UNITS for unit in failure_units) or any(
            token in row["row_or_bucket_id"]
            for token in ("MODULE", "MOUNTING")
        ):
            row["role_in_loss"] = "included_screening_direct_or_civil"
            row["include_in_direct_denominator"] = "true"
            row["allocation_rule"] = "Multiply the named row value by its v2.1 same-unit screening DR once."
            row["status"] = "included_v2_1_screening_reference"
        if row["failure_unit_id"] == evaluator.SUPPORT_UNIT:
            row["role_in_loss"] = "included_support_once"
            row["allocation_rule"] = "Apply the value-weighted direct-and-civil DR once; do not damage support independently."
            row["status"] = "included_v2_1_derived_support_rule"
        if row["row_or_bucket_id"] == "SOLAR_VALUE_SUMMARY_PHYSICAL":
            row["status"] = "v2_1_full_physical_reference_denominator"
        if row["row_or_bucket_id"] == "SOLAR_VALUE_SUMMARY_INSTALLED":
            row["status"] = "v2_1_installed_capex_reporting_denominator"
    return rows


def old_new_rows(artifact: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for architecture, request_factory in (
        ("fixed_tilt", full_fixed_request),
        ("tracker", full_tracker_request),
    ):
        for ratio in (0.0, 0.5, 1.0, 1.5, 2.0):
            result = evaluator.evaluate_damage_call(artifact, request_factory(ratio, ratio))
            scenarios = result["physical_damage_assembly"]["scenario_results"]
            rows.append(
                {
                    "architecture": architecture,
                    "array_demand_ratio": ratio,
                    "site_demand_ratio": ratio,
                    "model_v2_0_full_plant_dr": "WITHHELD",
                    "model_v2_1_lower_resistance_physical_dr": scenarios["lower_resistance"]["physical_replacement_dr"],
                    "model_v2_1_central_physical_dr": scenarios["central_screening"]["physical_replacement_dr"],
                    "model_v2_1_upper_resistance_physical_dr": scenarios["upper_resistance"]["physical_replacement_dr"],
                    "reason": "coverage-complete common-unit proxies plus named value assembly",
                }
            )
    return rows


def curve_table_rows(artifact: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for architecture, request_factory in (
        ("fixed_tilt", full_fixed_request),
        ("tracker", full_tracker_request),
    ):
        for step in range(41):
            ratio = step / 20
            result = evaluator.evaluate_damage_call(artifact, request_factory(ratio, ratio))
            for scenario_id, scenario in result["physical_damage_assembly"]["scenario_results"].items():
                rows.append(
                    {
                        "architecture": architecture,
                        "array_demand_ratio": ratio,
                        "site_facility_demand_ratio": ratio,
                        "scenario_id": scenario_id,
                        "physical_replacement_dr": scenario["physical_replacement_dr"],
                        "physical_loss_2024_usd_per_kwdc": scenario["physical_loss_2024_usd_per_kwdc"],
                        "installed_capex_physical_loss_fraction": scenario["installed_capex_physical_loss_fraction"],
                    }
                )
    return rows


def build_workbook(
    artifact: Mapping[str, Any],
    kats: Mapping[str, Any],
    sources: list[Mapping[str, Any]],
    claims: list[Mapping[str, Any]],
    parameters: list[Mapping[str, Any]],
    values: list[Mapping[str, Any]],
    old_new: list[Mapping[str, Any]],
    curve_table: list[Mapping[str, Any]],
) -> None:
    def workbook_rows(rows: list[Mapping[str, Any]]) -> list[dict[str, Any]]:
        return [
            {
                key: (
                    json.dumps(value, sort_keys=True)
                    if isinstance(value, (dict, list))
                    else value
                )
                for key, value in row.items()
            }
            for row in rows
        ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "README"
    sheet.append(["TC wind x solar", MODEL, DOCS])
    sheet.append(["Purpose", "coverage-complete screening physical DR and scenario loss"])
    sheet.append(["Canonical", False])
    sheet.append(["Curve records", len(artifact["pathways"][0]["curve_records"])])
    sheet.append(["Physical value coverage", 1.0])
    for title, rows in (
        ("Curve_Records", artifact["pathways"][0]["curve_records"]),
        ("Value_Crosswalk", values),
        ("Parameter_Tiers", parameters),
        ("Sources", sources),
        ("Claims", claims),
        ("Old_vs_New", old_new),
        ("Plant_Curve_Table", curve_table),
    ):
        worksheet = workbook.create_sheet(title)
        v20_builder.append_mapping_table(worksheet, workbook_rows(rows))
    worksheet = workbook.create_sheet("KATs")
    worksheet.append(["test_id", "type", "status_or_error"])
    for test in kats["runtime_known_answer_tests"]:
        worksheet.append([test["test_id"], "runtime", "PASS_EXPECTED"])
    for test in kats["rejection_tests"]:
        worksheet.append([test["test_id"], "rejection", test["expected_error_code"]])
    worksheet = workbook.create_sheet("QA")
    worksheet.append(["check", "observed", "expected", "status"])
    qa = [
        ("model", artifact["semantic_damage_model_version"], MODEL),
        ("curve records", len(artifact["pathways"][0]["curve_records"]), 10),
        ("common curves", len(common_records()), 5),
        ("withheld units", len(artifact["capability_declaration"]["pathway_capabilities"][0]["withheld_failure_units"]), 0),
        ("scenario loss support", artifact["capability_declaration"]["pathway_capabilities"][0]["scenario_loss_given_value_basis"], "supported_with_explicit_failure_unit_value_and_exposure_basis"),
        ("physical value", evaluator.PHYSICAL_REPLACEMENT_VALUE_PER_KWDC, 877.7957023626668),
        ("runtime KATs", len(kats["runtime_known_answer_tests"]), 5),
        ("rejection KATs", len(kats["rejection_tests"]), 5),
    ]
    for name, observed, expected in qa:
        worksheet.append([name, observed, expected, "PASS" if observed == expected else "FAIL"])
    for worksheet in workbook.worksheets:
        v20_builder.style_sheet(worksheet)
    v20_builder.save_deterministic_workbook(workbook, WORKBOOK)


def main() -> int:
    artifact, capability, new_records = build_artifact()
    sources = v20_builder.consolidated_sources()
    claims = v21_claims()
    parameters = v21_parameter_rows(
        [*artifact["pathways"][0]["curve_records"][1:5], *new_records]
    )
    values = v21_value_rows()
    kats = build_kats(artifact)
    old_new = old_new_rows(artifact)
    curve_table = curve_table_rows(artifact)
    v20_builder.write_json(CAPABILITY, capability)
    v20_builder.write_json(ARTIFACT, artifact)
    v20_builder.write_json(KATS, kats)
    v20_builder.write_csv(SOURCES, v20_builder.SOURCE_FIELDS, sources)
    v20_builder.write_csv(CLAIMS, v20_builder.CLAIM_FIELDS, claims)
    v20_builder.write_csv(PARAMETERS, v20_builder.PARAMETER_FIELDS, parameters)
    v20_builder.write_csv(VALUES, list(values[0]), values)
    v20_builder.write_csv(OLD_NEW, list(old_new[0]), old_new)
    v20_builder.write_csv(CURVE_TABLE, list(curve_table[0]), curve_table)
    build_workbook(artifact, kats, sources, claims, parameters, values, old_new, curve_table)
    print(f"artifact_sha256={digest(ARTIFACT)}")
    print(f"capability_sha256={digest(CAPABILITY)}")
    print(f"known_answer_tests_sha256={digest(KATS)}")
    print(f"workbook_sha256={digest(WORKBOOK)}")
    print(f"curve_records={len(artifact['pathways'][0]['curve_records'])}")
    print(f"runtime_kats={len(kats['runtime_known_answer_tests'])}")
    print(f"curve_table_rows={len(curve_table)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
