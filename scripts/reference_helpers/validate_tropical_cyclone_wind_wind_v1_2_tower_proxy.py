#!/usr/bin/env python3
"""Validate the TC-wind × Wind Farm model-v1.2 tower-only proposal."""

from __future__ import annotations

import csv
import json
import math
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from tropical_cyclone_wind_wind_curve_eval import (
    PROXY_ARCHETYPE_ID,
    PROXY_FAILURE_UNIT,
    SUPPORTED_FAILURE_UNIT,
    TOWER_PROXY_ARCHETYPE_ID,
    TOWER_PROXY_ASSET_PROFILE_ID,
    TOWER_PROXY_COVERED_VALUE_SHARE,
    TOWER_PROXY_POLICY_ID,
    TOWER_PROXY_VALUE_BASIS_ID,
    TropicalCycloneWindEvaluationError,
    evaluate_damage_call,
)


ROOT = Path(__file__).resolve().parents[2]
CELL = ROOT / "docs/cells/tropical_cyclone_wind_wind"
PROPOSED = CELL / "proposed"
BASELINE = CELL / "archive/model_v1_0__docs_r1"
STEM = "model_v1_2__docs_r2"
ARTIFACT = PROPOSED / f"tropical_cyclone_wind_wind__{STEM}__curve_artifact.json"
CAPABILITY = PROPOSED / f"tropical_cyclone_wind_wind__{STEM}__capability.json"
KATS = PROPOSED / f"known_answer_tests_tropical_cyclone_wind_wind__{STEM}.json"
VALUES = PROPOSED / f"VALUE_CROSSWALK_tropical_cyclone_wind_wind__{STEM}.csv"
WORKBOOK = PROPOSED / f"damage_curve_records_tropical_cyclone_wind_wind__{STEM}.xlsx"
REPORT = PROPOSED / f"VALIDATION_REPORT_tropical_cyclone_wind_wind__{STEM}.md"
BASELINE_ARTIFACT = BASELINE / "tropical_cyclone_wind_wind__model_v1_0__docs_r1__curve_artifact.json"
BUNDLE_SCHEMA = ROOT / "docs/contracts/schemas/curve_artifact_bundle.v3.schema.json"
CAPABILITY_SCHEMA = ROOT / "docs/contracts/schemas/capability_declaration.v3.schema.json"
EMIT_SCHEMA = ROOT / "docs/contracts/schemas/damage_emit.v2.schema.json"


def load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def result(emit: Mapping[str, Any], unit_id: str) -> Mapping[str, Any]:
    matches = [item for item in emit["failure_unit_results"] if item["failure_unit_id"] == unit_id]
    require(len(matches) == 1, f"expected one result for {unit_id}")
    return matches[0]


def validate_structure(baseline: Mapping[str, Any], artifact: Mapping[str, Any], capability: Mapping[str, Any]) -> None:
    require(artifact["semantic_damage_model_version"] == "model v1.2", "model version mismatch")
    require(artifact["documentation_revision"] == "docs r2", "docs revision mismatch")
    require(artifact["canonical_runtime_artifact"] is False, "proposal is unexpectedly canonical")
    require(artifact["promotion_status"] == "proposed", "proposal status mismatch")
    require(capability == artifact["capability_declaration"], "embedded/external capability drift")
    baseline_records = baseline["pathways"][0]["curve_records"]
    proposed_records = artifact["pathways"][0]["curve_records"]
    require(proposed_records[: len(baseline_records)] == baseline_records, "source-native records changed")
    require(len(proposed_records) == len(baseline_records) + 1, "expected three source records plus one proxy")
    proxy = proposed_records[-1]
    source = baseline_records[-1]
    require(proxy["failure_unit_id"] == SUPPORTED_FAILURE_UNIT, "proxy is not on the source tower unit")
    require(proxy["parameters"] == source["parameters"], "proxy parameters changed")
    require(proxy["x_axis"] == source["x_axis"], "proxy axis changed")
    require(proxy["selector_match"]["turbine_archetype_id"] == TOWER_PROXY_ARCHETYPE_ID, "proxy archetype changed")
    require(all(record["selector_match"]["turbine_archetype_id"] != PROXY_ARCHETYPE_ID for record in proposed_records), "old 0.63 proxy remained")
    contract = artifact["derivation_rationale"]["owner_approved_proxy_contract"]
    require(contract["proxy_policy_id"] == TOWER_PROXY_POLICY_ID, "proxy policy mismatch")
    require(contract["canonical_asset_profile_id"] == TOWER_PROXY_ASSET_PROFILE_ID, "asset profile mismatch")
    require(contract["covered_value_basis_id"] == TOWER_PROXY_VALUE_BASIS_ID, "value basis mismatch")
    require(contract["failure_unit_id"] == SUPPORTED_FAILURE_UNIT, "contract unit mismatch")
    require(contract["covered_value_share_of_project_tiv"] == 0.16, "covered share mismatch")
    require(contract["uncovered_value_share_of_project_tiv"] == 0.84, "uncovered share mismatch")


def validate_reproduction(baseline: Mapping[str, Any], artifact: Mapping[str, Any], kats: Mapping[str, Any]) -> int:
    for test in kats["v1_0_reproduction_tests"]:
        old = result(evaluate_damage_call(baseline, test["input"]), test["input"]["failure_unit_id"])
        new = result(evaluate_damage_call(artifact, test["input"]), test["input"]["failure_unit_id"])
        require(new == old, f"source reproduction failed: {test['test_id']}")
    return len(kats["v1_0_reproduction_tests"])


def validate_proxy(artifact: Mapping[str, Any], kats: Mapping[str, Any]) -> int:
    for test in kats["proxy_known_answer_tests"]:
        emit = evaluate_damage_call(artifact, test["input"])
        actual = result(emit, SUPPORTED_FAILURE_UNIT)
        expected = test["expected"]
        require(actual["status"] == "supported", f"proxy withheld: {test['test_id']}")
        require(actual["curve_id"] == expected["curve_id"], f"curve drift: {test['test_id']}")
        require(math.isclose(actual["scalar_central_dr"], expected["failure_unit_damage_ratio"], rel_tol=0, abs_tol=kats["absolute_tolerance"]), f"DR drift: {test['test_id']}")
        require(emit["exposure_used"]["covered_value_share_of_project_tiv"] == 0.16, "covered share drift")
        require(emit["exposure_used"]["covered_subsystems"] == ["tower"], "covered subsystem drift")
        flags = set(actual["metadata_flags"])
        require("PARTIAL_TOWER_VALUE_COVERAGE_16PCT" in flags, "tower coverage flag missing")
        require("UNCOVERED_PROJECT_VALUE_84PCT_WITHHELD_NOT_ZERO" in flags, "withheld flag missing")
        boundary_flag = expected.get("boundary_flag")
        if boundary_flag:
            require(boundary_flag in flags, f"boundary flag missing: {test['test_id']}")
    return len(kats["proxy_known_answer_tests"])


def validate_negative(artifact: Mapping[str, Any], kats: Mapping[str, Any]) -> int:
    for test in kats["negative_contract_tests"]:
        try:
            evaluate_damage_call(artifact, test["input"])
        except TropicalCycloneWindEvaluationError as exc:
            require(exc.code == test["expected_error_code"], f"wrong error for {test['test_id']}: {exc.code}")
        else:
            raise AssertionError(f"negative test passed unexpectedly: {test['test_id']}")
    sample = dict(kats["proxy_known_answer_tests"][5]["input"])
    sample["failure_unit_id"] = PROXY_FAILURE_UNIT
    equipment = result(evaluate_damage_call(artifact, sample), PROXY_FAILURE_UNIT)
    require(equipment["status"] == "withheld", "equipment assembly received a numeric proxy")
    require(equipment["scalar_central_dr"] is None, "equipment assembly DR must be null")
    return len(kats["negative_contract_tests"]) + 1


def validate_value(kats: Mapping[str, Any]) -> int:
    with VALUES.open(newline="") as handle:
        rows = list(csv.DictReader(handle))
    shares = {row["subsystem"]: float(row["share_of_project_tiv"]) for row in rows}
    require(shares == {"tower": 0.16, "covered_total": 0.16, "all_non_tower_value": 0.84}, "value table changed")
    for test in kats["value_crosswalk_tests"]:
        tiv = test["project_tiv_usd"]
        require(math.isclose(tiv * 0.16, test["expected_covered_value_usd"]), "covered value KAT failed")
        require(math.isclose(tiv * 0.84, test["expected_uncovered_value_usd"]), "uncovered value KAT failed")
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    require(workbook.sheetnames == ["README", "Curve records", "Proxy KATs", "Value crosswalk"], "workbook topology changed")
    return len(kats["value_crosswalk_tests"])


def validate_schemas(artifact: Mapping[str, Any], capability: Mapping[str, Any], kats: Mapping[str, Any]) -> str:
    from jsonschema import Draft202012Validator
    from referencing import Registry, Resource

    schemas = [load(BUNDLE_SCHEMA), load(CAPABILITY_SCHEMA), load(EMIT_SCHEMA)]
    registry = Registry().with_resources([(schema["$id"], Resource.from_contents(schema)) for schema in schemas])
    Draft202012Validator(schemas[1], registry=registry).validate(capability)
    Draft202012Validator(schemas[0], registry=registry).validate(artifact)
    emit = evaluate_damage_call(artifact, kats["proxy_known_answer_tests"][0]["input"])
    Draft202012Validator(schemas[2], registry=registry).validate(emit)
    return "bundle v3, capability v3 and emit v2 passed"


def main() -> None:
    baseline = load(BASELINE_ARTIFACT)
    artifact = load(ARTIFACT)
    capability = load(CAPABILITY)
    kats = load(KATS)
    validate_structure(baseline, artifact, capability)
    reproduction = validate_reproduction(baseline, artifact, kats)
    proxy = validate_proxy(artifact, kats)
    negative = validate_negative(artifact, kats)
    value = validate_value(kats)
    schemas = validate_schemas(artifact, capability, kats)
    REPORT.write_text(f"""# Validation report — model v1.2/docs r2

Status: **PASS as a noncanonical proposal**.

| Gate | Result |
|---|---:|
| source-native reproduction answers | {reproduction} passed |
| tower-proxy known answers | {proxy} passed |
| negative/fail-closed contracts | {negative} passed |
| value/cap answers | {value} passed |
| schemas | {schemas} |

The Jaimes parameters are unchanged. The proposed canonical bridge covers only tower value (0.16 of
project TIV), with 0.84 explicitly withheld. Promotion still requires the governed Hazard consumer rebuild.
""")
    print("PASS tropical_cyclone_wind_wind model v1.2 tower-only proposal")
    print(f"source_reproduction={reproduction}")
    print(f"tower_proxy_known_answers={proxy}")
    print(f"negative_contracts={negative}")
    print(f"value_tests={value}")
    print(schemas)


if __name__ == "__main__":
    main()
